# === Standard Library ===
import os
from datetime import datetime

# === Third-Party Libraries ===
import pandas as pd
from tqdm import tqdm
from scipy.stats import trim_mean

# add near the other imports

from dropbox_utils import upload_file
from config import get_dropbox_path


# === Excel Handling ===
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# === Defaults ===
from config import DEFAULT_START_DATE

# === Contract map: ticker -> (security type, exchange, currency) ===
from config import CONTRACT_MAP

# === IB Data Fetch ===
def get_ib_data(tickers="SPY", start_date="2003-01-01"):
    from ib_insync import IB, Stock, Index
    from datetime import datetime
    import pandas as pd
    from tqdm import tqdm

    if isinstance(tickers, str):
        tickers = [tickers]

    results = {}

    for ticker in tickers:
        ib = IB()
        try:
            # unique clientId per ticker
            cid = abs(hash(ticker)) % 9000
            ib.connect("127.0.0.1", 7496, clientId=cid)
            print(f"[get_ib_data] Connected to IB (clientId={cid}) for {ticker}")
        except Exception as e:
            print(f"[get_ib_data] Connection failed for {ticker}: {e}")
            continue

        try:
            stype, exch, curr = CONTRACT_MAP[ticker]
            contract = Stock(ticker, exch, curr) if stype == "Stock" else Index(ticker, exch)

            print(f"[get_ib_data] Requesting bars for {ticker} since {start_date}...")
            bars = ib.reqHistoricalData(
                contract,
                endDateTime="",
                durationStr="20 Y",
                barSizeSetting="1 day",
                whatToShow="TRADES",
                useRTH=True,
                formatDate=1,
            )
            df = pd.DataFrame(bars)
            if not df.empty:
                df["date"] = pd.to_datetime(df["date"])
                results[ticker] = df
                print(f"[get_ib_data] Retrieved {len(df)} rows for {ticker}")
            else:
                print(f"[get_ib_data] No data for {ticker}")
        finally:
            try:
                ib.disconnect()
                print(f"[get_ib_data] Disconnected {ticker}")
            except Exception:
                pass

    return results


# === Known Corrections (SPY only) ===
SPY_CORRECTIONS = {
    "2004-04-19": {"open": 113.55, "high": 113.99, "low": 113.27, "close": 113.83, "volume": 28277600},
    "2004-04-27": {"open": 114.23, "high": 115.12, "low": 113.96, "close": 114.30, "volume": 43485500},
    "2004-05-05": {"open": 112.41, "high": 112.96, "low": 112.16, "close": 112.78, "volume": 34405000},
    "2006-03-03": {"open": 128.67, "high": 130.07, "low": 128.65, "close": 128.76, "volume": 73402496},
    "2006-05-18": {"open": 127.35, "high": 127.75, "low": 126.11, "close": 126.21, "volume": 37906304},
    "2006-03-23": {"open": 130.26, "high": 130.39, "low": 129.66, "close": 130.11, "volume": 46704200},
    "2007-04-02": {"open": 142.16, "high": 142.46, "low": 140.89, "close": 142.16, "volume": 79454680},
    "2007-05-25": {"open": 151.49, "high": 152.02, "low": 151.18, "close": 151.69, "volume": 33309621},
    "2007-06-29": {"open": 150.90, "high": 151.65, "low": 149.15, "close": 150.43, "volume": 39701730},
    "2007-07-23": {"open": 154.18, "high": 154.72, "low": 153.30, "close": 153.97, "volume": 21183836},
    "2007-07-26": {"open": 150.19, "high": 150.80, "low": 146.39, "close": 148.02, "volume": 57592490},
    "2007-08-10": {"open": 144.39, "high": 146.50, "low": 143.12, "close": 144.71, "volume": 11018370},
    "2022-01-24": {"open": 432.03, "high": 440.38, "low": 420.76, "close": 439.84, "volume": 52496719},
    "2022-01-26": {"open": 440.72, "high": 444.04, "low": 428.86, "close": 433.38, "volume": 36391088},
    "2022-02-24": {"open": 411.02, "high": 428.76, "low": 410.64, "close": 428.30, "volume": 13942946},
    "2022-05-04": {"open": 417.08, "high": 429.66, "low": 413.71, "close": 429.06, "volume": 44247895},
    "2022-05-05": {"open": 424.55, "high": 425.00, "low": 409.44, "close": 413.81, "volume": 72929106},
    "2022-05-20": {"open": 393.25, "high": 397.03, "low": 380.54, "close": 389.63, "volume": 31432197},
    "2022-10-13": {"open": 349.21, "high": 367.51, "low": 348.11, "close": 365.97, "volume": 47254518},
    "2024-08-01": {"open": 552.57, "high": 554.87, "low": 539.43, "close": 543.01, "volume": 76428732},
    "2024-09-11": {"open": 548.70, "high": 555.36, "low": 539.96, "close": 554.42, "volume": 75248608},
    "2024-12-18": {"open": 603.98, "high": 606.41, "low": 585.89, "close": 586.28, "volume": 38248729},
    "2025-03-03": {"open": 596.18, "high": 597.34, "low": 579.90, "close": 583.77, "volume": 74249199},
    "2025-04-04": {"open": 523.67, "high": 525.87, "low": 505.06, "close": 505.28, "volume": 17965131},
    "2025-04-07": {"open": 489.19, "high": 523.17, "low": 481.80, "close": 504.38, "volume": 56611355},
    "2025-04-08": {"open": 521.86, "high": 524.98, "low": 489.16, "close": 496.48, "volume": 55816581},
    "2025-04-09": {"open": 493.44, "high": 548.62, "low": 493.05, "close": 548.62, "volume": 11867317},
    "2025-04-10": {"open": 532.17, "high": 533.50, "low": 509.32, "close": 524.58, "volume": 52331225},
    "2025-04-11": {"open": 523.01, "high": 536.43, "low": 520.07, "close": 533.94, "volume": 37866334},
    "2025-04-30": {"open": 547.57, "high": 556.52, "low": 541.52, "close": 554.54, "volume": 33101463},
}

# === Gap Data Calculation ===
def compute_gap_data(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    gap_types = []
    days_to_target = []
    target_dates = []

    df["year"] = pd.to_datetime(df["date"]).dt.year
    df["dow"] = pd.to_datetime(df["date"]).dt.day_name()
    df["previous_date"] = df["date"].shift(1)
    df["previous_open"] = df["open"].shift(1)
    df["previous_high"] = df["high"].shift(1)
    df["previous_low"] = df["low"].shift(1)
    df["previous_close"] = df["close"].shift(1)
    df["gap value"] = abs(df["open"] - df["previous_close"])

    data = df.to_dict("records")
    for i in tqdm(range(1, len(data)), desc="Computing gap data"):
        prev_bar = data[i - 1]
        curr_bar = data[i]

        if prev_bar["close"] < curr_bar["open"] <= prev_bar["high"]:
            gap_type = "price_gap_up"
        elif prev_bar["close"] > curr_bar["open"] >= prev_bar["low"]:
            gap_type = "price_gap_down"
        elif curr_bar["open"] > prev_bar["high"]:
            gap_type = "bar_gap_up"
        elif curr_bar["open"] < prev_bar["low"]:
            gap_type = "bar_gap_down"
        else:
            gap_type = "no_gap"
        gap_types.append(gap_type)

        target_achieved = False
        days_count = 0
        target_achieved_date = None

        if gap_type == "no_gap":
            days_to_target.append(0)
            target_dates.append(prev_bar["date"])
        else:
            for j in range(i, len(data)):
                next_bar = data[j]
                days_count += 1
                if gap_type in ["price_gap_up", "bar_gap_up"] and next_bar["low"] <= prev_bar["close"]:
                    target_achieved = True
                    target_achieved_date = next_bar["date"]
                    break
                elif gap_type in ["price_gap_down", "bar_gap_down"] and next_bar["high"] >= prev_bar["close"]:
                    target_achieved = True
                    target_achieved_date = next_bar["date"]
                    break
            days_to_target.append(days_count if target_achieved else None)
            target_dates.append(target_achieved_date)

    df = df.iloc[1:].reset_index(drop=True).copy()
    df["gap_type"] = gap_types
    df["days to target"] = days_to_target
    df["target_achieved_date"] = target_dates
    df["trade_direction"] = df["gap_type"].apply(
        lambda x: "not tradable" if "no_gap" in str(x).lower()
        else "long" if "down" in str(x).lower()
        else "short" if "up" in str(x).lower()
        else None
    )
    df["original_index"] = range(len(df))
    return df

# === Rolling Statistics ===
def calculate_statistics(df, trim_percentage=0.1, window=50):
    df = df.copy()
    df[f"MA {window} Days to Close"] = df["days to target"].rolling(window=window, min_periods=1).mean()
    df[f"Rolling Stdev {window} Days"] = df["days to target"].rolling(window=window, min_periods=1).std()
    df["Median"] = df["days to target"].rolling(window=window, min_periods=1).median()
    df[f"Trim Avg Days to Close {window} Days"] = df["days to target"].rolling(
        window=window, min_periods=1).apply(lambda x: trim_mean(x, proportiontocut=trim_percentage), raw=False)
    df[f"Trim Stdev Days to Close {window} Days"] = df["days to target"].rolling(
        window=window, min_periods=1).apply(lambda x: x.sort_values()
                                            .iloc[int(len(x) * trim_percentage): int(len(x) * (1 - trim_percentage))].std(ddof=1), raw=False)
    df["1st Quartile (25%)"] = df["days to target"].rolling(window=window, min_periods=1).quantile(0.25)
    df["2nd Quartile (50%)"] = df["days to target"].rolling(window=window, min_periods=1).quantile(0.50)
    df["3rd Quartile (75%)"] = df["days to target"].rolling(window=window, min_periods=1).quantile(0.75)
    df["IQR"] = df["3rd Quartile (75%)"] - df["1st Quartile (25%)"]
    df["Upper Bound"] = df["3rd Quartile (75%)"] + (1.5 * df["IQR"])
    return df

# === Summaries + MAM Stats ===
def create_days_to_target_summary_table_gap_types_df(df, max_day=10):
    df = df[df["days to target"].notna()].copy()
    df["gap_type"] = df["gap_type"].astype(str)
    df["days to target"] = pd.to_numeric(df["days to target"], errors="coerce").fillna(0).astype(int)
    summary_rows = []
    for gap_type in sorted(df["gap_type"].unique()):
        if gap_type == "no_gap":
            continue
        sub_df = df[df["gap_type"] == gap_type]
        total = len(sub_df)
        if total == 0:
            continue
        cumulative_counts = [(sub_df["days to target"] <= i).sum() for i in range(1, max_day + 1)]
        cumulative_pct = [round(c / total, 4) for c in cumulative_counts]
        row = [gap_type] + cumulative_pct
        summary_rows.append(row)
    columns = ["gap_type"] + [f"Days {i}" for i in range(1, max_day + 1)]
    return pd.DataFrame(summary_rows, columns=columns)

def compute_gap_type_stats_df(df):
    df = df.copy()
    df["gap_type"] = df["gap_type"].astype(str)
    df = df[(df["days to target"].notna()) & (df["gap_type"] != "no_gap")].copy()
    return df.groupby("gap_type")["days to target"].agg([
        ("mean", "mean"),
        ("std", "std"),
        ("trim_mean", lambda x: trim_mean(x, proportiontocut=0.1)),
        ("trim_std", lambda x: x[(x >= x.quantile(0.1)) & (x <= x.quantile(0.9))].std()),
        ("q1", lambda x: x.quantile(0.25)),
        ("median", "median"),
        ("q3", lambda x: x.quantile(0.75)),
        ("upper_bound", lambda x: x.quantile(0.75) + 1.5 * (x.quantile(0.75) - x.quantile(0.25))),
        ("mad", lambda x: (x - x.median()).abs().mean()),
    ]).reset_index()

def generate_mam_stats_by_gap_type_df(df):
    df = df.copy()
    df["gap_type"] = df["gap_type"].astype(str)
    df = df[(df["max_adverse_move"].notna()) & (df["gap_type"] != "no_gap")].copy()
    grouped = df.groupby("gap_type")["max_adverse_move"].agg([
        ("mean_mam_pts", "mean"),
        ("std_mam_pts", "std"),
        ("median_mam_pts", "median"),
        ("q1_mam_pts", lambda x: x.quantile(0.25)),
        ("q3_mam_pts", lambda x: x.quantile(0.75)),
        ("mad_mam_pts", lambda x: (x - x.median()).abs().mean()),
    ]).reset_index()
    grouped["iqr_mam_pts"] = grouped["q3_mam_pts"] - grouped["q1_mam_pts"]
    grouped["upper_bound"] = grouped["q3_mam_pts"] + 1.5 * grouped["iqr_mam_pts"]
    return grouped

def generate_mam_days_stats_by_gap_type_df(df):
    df = df.copy()
    df["gap_type"] = df["gap_type"].astype(str)
    df = df[(df["days_to_mam"].notna()) & (df["gap_type"] != "no_gap")].copy()
    grouped = df.groupby("gap_type")["days_to_mam"].agg([
        ("mean_days_to_mam", "mean"),
        ("std_days_to_mam", "std"),
        ("median_days_to_mam", "median"),
        ("q1_days_to_mam", lambda x: x.quantile(0.25)),
        ("q3_days_to_mam", lambda x: x.quantile(0.75)),
        ("mad_days_to_mam", lambda x: (x - x.median()).abs().mean()),
    ]).reset_index()
    grouped["iqr_days_to_mam"] = grouped["q3_days_to_mam"] - grouped["q1_days_to_mam"]
    grouped["upper_bound_mam"] = grouped["q3_days_to_mam"] + 1.5 * grouped["iqr_days_to_mam"]
    return grouped

# === MAM Calculation ===
def compute_gap_data_with_mam_and_days_to_mam(df):
    from tqdm import tqdm
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"])
    max_adverse_moves, mam_dates, mam_indexes, days_to_mam = [None], [None], [None], [None]

    for i in tqdm(range(1, len(df)), desc="Computing MAM"):
        curr_bar = df.iloc[i]
        gap_type = str(curr_bar.get("gap_type", "")).strip().lower()
        entry_price = curr_bar["open"]
        target_achieved_date = curr_bar.get("target_achieved_date", pd.NaT)

        if gap_type == "no_gap":
            max_adverse_moves.append(0)
            mam_dates.append(curr_bar["date"])
            mam_indexes.append(i)
            days_to_mam.append(0)
            continue

        if pd.isnull(target_achieved_date):
            j_end = len(df) - 1
        else:
            match = df[df["date"] == target_achieved_date]
            j_end = match.index[0] if not match.empty else len(df) - 1

        if gap_type in ("price_gap_down", "bar_gap_down"):
            min_low, min_idx = float("inf"), None
            for j in range(i, j_end + 1):
                low = df.iloc[j]["low"]
                if low < min_low:
                    min_low, min_idx = low, j
            mam_val = entry_price - min_low
            mam_date = df.iloc[min_idx]["date"] if min_idx is not None else None
            mam_idx = min_idx
        elif gap_type in ("price_gap_up", "bar_gap_up"):
            max_high, max_idx = -float("inf"), None
            for j in range(i, j_end + 1):
                high = df.iloc[j]["high"]
                if high > max_high:
                    max_high, max_idx = high, j
            mam_val = max_high - entry_price
            mam_date = df.iloc[max_idx]["date"] if max_idx is not None else None
            mam_idx = max_idx
        else:
            mam_val, mam_date, mam_idx = None, None, None

        max_adverse_moves.append(mam_val)
        mam_dates.append(mam_date)
        mam_indexes.append(mam_idx)
        days_to_mam.append(mam_idx - i + 1 if mam_idx is not None else None)

    df["max_adverse_move"] = max_adverse_moves
    df["mam_date"] = mam_dates
    df["mam_index"] = mam_indexes
    df["days_to_mam"] = days_to_mam
    return df

# === Excel Helpers ===
def format_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

def format_probabilities_in_workbook(wb):
    if "Days to Target" in wb.sheetnames:
        ws = wb["Days to Target"]
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if isinstance(cell.value, (float, int)):
                    cell.number_format = "0.0%"

def save_multiple_sheets_with_formatting(sheets_dict, output_path):
    wb = Workbook()
    first_sheet = True
    for sheet_name, df in sheets_dict.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        headers = ["index"] + list(map(str, df.columns)) if sheet_name == "All Data" else list(map(str, df.columns))
        ws.append(headers)
        for idx, row in df.iterrows():
            values = [idx] + row.tolist() if sheet_name == "All Data" else row.tolist()
            ws.append(values)
        format_sheet(wb, sheet_name)
    format_probabilities_in_workbook(wb)
    wb.save(output_path)

# === Runner ===


def run_gap_analysis_for_contracts(contract_map: dict, dfs: dict,
                                   start_date=DEFAULT_START_DATE):
    results = {}

    for ticker in contract_map:
        print(f"=== {ticker} ===")
        df = dfs.get(ticker)
        if df is None or df.empty:
            print(f"No data for {ticker}, skipping")
            continue

        # --- SPY corrections ---
        if ticker == "SPY":
            df["date"] = pd.to_datetime(df["date"])
            for date_str, values in SPY_CORRECTIONS.items():
                mask = df["date"] == pd.to_datetime(date_str)
                for col, val in values.items():
                    df.loc[mask, col] = val

        # --- Core gap computations ---
        df = compute_gap_data(df)
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df["target_achieved_date"] = pd.to_datetime(df["target_achieved_date"], errors="coerce")
        df = compute_gap_data_with_mam_and_days_to_mam(df)
        df = calculate_statistics(df)

        for col in ["date", "previous_date", "target_achieved_date", "mam_date"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce").dt.date

        gap_type_stats = compute_gap_type_stats_df(df)
        days_to_target = create_days_to_target_summary_table_gap_types_df(df)
        mam_stats_pts = generate_mam_stats_by_gap_type_df(df)
        mam_stats_days = generate_mam_days_stats_by_gap_type_df(df)

        # --- Save locally (temporary file) ---
        local_filename = f"{ticker.lower()} gap analysis.xlsx"
        save_multiple_sheets_with_formatting(
            {
                "All Data": df,
                "Gap Type Stats (DAYS)": gap_type_stats,
                "Days to Target": days_to_target,
                "MAM (PTS)": mam_stats_pts,
                "MAM (DAYS)": mam_stats_days,
            },
            local_filename
        )

        # --- Upload to Dropbox ---
        dropbox_path = get_dropbox_path(ticker, "gaps", local_filename)
        upload_file(local_filename, dropbox_path)
        print(f"[Dropbox] Uploaded {ticker} gap analysis → {dropbox_path}")

        results[ticker] = dropbox_path
        print(f"[Dropbox] Uploaded {ticker} → {dropbox_path}")

    return results


# === Main Runner ===
if __name__ == "__main__":
    dfs = get_ib_data(list(CONTRACT_MAP.keys()), start_date=DEFAULT_START_DATE)
    results = run_gap_analysis_for_contracts(CONTRACT_MAP, dfs)
    print("Gap analysis complete:", results)

