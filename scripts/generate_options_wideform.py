import os
import pandas as pd
from tqdm import tqdm
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# === CONFIG ===
FOLDER = r"C:\2mdt\2mindt-site\public\spy-options"
FILE_PREFIX = "spy_options_with_greeks_and_montecarlo_"
OUT_PATH = os.path.join(FOLDER, "wideform_by_strike_type.xlsx")

KEY_COLUMNS = ["strike", "type", "dte"]
TRACKED_COLUMNS = [
    "spot", "bid", "ask", "volume", "openInterest", "inTheMoney", "impliedVolatility",
    "moneyness", "abs_diff", "pinning_strength", "total_activity",
    "volume_based_pin_rank", "volume_based_pin_candidate",
    "influence_based_pinning_rank", "influence_based_pinning_candidate"
]

def reorder_columns_by_metric(df, key_columns, tracked_metrics):
    all_cols = df.columns.tolist()
    metric_blocks = []
    for metric in tracked_metrics:
        metric_cols = sorted([col for col in all_cols if col.startswith(metric + "_")])
        metric_blocks.extend(metric_cols)
    final_order = key_columns + [col for col in metric_blocks if col in df.columns]
    return df[final_order]

def extract_timestamp(filename):
    base = os.path.basename(filename).replace(".xlsx", "")
    parts = base.split("_")
    date_part = parts[6]
    time_part = "_".join(parts[-2:])
    return f"{date_part}_{time_part}"

def get_recent_files():
    print("Scanning for recent files...")
    files = [f for f in os.listdir(FOLDER) if f.startswith(FILE_PREFIX) and f.endswith(".xlsx")]
    dated = []
    for f in files:
        try:
            date_str = f[len(FILE_PREFIX):].split("_")[0]
            dt = pd.to_datetime(date_str).date()
            dated.append((dt, os.path.join(FOLDER, f)))
        except:
            continue
    dated.sort(reverse=True)
    grouped = {}
    for dt, path in dated:
        grouped.setdefault(dt, []).append(path)
    print(f"✅ Found files for {len(grouped)} recent day(s): {list(grouped.keys())[:10]}")
    recent_paths = []
    for day in list(grouped.keys())[:10]:
        recent_paths.extend(sorted(grouped[day]))
    return recent_paths

def format_excel_sheet(filepath, sheet_name="Sheet1"):
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    ws.freeze_panes = ws["A2"]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(filepath)
    print(f"✅ Formatted sheet: {filepath}")

# === MAIN PROCESS ===
if __name__ == "__main__":
    start_time = datetime.now()
    print(f"Started at: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")

    files = get_recent_files()
    spot_log = []

    # First pass: determine global min/max spot
    min_spot, max_spot = float('inf'), float('-inf')
    for fpath in files:
        try:
            df = pd.read_excel(fpath, sheet_name="options")
            if "spot" in df.columns and not df["spot"].dropna().empty:
                spot = df["spot"].dropna().iloc[0]
                min_spot = min(min_spot, spot)
                max_spot = max(max_spot, spot)
        except Exception as e:
            print(f"❌ Error scanning {fpath}: {e}")
            continue

    lower_bound = int(min_spot) - 10
    upper_bound = int(max_spot) + 10
    print(f"📌 Global strike bounds: {lower_bound} to {upper_bound}")

    wide_df = None

    for fpath in tqdm(files, desc="Processing snapshots"):
        try:
            df = pd.read_excel(fpath, sheet_name="options")
            if "spot" not in df.columns or df["spot"].dropna().empty:
                continue

            spot = df["spot"].dropna().iloc[0]
            timestamp = extract_timestamp(fpath)
            spot_log.append({"timestamp": timestamp, "spot": spot})

            df = df[
                (df["dte"] <= 10) &
                (df["strike"] >= lower_bound) &
                (df["strike"] <= upper_bound)
            ]

            for col in ["strike", "dte"]:
                if col in df.columns:
                    df[col] = df[col].astype(float)

            # Exclude 'spot' from wideform, but still log its time series separately
            columns_to_track = [col for col in TRACKED_COLUMNS if col != "spot"]
            rename_cols = {
                col: f"{col}_{timestamp}" for col in columns_to_track if
                col in df.columns
            }

            df_renamed = df[KEY_COLUMNS + list(rename_cols.keys())].rename(columns=rename_cols)

            if wide_df is None:
                wide_df = df_renamed
            else:
                wide_df = pd.merge(wide_df, df_renamed, on=KEY_COLUMNS, how="outer")

        except Exception as e:
            print(f"❌ Failed to process {fpath}: {e}")
            continue

    if wide_df is not None:
        wide_df = reorder_columns_by_metric(wide_df, KEY_COLUMNS, TRACKED_COLUMNS)
        with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
            wide_df.to_excel(writer, index=False, sheet_name="options_by_strike_type")
            pd.DataFrame(spot_log).to_excel(writer, index=False, sheet_name="spot_history")
        format_excel_sheet(OUT_PATH, sheet_name="options_by_strike_type")
        print(f"\n✅ Exported to:\n{OUT_PATH}")
    else:
        print("❌ No valid snapshots processed.")

    end_time = datetime.now()
    duration = end_time - start_time
    print(f"Finished at: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Elapsed time: {duration}")
