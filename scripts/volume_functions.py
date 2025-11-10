# volume_functions.py — 30-min bands (RTH + ETH) per ticker in TICKER_MAP

import os
from datetime import datetime
from copy import copy

import pandas as pd

from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import TICKER_MAP
from config import DROPBOX_ROOT
US_EASTERN = ZoneInfo("America/New_York")

# ----- Band definitions -----
RTH_BANDS = [
    ("09:30","10:00"),("10:00","10:30"),("10:30","11:00"),("11:00","11:30"),
    ("11:30","12:00"),("12:00","12:30"),("12:30","13:00"),("13:00","13:30"),
    ("13:30","14:00"),("14:00","14:30"),("14:30","15:00"),("15:00","15:30"),
    ("15:30","16:00"),
]
ETH_BANDS = [
    # pre
    ("04:00","04:30"),("04:30","05:00"),("05:00","05:30"),("05:30","06:00"),
    ("06:00","06:30"),("06:30","07:00"),("07:00","07:30"),("07:30","08:00"),
    ("08:00","08:30"),("08:30","09:00"),("09:00","09:30"),
    # post
    ("16:00","16:30"),("16:30","17:00"),("17:00","17:30"),("17:30","18:00"),
    ("18:00","18:30"),("18:30","19:00"),("19:00","19:30"),("19:30","20:00"),
]
RTH_SET = {f"{s}–{e}" for s, e in RTH_BANDS}
ETH_SET = {f"{s}–{e}" for s, e in ETH_BANDS}
ALL_BANDS = RTH_BANDS + ETH_BANDS
ALL_SET = [f"{s}–{e}" for s, e in ALL_BANDS]
SESSION_MAP = ({b: "RTH" for b in RTH_SET} | {b: "ETH" for b in ETH_SET})


def compute_relative_flow(current_volume: float,
                          elapsed_minutes: float,
                          avg_20d_volume: float,
                          bar_minutes: int = 30) -> float:
    """
    Compute real-time relative flow.
    current_volume : volume so far in this bar
    elapsed_minutes : minutes completed in the bar
    avg_20d_volume : 20-day average total volume for this band
    bar_minutes : bar length in minutes (default 30)
    """
    if elapsed_minutes <= 0 or avg_20d_volume <= 0:
        return float("nan")
    # Current flow rate vs. historical flow rate
    current_rate = current_volume / elapsed_minutes
    avg_rate = avg_20d_volume / bar_minutes
    return current_rate / avg_rate

# ----- Fetch 30-min bars -----
def _fetch_30m(ticker: str, days: int) -> pd.DataFrame:
    from ib_insync import IB, Stock, Index, util
    ib = IB()
    ib.connect("127.0.0.1", 7496, clientId=abs(hash(ticker)) % 9000)

    secType, exchange, currency = TICKER_MAP[ticker]
    if secType == "Stock":
        contract = Stock(ticker, exchange, currency)
    else:
        contract = Index(ticker, exchange)

    bars = ib.reqHistoricalData(
        contract,
        endDateTime="",
        durationStr=f"{days} D",
        barSizeSetting="30 mins",
        whatToShow="TRADES",
        useRTH=False,   # ETH + RTH
        formatDate=2,
    )
    print(
        f"[DEBUG] {ticker}: last bar timestamp → {bars[-1].date if bars else 'None'}")
    ib.disconnect()

    df = util.df(bars)
    if df.empty:
        return pd.DataFrame(columns=["timestamp","open","high","low","close","volume","barCount","average"])

    df = df.rename(columns={"date": "timestamp"})
    ts = (pd.to_datetime(df["timestamp"], utc=True)
            .dt.tz_convert(US_EASTERN)
            .dt.tz_localize(None))
    df["timestamp"] = ts
    print(f"Fetching Volume for {ticker}")
    cutoff = pd.Timestamp.now(tz=US_EASTERN).floor("30min").tz_localize(None)
    df = df[df["timestamp"] <= cutoff]
    return df


# ----- Excel formatting -----
def format_sheet(wb, sheetname: str):
    ws = wb[sheetname]
    for cell in ws[1]:
        f = copy(cell.font); f.bold = True; cell.font = f
    ws.freeze_panes = "A2"
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def format_timebands_book(path: str):
    wb = load_workbook(path)
    if "Timebands" in wb.sheetnames:
        ws = wb["Timebands"]
        header = {cell.value: cell.column for cell in ws[1] if cell.value}
        fmt_map = {
            "timestamp": "yyyy-mm-dd hh:mm",
            "generated_at": "yyyy-mm-dd hh:mm",
            "volume": "#,##0",
            "avg_20d": "#,##0",
            "ratio_to_avg_20d": "0.00%",
        }
        for name, fmt in fmt_map.items():
            if name in header:
                col_idx = header[name]
                col_letter = get_column_letter(col_idx)
                for r in range(2, ws.max_row + 1):
                    ws[f"{col_letter}{r}"].number_format = fmt
        format_sheet(wb, "Timebands")
    wb.save(path)


# ----- Main runner -----
def run_timebands_30m(ticker, days=20, include_rth=True, include_eth=True, parent_ticker=None):

    """
    Pull 30-minute historical bars for the given ticker (ETF or futures),
    compute rolling 20-day averages and z-scores, and upload results to Dropbox.
    """
    print(f"\n[run_timebands_30m] Starting for {ticker}...")
    from ib_insync import IB, Stock, Index, Future
    from config import TICKER_MAP, CONTRACT_MAP, get_dropbox_path
    import pandas as pd
    import os
    import numpy as np
    from datetime import datetime
    from zoneinfo import ZoneInfo
    from dropbox_utils import upload_file

    US_EASTERN = ZoneInfo("America/New_York")

    ib = IB()
    try:
        cid = abs(hash(ticker)) % 9000
        ib.connect("127.0.0.1", 7496, clientId=cid)
        print(f"[run_timebands_30m] Connected to IB (clientId={cid})")

        # === Build contract ===
        if ticker in TICKER_MAP:
            stype, exch, curr = TICKER_MAP[ticker]
            if stype == "Stock":
                contract = Stock(ticker, exch, curr)
            elif stype == "Index":
                contract = Index(ticker, exch)
            else:
                raise ValueError(f"Unsupported secType {stype} for {ticker}")
        else:
            if ticker not in CONTRACT_MAP:
                print(f"[run_timebands_30m] No contract info for {ticker}. Skipping.")
                return None

            fut_cfg = CONTRACT_MAP[ticker]
            base_future = Future(
                symbol=fut_cfg["symbol"],
                exchange=fut_cfg["exchange"],
                currency=fut_cfg["currency"],
                multiplier=str(fut_cfg["multiplier"]),
            )
            print(f"[DEBUG] {ticker}: Requesting contract details...")
            contracts = ib.reqContractDetails(base_future)
            if not contracts:
                print(f"[run_timebands_30m] No futures found for {ticker}.")
                return None

            valid = [
                c.contract for c in contracts
                if c.contract.lastTradeDateOrContractMonth
                and len(c.contract.lastTradeDateOrContractMonth) >= 6
            ]
            valid.sort(key=lambda c: c.lastTradeDateOrContractMonth)
            front = valid[0] if valid else contracts[0].contract
            print(f"[DEBUG] {ticker}: available expiries → {[c.lastTradeDateOrContractMonth for c in valid]}")
            contract = front
            print(f"[run_timebands_30m] Using {ticker} front contract: {contract.localSymbol} ({contract.lastTradeDateOrContractMonth})")

        # === Request data ===
        print(f"[run_timebands_30m] Requesting 30m bars for {ticker}...")
        bars = ib.reqHistoricalData(
            contract,
            endDateTime="",
            durationStr=f"{days} D",
            barSizeSetting="30 mins",
            whatToShow="TRADES",
            useRTH=0,
            formatDate=1,
        )
        print(f"[DEBUG] {ticker}: bars retrieved → {len(bars)}")
        if not bars:
            print(f"[run_timebands_30m] No bars returned for {ticker}.")
            return None

        # === Parse dataframe ===
        print(f"[DEBUG] {ticker}: Building DataFrame for export...")
        df = pd.DataFrame(bars)
        raw_ts = pd.to_datetime(df["date"], errors="coerce", utc=True)
        ts = raw_ts.dt.tz_convert(US_EASTERN).dt.tz_localize(None)
        df["timestamp"] = ts
        df["date"] = ts.dt.date
        df["time"] = ts.dt.strftime("%H:%M")
        df["start"] = ts.dt.strftime("%H:%M")
        df["end"] = (ts + pd.Timedelta(minutes=30)).dt.strftime("%H:%M")
        df["band"] = df["start"] + "–" + df["end"]
        df["granularity"] = "30min"
        df["generated_at"] = datetime.now(tz=US_EASTERN).strftime("%Y-%m-%d %H:%M:%S")
        df["session"] = ts.dt.time.between(
            pd.to_datetime("09:30").time(),
            pd.to_datetime("16:00").time(),
        )
        df["session"] = df["session"].map({True: "RTH", False: "ETH"})

        min_date = df["date"].min()
        max_date = df["date"].max()
        print(f"[DEBUG] {ticker}: DataFrame created with {len(df)} rows. Date range: {min_date} → {max_date}")

        # === Rolling stats ===
        print(f"[DEBUG] {ticker}: Calculating rolling 20-day stats...")
        df["avg_20d"] = df.groupby("band", group_keys=False)["volume"].transform(
            lambda x: x.rolling(20, min_periods=1).mean()
        )
        df["stdev_20d"] = df.groupby("band", group_keys=False)["volume"].transform(
            lambda x: x.rolling(20, min_periods=1).std()
        )
        df["zscore_20d"] = (df["volume"] - df["avg_20d"]) / df["stdev_20d"]
        df["ratio_to_avg_20d"] = df["volume"] / df["avg_20d"]

        # --- projected features ---
        df["est_vol_at_close"] = df["ratio_to_avg_20d"]  # default for completed bars

        now = datetime.now(tz=US_EASTERN)
        for i, r in df.iterrows():
            try:
                start_dt = datetime.combine(
                    pd.to_datetime(r["date"]).date(),
                    pd.to_datetime(r["start"]).time(),
                ).replace(tzinfo=US_EASTERN)
                end_dt = datetime.combine(
                    pd.to_datetime(r["date"]).date(),
                    pd.to_datetime(r["end"]).time(),
                ).replace(tzinfo=US_EASTERN)
                if start_dt <= now < end_dt:
                    elapsed = max(0.01, (now - start_dt).total_seconds() / 60)
                    df.at[i, "est_vol_at_close"] = compute_relative_flow(
                        r["volume"], elapsed, r["avg_20d"], bar_minutes=30
                    )
                    break
            except Exception:
                continue

        # === Compute projected z-score AFTER est_vol_at_close ===
        with np.errstate(divide="ignore", invalid="ignore"):
            df["projected_zscore"] = (df["est_vol_at_close"] - 1.0) * (
                df["avg_20d"] / df["stdev_20d"]
            )
        df["projected_zscore"] = pd.to_numeric(df["projected_zscore"], errors="coerce").astype(float)

        print(f"Projected Z dtype: {df['projected_zscore'].dtype}")
        print(f"NaN ratio: {df['projected_zscore'].isna().mean()}")

        # === Write Excel ===
        filename = f"{ticker.lower()}_timeband_volume.xlsx"
        local_filename = os.path.join(r"C:\2mdt\2mindt-site\scripts", filename)
        print(f"[DEBUG] {ticker}: Writing Excel file → {local_filename}")
        with pd.ExcelWriter(local_filename, engine="openpyxl", mode="w") as w:
            df.to_excel(w, sheet_name="Timebands", index=False)

        # === Upload to Dropbox ===
        parent = (parent_ticker or ticker)
        if ticker.lower() == "es":
            category = "es-timebands"
        elif ticker.lower() == "mes":
            category = "mes-timebands"
        else:
            category = "timebands"
        dropbox_path = get_dropbox_path(parent, category, filename)

        print(f"[DEBUG] {ticker}: Uploading file to Dropbox → {dropbox_path}")
        upload_file(local_filename, dropbox_path)
        print(f"[Dropbox] Uploaded {ticker} timebands → {dropbox_path}")

        # === Confirm Excel write success ===
        exported_df = pd.read_excel(local_filename, nrows=5)
        print(f"[DEBUG] {ticker}: Verified Excel contains {len(exported_df)}+ rows. Preview:")
        print(exported_df.head())

        os.remove(local_filename)
        print(f"[Local] Deleted temporary file: {local_filename}")
        print(f"[DEBUG] {ticker}: Final confirmed last date written to Excel → {max_date}")

    except Exception as e:
        print(f"[ERROR] {ticker}: {e}")
    finally:
        try:
            ib.disconnect()
            print(f"[run_timebands_30m] Disconnected from IB for {ticker}.")
        except Exception:
            pass
