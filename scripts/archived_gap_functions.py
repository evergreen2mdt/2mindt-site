# === Standard Libraries ===
import os
import shutil
import sqlite3
from datetime import datetime, timedelta
import warnings
from scipy.stats import trim_mean
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
# === Data Science / Analysis ===
import numpy as np
import pandas as pd
from scipy.stats import trim_mean, iqr
from tqdm import tqdm
from ib_insync import IB
import pandas as pd
import os
from datetime import datetime

from dotenv import load_dotenv
# === Finance / Data Feeds ===
import yfinance as yf
from ib_insync import IB, Stock
from ta import add_all_ta_features

# === Visualization ===
import matplotlib.pyplot as plt
import seaborn as sns

# === Excel Handling ===
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# === External Services / HTTP ===
import requests

# === Supabase (env-driven, single source of truth) ===
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv(dotenv_path="C:/2mdt/2mindt-site/.env")


SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
    raise ValueError("Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in environment variables.")

HEADERS = {
    "apikey": SUPABASE_SERVICE_ROLE_KEY,
    "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
    "Content-Type": "application/json",
}

supabase: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)





#===============================
#Record Time and Create Backup
#===============================




def create_backup(file_path):


    # Record the start time
    start_time = datetime.now()
    print(f"Program started at: {start_time}")

    # Only create backup if the file exists
    if not os.path.exists(file_path):
        print(f"No backup created. File does not exist yet: {file_path}")
        return

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_file_path = f"{file_path.replace('.xlsx', f'_{timestamp}_backup.xlsx')}"

    try:
        shutil.copy(file_path, backup_file_path)
        print(f"Backup created successfully: {backup_file_path}")
    except Exception as e:
        print(f"Error creating backup: {e}")




def get_ib_spy_data(df=None, start_date="2003-09-10"):
    ib = IB()
    try:
        ib.connect('127.0.0.1', 7496, clientId=1)
        print('Connected to IB')
    except Exception as e:
        print(f"Connection failed: {e}")
        return None

    contract = Stock('SPY', 'SMART', 'USD')
    ib_start_date = datetime.strptime(start_date, '%Y-%m-%d')
    ib_end_date = datetime.now()
    duration_days = (ib_end_date - ib_start_date).days
    duration_years = (duration_days // 365) + 1
    duration_str = f'{duration_years} Y'
    print(f"end date: {ib_end_date}")
    bars = ib.reqHistoricalData(
        contract,
        endDateTime=ib_end_date.strftime("%Y%m%d %H:%M:%S"),

        durationStr=duration_str,
        barSizeSetting='1 day',
        whatToShow='TRADES',
        useRTH=True,
        formatDate=1
    )

    if not bars:
        print("No historical data fetched from IB.")
        return None

    all_data = []
    for bar in tqdm(bars, desc="Downloading IB SPY bars"):
        all_data.append({
            'date': bar.date,
            'open': bar.open,
            'high': bar.high,
            'low': bar.low,
            'close': bar.close,
            'volume': bar.volume
        })

    df = pd.DataFrame(all_data)
    df['date'] = pd.to_datetime(df['date']).dt.date
    print("SPY data retrieved from IB.")
    return df



from datetime import timedelta

def get_yf_vix_data(df, start_date):
    df['date'] = pd.to_datetime(df['date']).dt.normalize()
    end_date = df['date'].max()

    print(f"⬇️ Downloading VIX in chunks: {start_date} to {end_date.date()}")

    vix_all = []

    chunk_start = pd.to_datetime(start_date)
    while chunk_start < end_date:
        chunk_end = min(chunk_start + timedelta(days=365 * 5), end_date)

        try:
            print(f"📆 Chunk: {chunk_start.date()} → {chunk_end.date()}")
            vix_chunk = yf.download('^VIX', start=chunk_start.strftime('%Y-%m-%d'),
                                    end=chunk_end.strftime('%Y-%m-%d'),
                                    interval='1d', auto_adjust=True,
                                    threads=False, progress=False)
            if not vix_chunk.empty:
                vix_all.append(vix_chunk)
            else:
                print("⚠️ Empty chunk.")
        except Exception as e:
            print(f"❌ Failed chunk: {e}")
        chunk_start = chunk_end + timedelta(days=1)

    if not vix_all:
        print("❌ All VIX chunks failed or returned empty.")
        return df

    vix = pd.concat(vix_all)
    vix.reset_index(inplace=True)
    vix['date'] = pd.to_datetime(vix['Date']).dt.normalize()
    vix.rename(columns={'Close': 'VIX'}, inplace=True)

    df = pd.merge(df, vix[['date', 'VIX']], on='date', how='left')
    print("✅ VIX successfully merged into SPY DataFrame.")
    return df







def assign_vix_quantile_regimes(df, column='VIX'):
    if column not in df:
        return df

    # drop NA to compute quantiles
    vix_non_na = df[column].dropna()
    q = vix_non_na.quantile([0.25, 0.5, 0.75])

    def regime(v):
        if pd.isna(v):
            return 'unknown'
        elif v <= q[0.25]:
            return 'low'
        elif v <= q[0.5]:
            return 'moderate'
        elif v <= q[0.75]:
            return 'high'
        else:
            return 'extreme'

    tqdm.pandas(desc="Assigning VIX Regimes")
    df['vix_regime'] = df[column].progress_apply(regime)

    return df




def compute_gap_data(df):
    """
    Compute previous day move types, gap type, current day move types,
    trade categories, days to target, and same-day move types.
    """
    df = df.copy()

    prev_day_move_types = []
    gap_types = []
    days_to_target = []
    target_dates = []

    df['year'] = pd.to_datetime(df['date']).dt.year
    df['dow'] = pd.to_datetime(df['date']).dt.day_name()
    df['previous_date'] = df['date'].shift(1)
    df['previous_open'] = df['open'].shift(1)
    df['previous_high'] = df['high'].shift(1)
    df['previous_low'] = df['low'].shift(1)
    df['previous_close'] = df['close'].shift(1)
    df['gap value'] = abs(df['open'] - df['previous_close'])

    data = df.to_dict('records')

    for i in tqdm(range(1, len(data)), desc="Computing gap data"):
        prev_bar = data[i - 1]
        curr_bar = data[i]

        if prev_bar['close'] == prev_bar['open']:
            prev_day_move_type = 'doji'
        elif prev_bar['close'] > prev_bar['open']:
            prev_day_move_type = 'move_up'
        else:
            prev_day_move_type = 'move_down'
        prev_day_move_types.append(prev_day_move_type)

        if prev_bar['close'] < curr_bar['open'] <= prev_bar['high']:
            gap_type = 'price_gap_up'
        elif prev_bar['close'] > curr_bar['open'] >= prev_bar['low']:
            gap_type = 'price_gap_down'
        elif curr_bar['open'] > prev_bar['high']:
            gap_type = 'bar_gap_up'
        elif curr_bar['open'] < prev_bar['low']:
            gap_type = 'bar_gap_down'
        else:
            gap_type = 'no_gap'
        gap_types.append(gap_type)

        target_achieved = False
        days_count = 0
        target_achieved_date = None

        if gap_type == 'no_gap':
            days_to_target.append(0)
            target_dates.append(prev_bar['date'])
        else:
            for j in range(i, len(data)):
                next_bar = data[j]
                days_count += 1
                if gap_type in ['price_gap_up', 'bar_gap_up'] and next_bar['low'] <= prev_bar['close']:
                    target_achieved = True
                    target_achieved_date = next_bar['date']
                    break
                elif gap_type in ['price_gap_down', 'bar_gap_down'] and next_bar['high'] >= prev_bar['close']:
                    target_achieved = True
                    target_achieved_date = next_bar['date']
                    break

            days_to_target.append(days_count if target_achieved else None)
            target_dates.append(target_achieved_date)

    df = df.iloc[1:].reset_index(drop=True).copy()
    df['prev_day_move_types'] = prev_day_move_types
    df['gap_type'] = gap_types
    df['days to target'] = days_to_target

    df['target_achieved_date'] = target_dates

    df['gap_type'] = df['gap_type']
    df['gap value'] = df['gap value']
    df['original_index'] = range(len(df))

    df['trade_direction'] = df['gap_type'].apply(
        lambda x: "not tradable" if "no_gap" in str(x).lower()
        else "long" if "down" in str(x).lower()
        else "short" if "up" in str(x).lower()
        else None
    )

    df["cat_key"] = df["prev_day_move_types"].astype(str) + " + " + df["gap_type"].astype(str)

    def classify_same_day_move(row):
        if row["trade_direction"] == "long":
            if row["open"] > row["close"]:
                return "and_go"
            elif row["open"] < row["close"]:
                return "and_comeback"
            else:
                return "none"
        elif row["trade_direction"] == "short":
            if row["open"] < row["close"]:
                return "and_go"
            elif row["open"] > row["close"]:
                return "and_comeback"
            else:
                return "none"
        else:
            return "none"

    df["same_day_move_type"] = df.apply(classify_same_day_move, axis=1)

    return df




def calculate_statistics(df, window_sizes=[50, 200, 500], trim_percentage=0.1):
    df = df.copy()
    if 'days to target' not in df.columns:
        raise ValueError("'days to target' column is required for calculate_statistics()")

    for window in tqdm(window_sizes, desc="Calculating rolling stats"):
        df[f'MA {window} Days to Close'] = df['days to target'].rolling(window=window, min_periods=1).mean()
        df[f'Rolling Stdev {window} Days'] = df['days to target'].rolling(window=window, min_periods=1).std()

    tqdm.write("Calculating median, trimmed mean/stdev, and quantiles...")
    df['Median'] = df['days to target'].rolling(window=50, min_periods=1).median()
    df['Trim Avg Days to Close 50 Days'] = df['days to target'].rolling(window=50, min_periods=1).apply(
        lambda x: trim_mean(x, proportiontocut=trim_percentage), raw=False
    )
    df['Trim Stdev to Days to Close 50 Days'] = df['days to target'].rolling(window=50, min_periods=1).std()

    df['1st Quartile (25%)'] = df['days to target'].rolling(window=50, min_periods=1).quantile(0.25)
    df['2nd Quartile (50%)'] = df['days to target'].rolling(window=50, min_periods=1).quantile(0.50)
    df['3rd Quartile (75%)'] = df['days to target'].rolling(window=50, min_periods=1).quantile(0.75)
    df['IQR'] = df['3rd Quartile (75%)'] - df['1st Quartile (25%)']
    df['Upper Bound'] = df['3rd Quartile (75%)'] + (1.5 * df['IQR'])

    tqdm.write("Statistics calculation complete.")
    return df




def generate_trade_category_sheets(df1, file_path, calculate_statistics):
    """
    Generate a spreadsheet with sheets for all data and for each unique trade category,
    computing statistics for each category.

    :param df1: DataFrame with a 'trade category' column.
    :param file_path: Path to save the Excel file.
    :param calculate_statistics: Function to calculate statistics on a DataFrame.
    :return: The processed DataFrame.
    """
    # Ensure 'trade category' column exists
    if 'trade category' not in df1.columns:
        print("Error: The DataFrame does not contain a 'trade category' column.")
        return df1

    # Get category counts
    category_counts = df1['trade category'].value_counts()

    # Print category names and their occurrence counts
    print("\nTrade Categories and Counts:")
    for category, count in category_counts.items():
        print(f"{category}: {count} occurrences")


    # Get all unique trade categories
    trade_categories = df1['trade category'].unique()
    print(f"Processing {len(trade_categories)} unique trade categories...")


    return df1  # Returning original DataFrame (df1) if needed




def format_worksheet(workbook, file_path):
    for sheet_name in tqdm(workbook.sheetnames, desc="Formatting sheets"):
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = 'A2'

        # Format column widths
        for col in tqdm(worksheet.columns, desc=f"Adjusting column widths for '{sheet_name}'", leave=False):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            worksheet.column_dimensions[col[0].column_letter].width = max_length + 2

        # Bold header row
        for cell in tqdm(worksheet[1], desc=f"Bolding headers in '{sheet_name}'", leave=False):
            cell.font = Font(bold=True)

    workbook.save(file_path)
    print("Excel formatting applied successfully!")



def create_summary_sheet(file_path):

    # Only continue if file exists
    if not os.path.exists(file_path):
        print(f"Summary sheet skipped. File doesn't exist yet: {file_path}")
        return
    workbook = load_workbook(file_path)
    summary_data = []

    selected_columns = [
        'MA 50 Days to Close', 'MA 200 Days to Close', 'Rolling Stdev 50 Days',
        'Rolling Stdev 200 Days',
        'Trim Avg Days to Close 50 Days', 'Trim Stdev to Days to Close 50 Days',
        '1st Quartile (25%)',
        '2nd Quartile (50%)', '3rd Quartile (75%)', 'IQR', 'Upper Bound'
    ]

    for sheet_name in workbook.sheetnames:
        if sheet_name != "Summary":
            df_sheet = pd.read_excel(file_path, sheet_name=sheet_name,
                                     index_col=0)  # Preserve category-specific index
            if not df_sheet.empty:
                last_row = df_sheet.iloc[-1]
                last_row = last_row[
                    selected_columns]  # Filter only statistical columns
                occurrence_count = len(df_sheet)
                summary_data.append(
                    [sheet_name, occurrence_count] + last_row.tolist())

    if summary_data:
        summary_df = pd.DataFrame(summary_data, columns=["Sheet Name",
                                                         "Occurrences"] + selected_columns)
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
        print("Summary sheet created successfully!")

    # Move Summary sheet to the first position
    workbook = load_workbook(file_path)

    sheet = workbook["Summary"]
    workbook._sheets.insert(0, workbook._sheets.pop(workbook._sheets.index(sheet)))
    workbook.save(file_path)
    print("Summary sheet moved to the first position.")




def commit_to_sqlite(df: pd.DataFrame,
                     db_dir: str = r"C:\2mdt\spy_db",
                     db_name: str = "SPY Gap Analysis.db",
                     table_name: str = "cat_polygon"):
    """
    Save DataFrame to a SQLite DB named 'SPY Gap Analysis.db'.

    Parameters:
    - df: DataFrame to store
    - db_dir: Directory where DB will be saved
    - db_name: File name of the SQLite DB
    - table_name: Table name in the database
    """
    # Ensure the directory exists
    os.makedirs(db_dir, exist_ok=True)
    db_path = os.path.join(db_dir, db_name)

    # Ensure datetime compatibility
    for col in ['date', 'target_achieved_date']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col])

    # Save to SQLite
    conn = sqlite3.connect(db_path)
    df.to_sql(table_name, conn, if_exists='replace', index=True)
    conn.commit()
    conn.close()
    print(f"Committed {len(df)} rows to '{table_name}' in '{db_path}'")

# def save_to_excel(df, file_path, calculate_statistics):
#     print("Starting Excel export...")
#
#     if "target_origination_date" not in df.columns and "date" in df.columns:
#         df["target_origination_date"] = df["date"]
#
#     # Strip timestamps
#     print("Processing datetime columns...")
#     for col in tqdm(['date', 'target_origination_date', 'target_achieved_date', 'mam_date'], desc="Stripping timestamps"):
#         if col in df.columns and pd.api.types.is_datetime64_any_dtype(df[col]):
#             df[col] = df[col].dt.date
#
#     # Round key float columns to two decimals
#     float_cols_to_round = [
#         "open", "high", "low", "close",
#         "previous_open", "previous_high", "previous_low", "previous_close", "gap value",
#         "Drawdown", "MA 50 Days to Close", "Rolling Stdev 50 Days",
#         "MA 200 Days to Close", "Rolling Stdev 200 Days",
#         "MA 500 Days to Close", "Rolling Stdev 500 Days",
#         "Median", "Trim Avg Days to Close 50 Days",
#         "Trim Stdev to Days to Close 50 Days", "max_adverse_move"
#     ]
#     for col in float_cols_to_round:
#         if col in df.columns:
#             df[col] = df[col].round(2)
#
#     # Preserve column order with 'original_index' first
#     column_order = ['original_index'] + [col for col in df.columns if col != 'original_index']
#
#     with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
#         print("Calculating statistics...")
#         df_all_data = df[column_order].copy()
#         df_all_data = calculate_statistics(df_all_data)
#         df_all_data.reset_index(drop=True, inplace=True)
#
#         print("Writing data to Excel...")
#         df_all_data.to_excel(writer, sheet_name='All Data', index=False)
#
#     print("Formatting workbook...")
#     workbook = load_workbook(file_path)
#     format_worksheet(workbook, file_path)
#
#     print(f"Excel file created successfully at {file_path}")
#     return file_path


def save_multiple_sheets_with_formatting(sheet_dict: dict, file_path: str):
    """
    Write multiple DataFrames to a single Excel workbook and format all sheets.

    Args:
        sheet_dict (dict): Dictionary of {sheet_name: DataFrame}
        file_path (str): Full path to output .xlsx file
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from tqdm import tqdm
    import os

    if os.path.exists(file_path):
        book = load_workbook(file_path)
    else:
        book = Workbook()
        book.remove(book.active)

    # Write each sheet
    for sheet_name, df in sheet_dict.items():
        # Remove old version
        if sheet_name in book.sheetnames:
            del book[sheet_name]
        ws = book.create_sheet(title=sheet_name)

        # Normalize datetime columns
        datetime_cols = df.select_dtypes(include=["datetime64[ns]", "datetime64[ns, UTC]"]).columns
        for col in datetime_cols:
            df[col] = df[col].dt.normalize()

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

    # Format all sheets
    for sheetname in tqdm(book.sheetnames, desc="Formatting sheets"):
        ws = book[sheetname]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells), default=0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    book.save(file_path)
    print(f"✅ Workbook written and all sheets formatted: {file_path}")



def open_excel_file(file_path):
    os.startfile(file_path)
    print("Successfully Opened File Path")


def disconnect_ib():
    ib = IB()
    ib.disconnect()
    print("Disconnected from IB")












def big_gaps(df1):
    """
    Identifies gap types in stock price data.
    """
    gap_types = []

    if df1.empty:
        df = df1
        return df

    for i in range(1, len(df1)):
        prev_bar = df1.iloc[i - 1]
        curr_bar = df1.iloc[i]

        if 'close' in prev_bar and 'open' in curr_bar:
            if prev_bar['close'] < curr_bar['open']:
                gap_type = 'gap_up'
            elif prev_bar['close'] > curr_bar['open']:
                gap_type = 'gap_down'
            else:
                gap_type = 'no_gap'
        else:
            gap_type = 'no_data'

        gap_types.append(gap_type)

        if gap_type == 'gap_up':
            df1 = check_short_thresholds(df1, i, prev_bar['close'])
        elif gap_type == 'gap_down':
            df1 = check_long_thresholds(df1, i, prev_bar['close'])
        elif gap_type == 'no_gap':
            df = df1
            return df

    df1 = df1.iloc[1:].reset_index(drop=True)
    df1['gap_type'] = gap_types

    df = df1
    return df


def check_long_thresholds(df1, index, prev_close):
    """
    Checks if the price drops 10 points below the previous day's close before surpassing it.
    """
    for j in range(index, len(df1)):
        if df1.iloc[j]['high'] >= prev_close:
            print(
                f"Target hit on {df1.iloc[j]['date']}: High reached {df1.iloc[j]['high']} before dropping 10 points.")
            return df1
        if df1.iloc[j]['low'] <= prev_close - 10:
            print(
                f"Drop detected on {df1.iloc[j]['date']}: Low reached {df1.iloc[j]['low']} before surpassing {prev_close}.")
            return df1
        df=df1
    return df1


def check_short_thresholds(df1, index, prev_close):
    """
    Checks if the price rises 10 points above the previous day's close before dropping below it.
    """
    for j in range(index, len(df1)):
        if df1.iloc[j]['low'] <= prev_close:
            print(
                f"Target hit on {df1.iloc[j]['date']}: Low reached {df1.iloc[j]['low']} before rising 10 points.")
            return df1
        if df1.iloc[j]['high'] >= prev_close + 10:
            print(
                f"Rise detected on {df1.iloc[j]['date']}: High reached {df1.iloc[j]['high']} before dropping below {prev_close}.")
            df=df1
            return df
    return df1


def calculate_drawdown(df):
    df = df.copy()
    df["Drawdown"] = pd.NA
    trades = df[(df["gap_type"] != "no_gap") & (df["days to target"].notna())]

    for i, row in tqdm(trades.iterrows(), total=trades.shape[0], desc="Calculating Drawdown"):
        try:
            entry_price = row["open"]
            days_to_target = int(row["days to target"])
            direction = "long" if "down" in str(row["gap_type"]).lower() else "short"

            # ✅ Include only rows from entry through target hit day
            if days_to_target <= 0:
                continue  # skip bad rows
            window = df.iloc[i : i + days_to_target]  # do NOT use +1

            if window.empty:
                continue

            if direction == "long":
                drawdown = entry_price - window["low"].min()
            else:
                drawdown = window["high"].max() - entry_price

            df.at[i, "Drawdown"] = drawdown

        except Exception as e:
            print(f"Error calculating drawdown at index {i}: {e}")
            df.at[i, "Drawdown"] = pd.NA

    return df




def compute_gap_data_with_mam_and_days_to_mam(df):
    """
    Computes max adverse move (MAM), MAM date, index, and days to MAM.
    Handles:
    - 'no gap': assigns default values
    - gap + no target hit: fallback MAM scan (entry → end of dataset)
    - gap + target hit: normal MAM scan (entry → target_achieved_date or fallback)
    """
    max_adverse_moves = [None]
    mam_dates = [None]
    mam_indexes = [None]
    days_to_mam = [None]

    df['date'] = pd.to_datetime(df['date'])

    for i in tqdm(range(1, len(df)), desc="Computing MAM"):
        curr_bar = df.iloc[i]
        gap_type = str(curr_bar.get('gap_type', '')).strip().lower().replace('_', ' ')
        entry_price = curr_bar["open"]
        target_achieved_date = curr_bar.get('target_achieved_date', pd.NaT)

        # CASE 1: no gap
        if gap_type == 'no gap':
            mam_dates.append(curr_bar['date'])
            mam_index = i
            mam_indexes.append(mam_index)
            days_to_mam.append(0)
            max_adverse_moves.append(0)
            continue

        # CASE 2: gap with no target achieved (fallback)
        if pd.isnull(target_achieved_date):
            j_end = len(df) - 1
        else:
            match = df[df['date'] == target_achieved_date]
            j_end = match.index[0] if not match.empty else len(df)

        if gap_type in ['price gap down', 'bar gap down']:  # LONG
            min_low = float('inf')
            min_low_index = None
            for j in range(i, j_end + 1):
                low = df.iloc[j]['low']
                if low < min_low:
                    min_low = low
                    min_low_index = j
            mam_value = entry_price - min_low
            mam_date = df.iloc[min_low_index]['date'] if min_low_index is not None else None
            mam_index = min_low_index

        elif gap_type in ['price gap up', 'bar gap up']:  # SHORT
            max_high = -float('inf')
            max_high_index = None
            for j in range(i, j_end + 1):
                high = df.iloc[j]['high']
                if high > max_high:
                    max_high = high
                    max_high_index = j
            mam_value = max_high - entry_price
            mam_date = df.iloc[max_high_index]['date'] if max_high_index is not None else None
            mam_index = max_high_index

        else:
            mam_value = None
            mam_date = None
            mam_index = None

        max_adverse_moves.append(mam_value)
        mam_dates.append(mam_date)
        mam_indexes.append(mam_index)
        days_to_mam.append(mam_index - i + 1 if mam_index is not None else None)

    df['max_adverse_move'] = max_adverse_moves
    df['mam_date'] = mam_dates
    df['mam_index'] = mam_indexes
    df['days_to_mam'] = days_to_mam

    return df



def compute_gap_data_with_mfe(df):
    df = df.copy()
    df.sort_values(by="date", inplace=True)
    df.reset_index(drop=True, inplace=True)

    mfe_values = []
    mfe_dates = []
    mfe_indexes = []
    days_to_mfe = []

    df["date"] = pd.to_datetime(df["date"]).dt.normalize()

    for i, row in tqdm(df.iterrows(), total=len(df), desc="Computing MFE"):
        signal_date = row["date"]
        target_date = row["target_achieved_date"]
        direction = row.get("trade_direction", None)

        if direction == "not tradable":
            mfe_values.append(0)
            mfe_dates.append(signal_date)
            mfe_indexes.append(i)
            days_to_mfe.append(0)
            continue

        # ✅ Use same-day open for entry
        entry_row = row
        entry_price = row["open"]
        entry_index = i
        entry_date = row["date"]

        # 🔍 DEBUG for 2025-05-20
        if row["date"] == pd.to_datetime("2025-05-20"):
            print("\n🔍 DEBUG for 2025-05-20")
            print("Signal Date:", signal_date)
            print("Entry Date:", entry_date)
            print("Entry Price:", entry_price)
            print("Target Achieved Date:", target_date)

        if pd.notna(target_date):
            target_date = pd.to_datetime(target_date).normalize()
            target_match = df[df["date"] == target_date]

            if not target_match.empty:
                target_index = target_match.index[0]
                trading_days = df[(df["date"] >= entry_date) & (df["date"] <= target_date)].shape[0]


                if row["date"] == pd.to_datetime("2025-05-20"):
                    print("Target Match Found:", not target_match.empty)
                    print("Trading Days to Target:", trading_days)
                    print("Target High:", target_match.iloc[0]["high"])

                if trading_days <= 10:
                    target_row = target_match.iloc[0]
                    if direction == "long":
                        mfe = target_row["high"] - entry_price
                    else:
                        mfe = entry_price - target_row["low"]

                    mfe_values.append(mfe)
                    mfe_dates.append(target_row["date"])
                    mfe_indexes.append(target_index)
                    days_to_mfe.append(trading_days)

                    if row["date"] == pd.to_datetime("2025-05-20"):
                        print("✅ MFE:", mfe)

                    continue

        # Fallback 10-day max window
        max_window = df[df.index >= entry_index].head(10)
        if pd.notna(target_date):
            target_index = df[df["date"] == target_date].index[0]
            max_window = max_window[max_window.index <= target_index].head(10)

        if max_window.empty:
            mfe_values.append(np.nan)
            mfe_dates.append(pd.NaT)
            mfe_indexes.append(None)
            days_to_mfe.append(None)
            continue

        if direction == "long":
            best_idx = max_window["high"].idxmax()
            best_row = max_window.loc[best_idx]
            mfe = best_row["high"] - entry_price
        elif direction == "short":
            best_idx = max_window["low"].idxmin()
            best_row = max_window.loc[best_idx]
            mfe = entry_price - best_row["low"]
        else:
            mfe = np.nan
            best_row = None

        mfe_values.append(mfe)
        mfe_dates.append(best_row["date"] if best_row is not None else pd.NaT)
        mfe_indexes.append(best_idx if best_row is not None else None)
        days_to_mfe.append(best_idx - i if best_row is not None else None)

    df["max_favorable_move"] = mfe_values
    df["mfe_date"] = pd.Series(pd.to_datetime(mfe_dates)).dt.normalize()
    df["mfe_index"] = mfe_indexes
    df["days_to_mfe"] = days_to_mfe

    return df















def add_ta_indicators(df):
    df = df.copy()

    # Ensure necessary columns exist
    required_cols = ['open', 'high', 'low', 'close', 'volume']
    if not all(col in df.columns for col in required_cols):
        print("Missing required OHLCV columns for technical indicators.")
        return df

    print("Adding technical indicators...")

    # -- Commenting out warning suppression to verify fix --

    # with warnings.catch_warnings():
    #     warnings.simplefilter("ignore", category=pd.errors.PerformanceWarning)
    df = add_all_ta_features(
        df,
        open="open",
        high="high",
        low="low",
        close="close",
        volume="volume",
        fillna=True
    )

    print("Technical indicators added using 'ta' library.")

    # ✅ This actually fixes the fragmentation
    return df.copy()


import pandas as pd
from openpyxl import load_workbook
def create_cat_key_summary_sheets_with_tables(df, output_path):
    df["cat_key"] = df["prev_day_move_types"].astype(str) + " + " + df["gap_type"].astype(str)

    grouped = df.groupby("cat_key")

    rolling_stats_cat_keys = {}
    days_to_target_cat_keys = {}
    mam_cat_keys = {}
    days_to_hit_chart = {}

    # Rolling stats
    for cat, group in grouped:
        group = group.copy()
        group["max_adverse_move"] = pd.to_numeric(group["max_adverse_move"], errors="coerce")
        stats = {
            "median_mam": group["max_adverse_move"].median(),
            "trimmed_mean_mam": group["max_adverse_move"].mean(),
            "std_mam": group["max_adverse_move"].std()
        }
        rolling_stats_cat_keys[cat] = stats

    # Days to target stats
    for cat, group in grouped:
        group = group.copy()
        counts = group["days_to_target"].value_counts().sort_index()
        dist = {f"day_{i}": counts.get(i, 0) for i in range(11)}
        days_to_target_cat_keys[cat] = dist

    # MAM stats
    for cat, group in grouped:
        group = group.copy()
        group["max_adverse_move"] = pd.to_numeric(group["max_adverse_move"], errors="coerce")
        stats = {
            "mean_mam": group["max_adverse_move"].mean(),
            "max_mam": group["max_adverse_move"].max(),
            "min_mam": group["max_adverse_move"].min()
        }
        mam_cat_keys[cat] = stats

    # Days-to-hit chart
    for cat, group in grouped:
        total = len(group)
        cumulative = []
        conditional = []
        cum_hits = 0

        for i in range(11):
            day_hits = (group["days_to_target"] == i).sum()
            cum_hits += day_hits
            cumulative.append(round(100 * cum_hits / total, 1))
            conditional.append(round(100 * day_hits / total, 1))

        days_to_hit_chart[cat] = {
            "cumulative": cumulative,
            "conditional": conditional
        }

    print("⚠️  Make sure Excel is closed before writing sheets.")
    print("📄 Writing summary sheets to Excel...")

    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        pd.DataFrame(rolling_stats_cat_keys).T.to_excel(writer, sheet_name='MAM Stats Cat Keys')
        pd.DataFrame(days_to_target_cat_keys).T.to_excel(writer, sheet_name='Days to Target Stats Cat Keys')
        pd.DataFrame(mam_cat_keys).T.to_excel(writer, sheet_name='%% MAM Stats Cat Keys')
        pd.DataFrame({k: v["cumulative"] for k, v in days_to_hit_chart.items()}).T.to_excel(writer, sheet_name='%% Days to Hit by Cat Keys')

    print("✅ Summary sheets written to Excel.")

    return {
        "rolling_stats_cat_keys": rolling_stats_cat_keys,
        "days_to_target_cat_keys": days_to_target_cat_keys,
        "mam_cat_keys": mam_cat_keys,
        "days_to_hit_chart": days_to_hit_chart
    }






def reset_excel_workbook(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)

    wb = Workbook()
    wb.save(file_path)



def write_all_data_sheet(file_path, df):
    book = load_workbook(file_path)
    sheet_name = "All Data"
    if sheet_name in book.sheetnames:
        del book[sheet_name]
    ws = book.create_sheet(title=sheet_name)

    for row in dataframe_to_rows(df.reset_index(), index=False, header=True):
        ws.append(row)

    book.save(file_path)

def create_cat_key_summary_sheets_with_tables(df, output_path):
    import numpy as np

    df["cat_key"] = df["prev_day_move_types"].astype(str) + " + " + df["gap_type"].astype(str)
    grouped = df.groupby("cat_key")

    rolling_stats_cat_keys = {}
    days_to_target_cat_keys = {}
    mam_cat_keys = {}
    days_to_hit_chart = {}

    total_rows = len(df)
    total_outliers = 0
    outlier_tracker = {}

    for cat, group in grouped:
        group = group.copy()
        group["max_adverse_move"] = pd.to_numeric(group["max_adverse_move"], errors="coerce")
        mam = group["max_adverse_move"].dropna()

        q1 = mam.quantile(0.25)
        q3 = mam.quantile(0.75)
        iqr = q3 - q1
        upper_bound = q3 + 1.5 * iqr

        outliers = mam > upper_bound
        outlier_count = outliers.sum()

        mad = np.median(np.abs(mam - np.median(mam))) if len(mam) > 1 else 0

        outlier_tracker[cat] = outlier_count
        total_outliers += outlier_count

        stats = {
            "median_mam": mam.median(),
            "trimmed_mean_mam": mam.mean(),
            "std_mam": mam.std(),
            "mad_mam": mad,
            "rolling_q1": q1,
            "rolling_q3": q3,
            "rolling_upper_bound": upper_bound,
            "total_occurrences": len(group),
            "outlier_count": outlier_count,
        }

        rolling_stats_cat_keys[cat] = stats

    # Add occurrence/outlier % and ratio
    for cat, stats in rolling_stats_cat_keys.items():
        occurrence_pct = stats["total_occurrences"] / total_rows if total_rows else 0
        outlier_pct = outlier_tracker[cat] / total_outliers if total_outliers else 0
        ratio = occurrence_pct / outlier_pct if outlier_pct else 0

        stats["occurrence_%"] = round(occurrence_pct * 100, 2)
        stats["outlier_%"] = round(outlier_pct * 100, 2)
        stats["occurrence_to_outlier_ratio"] = round(ratio, 2)

    # Days to target stats
    for cat, group in grouped:
        group = group.copy()
        group["days to target"] = pd.to_numeric(group["days to target"], errors="coerce")
        counts = group["days to target"].value_counts().sort_index()
        dist = {f"day_{i}": counts.get(i, 0) for i in range(11)}
        days_to_target_cat_keys[cat] = dist

    # MAM stats
    for cat, group in grouped:
        group = group.copy()
        group["max_adverse_move"] = pd.to_numeric(group["max_adverse_move"], errors="coerce")
        stats = {
            "mean_mam": group["max_adverse_move"].mean(),
            "max_mam": group["max_adverse_move"].max(),
            "min_mam": group["max_adverse_move"].min()
        }
        mam_cat_keys[cat] = stats

    # Hit chart data
    for cat, group in grouped:
        group = group.copy()
        group["days to target"] = pd.to_numeric(group["days to target"], errors="coerce")
        total = len(group)
        cumulative = []
        conditional = []
        cum_hits = 0

        for i in range(11):
            day_hits = (group["days to target"] == i).sum()
            cum_hits += day_hits
            cumulative.append(round(100 * cum_hits / total, 1))
            conditional.append(round(100 * day_hits / total, 1))

        days_to_hit_chart[cat] = {
            "cumulative": cumulative,
            "conditional": conditional
        }

    # Write to Excel
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        pd.DataFrame(rolling_stats_cat_keys).T.to_excel(writer, sheet_name="MAM Stats Cat Keys")
        pd.DataFrame(days_to_target_cat_keys).T.to_excel(writer, sheet_name="Days to Target Stats Cat Keys")
        pd.DataFrame(mam_cat_keys).T.to_excel(writer, sheet_name="%% MAM Stats Cat Keys")
        pd.DataFrame({k: v["cumulative"] for k, v in days_to_hit_chart.items()}).T.to_excel(
            writer, sheet_name="%% Days to Hit by Cat Keys"
        )

    return {
        "rolling_stats_cat_keys": rolling_stats_cat_keys,
        "days_to_target_cat_keys": days_to_target_cat_keys,
        "mam_cat_keys": mam_cat_keys,
        "days_to_hit_chart": days_to_hit_chart
    }





def create_days_to_target_summary_table_cat_key(file_path):


    # Load and prep data
    df = pd.read_excel(file_path, sheet_name="All Data")
    df["cat_key"] = df["prev_day_move_types"].astype(str) + " + " + df["gap_type"].astype(str)
    df = df[df["days to target"].notna()]

    max_day = 10
    headers = ["Label"] + [f"Days {i}" for i in range(max_day + 1)]
    all_rows = []
    # ALL GAPS summary block
    total = len(df)
    counts = [(df["days to target"] <= i).sum() for i in range(max_day + 1)]
    cumulative_pct = [round(c / total, 6) if total else 0.0 for c in counts]

    conditional_pct = []
    for i in range(len(counts)):
        if i == 0:
            conditional_pct.append(round(counts[0] / total, 6) if total else 0.0)
        else:
            new_hits = counts[i] - counts[i - 1]
            remaining = total - counts[i - 1]
            conditional = round(new_hits / remaining, 6) if remaining > 0 else 0.0
            conditional_pct.append(conditional)

    all_rows.append(["all_gaps - Total: {}".format(total)] + [""] * (max_day + 1))
    all_rows.append(["Count"] + counts)
    all_rows.append(["Cumulative %"] + cumulative_pct)
    all_rows.append(["Conditional %"] + conditional_pct)

    for key in sorted(df["cat_key"].unique()):
        sub_df = df[df["cat_key"] == key]
        total = len(sub_df)

        counts = [(sub_df["days to target"] <= i).sum() for i in range(max_day + 1)]
        cumulative_pct = [round(c / total, 6) if total else 0.0 for c in counts]

        conditional_pct = []
        for i in range(len(counts)):
            if i == 0:
                conditional_pct.append(round(counts[0] / total, 6) if total else 0.0)
            else:
                new_hits = counts[i] - counts[i - 1]
                remaining = total - counts[i - 1]
                conditional = round(new_hits / remaining, 6) if remaining > 0 else 0.0
                conditional_pct.append(conditional)

        all_rows.append([f"{key} - Total: {total}"] + [""] * (max_day + 1))
        all_rows.append(["Count"] + counts)
        all_rows.append(["Cumulative %"] + cumulative_pct)
        all_rows.append(["Conditional %"] + conditional_pct)

    # Convert to DataFrame
    summary_df = pd.DataFrame(all_rows, columns=headers)

    # Write to Excel
    book = load_workbook(file_path)
    sheet_name = "%% Days to Hit by Cat Keys"
    if sheet_name in book.sheetnames:
        del book[sheet_name]
    ws = book.create_sheet(title=sheet_name)

    # Define alternating fill colors (more readable)
    fills = [
        PatternFill(start_color="E6F4F1", end_color="E6F4F1", fill_type="solid"),
        PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
        PatternFill(start_color="F1E6F4", end_color="F1E6F4", fill_type="solid"),
    ]
    align_center = Alignment(horizontal="center", vertical="center")

    # Write headers
    for row in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(row)

    # Apply formatting row-by-row
    current_fill = None
    fill_index = 0
    for i in range(2, ws.max_row + 1, 4):  # Start after header, every 4 rows
        current_fill = fills[fill_index % len(fills)]
        fill_index += 1

        for r in range(i, min(i + 4, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = current_fill
                cell.alignment = align_center
                if ws.cell(row=r, column=1).value in ["Cumulative %", "Conditional %"] and c > 1:
                    cell.number_format = "0.0%"
            if ws.cell(row=r, column=1).value and "Total" in str(ws.cell(row=r, column=1).value):
                for cell in ws[r]:
                    cell.font = Font(bold=True)

    book.save(file_path)




def create_days_to_target_summary_table_gap_types(file_path):


    # Load and prep data
    df = pd.read_excel(file_path, sheet_name="All Data")
    df["gap_type"] = df["gap_type"].astype(str)
    df = df[df["days to target"].notna()]

    max_day = 10
    headers = ["Label"] + [f"Days {i}" for i in range(max_day + 1)]
    all_rows = []
    # ALL GAPS summary block
    total = len(df)
    counts = [(df["days to target"] <= i).sum() for i in range(max_day + 1)]
    cumulative_pct = [round(c / total, 6) if total else 0.0 for c in counts]

    conditional_pct = []
    for i in range(len(counts)):
        if i == 0:
            conditional_pct.append(round(counts[0] / total, 6) if total else 0.0)
        else:
            new_hits = counts[i] - counts[i - 1]
            remaining = total - counts[i - 1]
            conditional = round(new_hits / remaining, 6) if remaining > 0 else 0.0
            conditional_pct.append(conditional)

    all_rows.append(["all_gaps - Total: {}".format(total)] + [""] * (max_day + 1))
    all_rows.append(["Count"] + counts)
    all_rows.append(["Cumulative %"] + cumulative_pct)
    all_rows.append(["Conditional %"] + conditional_pct)

    for key in sorted(df["gap_type"].unique()):
        sub_df = df[df["gap_type"] == key]
        total = len(sub_df)

        counts = [(sub_df["days to target"] <= i).sum() for i in range(max_day + 1)]
        cumulative_pct = [round(c / total, 6) if total else 0.0 for c in counts]

        conditional_pct = []
        for i in range(len(counts)):
            if i == 0:
                conditional_pct.append(round(counts[0] / total, 6) if total else 0.0)
            else:
                new_hits = counts[i] - counts[i - 1]
                remaining = total - counts[i - 1]
                conditional = round(new_hits / remaining, 6) if remaining > 0 else 0.0
                conditional_pct.append(conditional)

        all_rows.append([f"{key} - Total: {total}"] + [""] * (max_day + 1))
        all_rows.append(["Count"] + counts)
        all_rows.append(["Cumulative %"] + cumulative_pct)
        all_rows.append(["Conditional %"] + conditional_pct)

    # Convert to DataFrame
    summary_df = pd.DataFrame(all_rows, columns=headers)

    # Write to Excel
    book = load_workbook(file_path)
    sheet_name = "Days to Target Gap Types"
    if sheet_name in book.sheetnames:
        del book[sheet_name]
    ws = book.create_sheet(title=sheet_name)

    for row in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(row)

    # Apply formatting
    fills = [
        PatternFill(start_color="E6F4F1", end_color="E6F4F1", fill_type="solid"),
        PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
        PatternFill(start_color="F1E6F4", end_color="F1E6F4", fill_type="solid"),
    ]
    align_center = Alignment(horizontal="center", vertical="center")

    current_fill = None
    fill_index = 0
    for i in range(2, ws.max_row + 1, 4):  # Start after header
        current_fill = fills[fill_index % len(fills)]
        fill_index += 1

        for r in range(i, min(i + 4, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = current_fill
                cell.alignment = align_center
                if ws.cell(row=r, column=1).value in ["Cumulative %", "Conditional %"] and c > 1:
                    cell.number_format = "0.0%"
            if ws.cell(row=r, column=1).value and "Total" in str(ws.cell(row=r, column=1).value):
                for cell in ws[r]:
                    cell.font = Font(bold=True)

    book.save(file_path)


def tag_outliers_by_cat_key_and_gap_type(file_path):


    df = pd.read_excel(file_path, sheet_name="All Data")
    df["cat_key"] = df["prev_day_move_types"].astype(str) + " + " + df["gap_type"].astype(str)
    df = df[df["days to target"].notna()]
    trading_days_total = len(df)

    def calc_stats(df, group_col):
        result = []
        grouped = df.groupby(group_col)
        total_outliers = 0
        interim = []

        for name, group in grouped:
            values = group["days to target"].dropna()
            q1 = values.quantile(0.25)
            q3 = values.quantile(0.75)
            iqr = q3 - q1
            upper_bound = q3 + 1.5 * iqr
            outliers = (values > upper_bound).sum()
            total = len(values)

            interim.append({
                "key": name,
                "Outlier threshold (days)": round(upper_bound, 2),
                "Total occurrences": total,
                "Number of outliers": outliers,
                "% Occurrence": total / trading_days_total if trading_days_total else 0,
            })
            total_outliers += outliers

        for row in interim:
            percent_outliers = row["Number of outliers"] / total_outliers if total_outliers else 0
            ratio = (row["% Occurrence"] / percent_outliers) if percent_outliers > 0 else 0
            row["% Outliers per line"] = percent_outliers
            row["Occurrence to Outlier Ratio"] = round(ratio, 2)
            result.append(row)

        return pd.DataFrame(result)

    # Build base tables
    cat_key_df = calc_stats(df, "cat_key").assign(type="cat_key")
    gap_type_df = calc_stats(df, "gap_type").assign(type="gap_type")

    # Add final row: all_gaps
    values = df["days to target"]
    q1 = values.quantile(0.25)
    q3 = values.quantile(0.75)
    iqr = q3 - q1
    upper_bound = q3 + 1.5 * iqr
    outliers = (values > upper_bound).sum()
    all_row = {
        "key": "all_gaps",
        "Outlier threshold (days)": round(upper_bound, 2),
        "Total occurrences": len(values),
        "Number of outliers": outliers,
        "% Occurrence": 1.0,
        "% Outliers per line": 1.0,
        "Occurrence to Outlier Ratio": 1.0,
        "type": "all"
    }

    final_df = pd.concat([cat_key_df, gap_type_df, pd.DataFrame([all_row])], ignore_index=True)

    # Format percentages
    final_df["% Occurrence"] = final_df["% Occurrence"].round(4)
    final_df["% Outliers per line"] = final_df["% Outliers per line"].round(4)

    # Write to Excel
    book = load_workbook(file_path)
    if "Outliers" in book.sheetnames:
        del book["Outliers"]
    ws = book.create_sheet("Outliers")

    for row in dataframe_to_rows(final_df, index=False, header=True):
        ws.append(row)

    # Add table
    col_count = ws.max_column
    row_count = ws.max_row
    last_col_letter = chr(64 + col_count) if col_count <= 26 else f"{chr(64 + (col_count - 1) // 26)}{chr(65 + (col_count - 1) % 26)}"
    table_ref = f"A1:{last_col_letter}{row_count}"
    table = Table(displayName="OutlierSummary", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    # Apply number formatting
    headers = [cell.value for cell in ws[1]]
    pct_cols = ["% Occurrence", "% Outliers per line"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_name in pct_cols:
            col_idx = headers.index(col_name)
            cell = row[col_idx]
            cell.number_format = "0.00%"

    # Apply conditional formatting to "Occurrence to Outlier Ratio"
    if "Occurrence to Outlier Ratio" in headers:
        ratio_col_idx = headers.index("Occurrence to Outlier Ratio") + 1
        col_letter = chr(64 + ratio_col_idx) if ratio_col_idx <= 26 else f"{chr(64 + (ratio_col_idx - 1) // 26)}{chr(65 + (ratio_col_idx - 1) % 26)}"
        data_range = f"{col_letter}2:{col_letter}{row_count}"

        ws.conditional_formatting.add(data_range, CellIsRule(
            operator='lessThan', formula=['1'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        ))
        ws.conditional_formatting.add(data_range, CellIsRule(
            operator='greaterThanOrEqual', formula=['1'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        ))

    book.save(file_path)








def analyze_cat_key_by_vix_regime(file_path):
    df = pd.read_excel(file_path, sheet_name="All Data")
    df["target_origination_date"] = pd.to_datetime(df["target_origination_date"])
    df["cat_key"] = df["prev_day_move_types"].astype(str) + " + " + df["gap_type"].astype(str)

    # Filter cat_keys with enough samples and valid target info
    valid_cat_keys = df["cat_key"].value_counts()
    valid_cat_keys = valid_cat_keys[valid_cat_keys > 800].index.tolist()
    df = df[df["cat_key"].isin(valid_cat_keys)]
    df = df[df["days to target"].notna()]

    #  Calculate reward/risk ratio safely
    df["reward_risk_ratio"] = np.where(
        (df["max_adverse_move"] > 0) & (df["max_favorable_move"].notna()),
        df["max_favorable_move"] / df["max_adverse_move"],
        np.nan
    )

    # Group stats
    grouped = df.groupby(["cat_key", "vix_regime"])
    summary = grouped["days to target"].agg(
        count="count",
        mean="mean",
        median="median",
        std="std",
        trimmed_mean=lambda x: trim_mean(x.dropna(), 0.1) if len(x.dropna()) >= 5 else None,
        q1=lambda x: np.percentile(x.dropna(), 25),
        q3=lambda x: np.percentile(x.dropna(), 75),
        iqr=lambda x: iqr(x.dropna()),
        mad=lambda x: np.median(np.abs(x.dropna() - np.median(x.dropna())))
    ).reset_index()

    mfe_summary = grouped["max_favorable_move"].mean().reset_index(name="avg_mfe")
    mae_summary = grouped["max_adverse_move"].mean().reset_index(name="avg_mae")
    rr_summary = grouped["reward_risk_ratio"].mean().reset_index(name="avg_reward_risk")

    summary = summary.merge(mfe_summary, on=["cat_key", "vix_regime"])
    summary = summary.merge(mae_summary, on=["cat_key", "vix_regime"])
    summary = summary.merge(rr_summary, on=["cat_key", "vix_regime"])

    # Add consistency flags
    stds = summary.groupby("cat_key")["std"].std().reset_index(name="std_across_regimes")
    stds["is_consistent"] = stds["std_across_regimes"] <= 1.5
    summary = pd.merge(summary, stds, on="cat_key", how="left")

    # Write to Excel as a table
    book = load_workbook(file_path)
    sheet_name = "Cat Key by VIX Regime"
    if sheet_name in book.sheetnames:
        del book[sheet_name]
    ws = book.create_sheet(title=sheet_name)

    for row in dataframe_to_rows(summary, index=False, header=True):
        ws.append(row)

    # Format as table
    table_range = f"A1:{chr(65 + len(summary.columns) - 1)}{len(summary) + 1}"
    tab = Table(displayName="CatKeyVIXTable", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    book.save(file_path)

    print(f"Saved to '{sheet_name}' as an Excel Table")

    # Save heatmap image
    pivot = summary.pivot(index="cat_key", columns="vix_regime", values="mean")
    plt.figure(figsize=(12, 8))

    sns.heatmap(pivot, annot=True, fmt=".1f", cmap="coolwarm", cbar_kws={"label": "Avg Days to Target"})
    plt.title("Avg Days to Target by Cat Key and VIX Regime")
    plt.ylabel("Cat Key")
    plt.xlabel("VIX Regime")
    plt.tight_layout()
    plt.savefig("vix_regime_heatmap.png", dpi=300)
    plt.close()

def compute_rolling_stats_by_gap_type(file_path):
    df = pd.read_excel(file_path, sheet_name="All Data")
    df["gap_type"] = df["gap_type"].astype(str)
    df = df[df["days to target"].notna()]
    df = df.copy()

    rolling_stats = df.groupby("gap_type")["days to target"].agg([
        ("rolling_mean", "mean"),
        ("rolling_std", "std"),
        ("rolling_trim_mean", lambda x: trim_mean(x, proportiontocut=0.1)),
        ("rolling_trim_std",
         lambda x: x[(x >= x.quantile(0.1)) & (x <= x.quantile(0.9))].std()),
        ("rolling_q1", lambda x: x.quantile(0.25)),
        ("rolling_median", "median"),
        ("rolling_q3", lambda x: x.quantile(0.75)),
        ("rolling_upper_bound", lambda x: x.quantile(0.75) + 1.5 * (
                    x.quantile(0.75) - x.quantile(0.25))),
        ("rolling_mad", lambda x: (x - x.median()).abs().mean()),
    ]).reset_index()

    counts = df.groupby("gap_type")["days to target"].agg([
        ("total_occurrences", "count"),
        ("outlier_count", lambda x: ((x > (x.quantile(0.75) + 1.5 * (x.quantile(0.75) - x.quantile(0.25)))).sum()))
    ]).reset_index()

    summary = pd.merge(rolling_stats, counts, on="gap_type", how="left")
    summary["occurrence_%"] = summary["total_occurrences"] / summary["total_occurrences"].sum()
    summary["outlier_%"] = summary["outlier_count"] / summary["total_occurrences"]
    summary["occurrence_to_outlier_ratio"] = summary["total_occurrences"] / (summary["outlier_count"] + 1)

    wb = load_workbook(file_path)
    if "Rolling Stats by Gap Type" in wb.sheetnames:
        del wb["Rolling Stats by Gap Type"]
    ws = wb.create_sheet("Rolling Stats by Gap Type")


    for row in dataframe_to_rows(summary, index=False, header=True):
        ws.append(row)

    # Add table formatting
    table_ref = f"A1:{chr(64 + summary.shape[1])}{len(summary)+1}"
    table = Table(displayName="GapTypeSummaryTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(file_path)
    print("✅ Rolling by gap_type written to 'Rolling Stats by Gap Type'")


def compute_dynamic_index_rolling_stats(file_path):
    df = pd.read_excel(file_path, sheet_name="All Data")
    df["target_origination_date"] = pd.to_datetime(df["target_origination_date"])
    df = df.sort_values("target_origination_date")
    df["cat_key"] = df["prev_day_move_types"].astype(str) + " + " + df["gap_type"].astype(str)

    all_cat_keys = df["cat_key"].unique()
    stats_records = []

    for cat_key in all_cat_keys:
        cat_df = df[df["cat_key"] == cat_key].copy()
        cat_df = cat_df[["target_origination_date", "days to target"]].dropna().reset_index(drop=True)

        total_points = len(cat_df)
        if total_points < 1:
            continue

        for w in range(1, total_points + 1):
            recent_data = cat_df["days to target"].iloc[-w:]
            median = np.median(recent_data)
            q1 = np.percentile(recent_data, 25)
            q3 = np.percentile(recent_data, 75)
            upper_bound = q3 + 1.5 * (q3 - q1)
            mean = np.mean(recent_data)
            std = np.std(recent_data, ddof=1) if len(recent_data) > 1 else 0
            mad = np.median(np.abs(recent_data - median)) if len(recent_data) > 1 else 0
            last_date = cat_df["target_origination_date"].iloc[-1]

            stats_records.append({
                "cat_key": cat_key,
                "rolling_window": w,
                "latest_date_in_window": last_date,
                "rolling_mean": mean,
                "rolling_std": std,
                "rolling_median": median,
                "rolling_q1": q1,
                "rolling_q3": q3,
                "rolling_upper_bound": upper_bound,
                "rolling_mad": mad
            })

    stats_df = pd.DataFrame(stats_records)

    # Save rolling stats
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        stats_df.to_excel(writer, sheet_name="rolling_by_index", index=False)

    # ------------------ Summary Table ------------------
    df_all = df.copy()
    df_all = df_all[~df_all["cat_key"].str.contains("no_gap", na=False)]

    summary_df = stats_df.groupby("cat_key")[[
        "rolling_mean", "rolling_std", "rolling_median", "rolling_q1",
        "rolling_q3", "rolling_upper_bound", "rolling_mad"
    ]].max().reset_index()

    # Add total occurrences and outlier counts
    occurrence_counts = df_all["cat_key"].value_counts().rename(
        "total_occurrences").reset_index()
    occurrence_counts.columns = ["cat_key", "total_occurrences"]

    outlier_counts = df_all[df_all["days to target"] > 10][
        "cat_key"].value_counts().rename("outlier_count").reset_index()
    outlier_counts.columns = ["cat_key", "outlier_count"]

    summary_df = summary_df.merge(occurrence_counts, on="cat_key", how="left")
    summary_df = summary_df.merge(outlier_counts, on="cat_key", how="left")
    summary_df["outlier_count"] = summary_df["outlier_count"].fillna(0).astype(
        int)

    total_days = df_all["target_origination_date"].nunique()
    total_outliers = summary_df["outlier_count"].sum()
    summary_df["occurrence_%"] = summary_df["total_occurrences"] / total_days
    summary_df["outlier_%"] = summary_df["outlier_count"] / total_outliers
    summary_df["occurrence_to_outlier_ratio"] = summary_df["occurrence_%"] / \
                                                summary_df["outlier_%"]

    # Write into rolling_by_index sheet starting at column L
    book = load_workbook(file_path)
    ws = book["rolling_by_index"]

    start_col = 12  # Column L
    for r_idx, row in enumerate(
            dataframe_to_rows(summary_df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Format as Excel Table
    from openpyxl.utils import get_column_letter
    end_col_letter = get_column_letter(start_col + len(summary_df.columns) - 1)
    table_range = f"{get_column_letter(start_col)}1:{end_col_letter}{len(summary_df) + 1}"

    summary_table = Table(displayName="CatKeySummaryTable", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    summary_table.tableStyleInfo = style
    ws.add_table(summary_table)

    book.save(file_path)

    print(
        "Rolling stats and enhanced summary table saved to 'rolling_by_index'.")



def compute_stability_summary(file_path):


    df = pd.read_excel(file_path, sheet_name="rolling_by_index")

    # Compute IQR
    df["iqr"] = df["rolling_q3"] - df["rolling_q1"]

    # Aggregate per cat_key
    agg = df.groupby("cat_key").agg(
        median_std=("rolling_median", "std"),
        avg_iqr=("iqr", "mean"),
        avg_mad=("rolling_mad", "mean"),
        mean_of_mean=("rolling_mean", "mean"),
        mean_q1=("rolling_q1", "mean"),
        mean_q3=("rolling_q3", "mean"),
        mean_upper_bound=("rolling_upper_bound", "mean"),
        count=("rolling_window", "count")
    ).reset_index()

    # Load All Data sheet to calculate outlier ratio
    df_all = pd.read_excel(file_path, sheet_name="All Data")
    df_all["cat_key"] = df_all["prev_day_move_types"].astype(str) + " + " + df_all["gap_type"].astype(str)

    # Occurrence and outlier counts
    occ = df_all["cat_key"].value_counts().rename("total_occurrences").reset_index()
    occ.columns = ["cat_key", "total_occurrences"]

    outliers = df_all[df_all["days to target"] > 10]["cat_key"].value_counts().rename("outlier_count").reset_index()
    outliers.columns = ["cat_key", "outlier_count"]

    # Merge into agg
    merged = agg.merge(occ, on="cat_key", how="left")
    merged = merged.merge(outliers, on="cat_key", how="left")
    merged["outlier_count"] = merged["outlier_count"].fillna(0).astype(int)
    merged["occurrence_to_outlier_ratio"] = merged["total_occurrences"] / merged["outlier_count"].replace(0, np.nan)

    # Filter to cat_keys with at least 450 rolling windows
    filtered = merged[merged["count"] >= 450].copy()

    # Normalize for score
    for col in ["median_std", "avg_iqr", "avg_mad"]:
        min_val = filtered[col].min()
        max_val = filtered[col].max()
        filtered[f"{col}_norm"] = (filtered[col] - min_val) / (max_val - min_val)

    # Stability score
    filtered["stability_score"] = (
        filtered["median_std_norm"] +
        filtered["avg_iqr_norm"] +
        filtered["avg_mad_norm"]
    ) / 3

    filtered.sort_values("stability_score", inplace=True)

    # Save to Excel
    book = load_workbook(file_path)
    if "stability" in book.sheetnames:
        del book["stability"]
    ws = book.create_sheet("stability")

    for r_idx, row in enumerate(dataframe_to_rows(filtered, index=False, header=True), start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    book.save(file_path)
    print("Stability summary with occurrence-to-outlier ratio saved to 'stability' sheet.")



def get_daily_mam_progression(df):
    """
    Build a daily log of MAM evolution for each trade.
    Output will be a DataFrame suitable for writing to Excel.
    """
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"])
    df["target_achieved_date"] = pd.to_datetime(df["target_achieved_date"])

    progression_records = []

    for i, row in df.iterrows():
        gap_type = str(row.get("gap_type", "")).lower()
        if "no_gap" in gap_type or pd.isna(row["days to target"]):
            continue

        entry_index = i
        entry_date = row["date"]
        entry_price = row["open"]
        direction = row["trade_direction"]
        exit_date = row["target_achieved_date"]

        # Define trade window: from day after entry to target hit (inclusive)
        trade_window = df[(df["date"] > entry_date)]
        if not trade_window.empty and pd.notna(exit_date):
            trade_window = trade_window[trade_window["date"] <= exit_date]

        if trade_window.empty:
            continue

        max_mam = 0
        for day_offset, (_, day_row) in enumerate(trade_window.iterrows(), start=1):
            current_date = day_row["date"]

            if direction == "long":
                adverse_move = entry_price - day_row["low"]
            elif direction == "short":
                adverse_move = day_row["high"] - entry_price
            else:
                adverse_move = 0

            max_mam = max(max_mam, adverse_move)

            progression_records.append({
                "trade_index": i,
                "date": current_date,
                "entry_date": entry_date,
                "entry_price": entry_price,
                "day_low": day_row["low"],
                "day_high": day_row["high"],
                "direction": direction,
                "running_mam": adverse_move,
                "max_mam_so_far": max_mam,
                "days_active": day_offset
            })

    return pd.DataFrame(progression_records)

def classify_single_candle(row):
    open_ = row['open']
    close = row['close']
    high = row['high']
    low = row['low']

    body = abs(close - open_)
    total_range = high - low
    if total_range == 0:
        return "Ambiguous Candle", 16

    upper_shadow = high - max(open_, close)
    lower_shadow = min(open_, close) - low
    body_ratio = body / total_range
    upper_ratio = upper_shadow / total_range
    lower_ratio = lower_shadow / total_range

    # Main candle types
    if body_ratio < 0.1:
        if upper_ratio > 0.4 and lower_ratio < 0.1:
            return "Gravestone Doji", 5
        elif lower_ratio > 0.4 and upper_ratio < 0.1:
            return "Dragonfly Doji", 4
        else:
            return "Doji", 3
    elif upper_shadow < 0.05 * total_range and lower_shadow < 0.05 * total_range:
        if close > open_:
            return "Bullish Marubozu", 1
        else:
            return "Bearish Marubozu", 2
    elif lower_ratio > 0.6 and upper_ratio < 0.2:
        if close > open_:
            return "Hammer", 7
        else:
            return "Hanging Man", 8
    elif upper_ratio > 0.6 and lower_ratio < 0.2:
        if close < open_:
            return "Shooting Star", 10
        else:
            return "Inverted Hammer", 9
    elif upper_ratio > 0.3 and lower_ratio > 0.3 and body_ratio < 0.4:
        return "Spinning Top", 6
    elif upper_ratio > 0.4 and lower_ratio > 0.4:
        return "High-Wave", 11

    # Secondary "Other" classifications
    if total_range < 0.005 * close:
        return "Pause Candle", 15
    elif body < 0.002 * close and upper_shadow < 0.003 * close and lower_shadow < 0.003 * close:
        return "Short Line Candle", 12
    elif abs(upper_shadow - lower_shadow) < 0.001 * close and body < 0.01 * close:
        return "Neutral Candle", 13
    elif body_ratio > 0.4 and upper_ratio < 0.3 and lower_ratio < 0.3:
        return "Continuation Candle", 14

    return "Ambiguous Candle", 16




def add_candle_classifications_to_excel(file_path, df):
    """
    Adds 'candle_type' and 'candle_code' columns to the 'All Data' sheet in the given Excel file.
    """
    book = load_workbook(file_path)
    if "All Data" not in book.sheetnames:
        print("Error: 'All Data' sheet not found.")
        return

    df_existing = pd.read_excel(file_path, sheet_name="All Data")

    # Align index
    df_existing = df_existing.reset_index(drop=True)
    df = df.reset_index(drop=True)

    # Add or update the columns
    df_existing["candle_type"] = df["candle_type"]
    df_existing["candle_code"] = df["candle_code"]

    # Save back to Excel
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_existing.to_excel(writer, sheet_name="All Data", index=False)

    print("Candle classifications successfully added to cat_.xlsx")

def detect_2_candle_patterns(df):
    patterns = []
    codes = []

    for i in range(len(df) - 1):
        c1 = df.iloc[i]
        c2 = df.iloc[i + 1]

        open1, close1 = c1['open'], c1['close']
        open2, close2 = c2['open'], c2['close']
        low1, high1 = c1['low'], c1['high']
        low2, high2 = c2['low'], c2['high']

        # Bullish Engulfing (loose)
        if close2 > open2 and close2 > open1 and open2 < close1:
            patterns.append("Bullish Engulfing")
            codes.append(1)
            continue

        # Bearish Engulfing (loose)
        if close2 < open2 and open2 > close1 and close2 < open1:
            patterns.append("Bearish Engulfing")
            codes.append(2)
            continue

        # Piercing Pattern (loose)
        if open2 < low1 and close2 > (open1 + close1)/2 and close2 < open1:
            patterns.append("Piercing Pattern")
            codes.append(3)
            continue

        # Dark Cloud Cover (loose)
        if open2 > close1 and close2 < (open1 + close1)/2 and close2 > open1:
            patterns.append("Dark Cloud Cover")
            codes.append(4)
            continue

        # Tweezers (low threshold)
        if abs(low1 - low2) < 0.0015 * low1:
            patterns.append("Tweezer Bottoms")
            codes.append(5)
            continue
        if abs(high1 - high2) < 0.0015 * high1:
            patterns.append("Tweezer Tops")
            codes.append(6)
            continue

        patterns.append(None)
        codes.append(None)

    patterns.append(None)
    codes.append(None)

    df['two_candle_pattern'] = patterns
    df['two_candle_code'] = codes
    return df


def detect_3_candle_patterns(df):
    patterns = []
    codes = []

    for i in range(len(df) - 2):
        c1 = df.iloc[i]
        c2 = df.iloc[i + 1]
        c3 = df.iloc[i + 2]

        open1, close1 = c1['open'], c1['close']
        open2, close2 = c2['open'], c2['close']
        open3, close3 = c3['open'], c3['close']

        body1 = abs(close1 - open1)
        body2 = abs(close2 - open2)

        # No direction check for c3. just midpoint logic
        # Morning Star (loose)
        if close1 < open1 and body2 < body1 * 0.5 and close3 > ((open1 + close1) / 2):
            patterns.append("Morning Star")
            codes.append(1)
            continue

        # Evening Star (loose)
        if close1 > open1 and body2 < body1 * 0.5 and close3 < ((open1 + close1) / 2):
            patterns.append("Evening Star")
            codes.append(2)
            continue

        # Three White Soldiers (loose)
        if close1 < close2 < close3 and open1 < close1 and open2 < close2 and open3 < close3:
            patterns.append("Three White Soldiers")
            codes.append(3)
            continue

        # Three Black Crows (loose)
        if close1 > close2 > close3 and open1 > close1 and open2 > close2 and open3 > close3:
            patterns.append("Three Black Crows")
            codes.append(4)
            continue

        patterns.append(None)
        codes.append(None)

    patterns.append(None)
    patterns.append(None)
    codes.append(None)
    codes.append(None)

    df['three_candle_pattern'] = patterns
    df['three_candle_code'] = codes
    return df



def detect_trend_context(df, ma_period=10):
    """
    Detects short-term trend context based on a moving average.
    Adds two columns:
      - 'ma_{period}' (e.g., ma_10)
      - 'trend_context' ('bullish', 'bearish', 'neutral')

    Parameters:
    - df (pd.DataFrame): Your DataFrame with at least 'close' column.
    - ma_period (int): Moving average period (default = 10 days).
    """

    ma_column = f"ma_{ma_period}"

    # Create 10-day simple moving average
    df[ma_column] = df['close'].rolling(window=ma_period).mean()

    # Define a neutral zone (very close to MA)
    neutral_threshold = 0.0025  # 0.25%

    trend_context = []
    for idx, row in df.iterrows():
        close = row['close']
        ma = row[ma_column]

        if pd.isna(ma):
            trend_context.append(None)
            continue

        if abs(close - ma) / ma <= neutral_threshold:
            trend_context.append("neutral")
        elif close > ma:
            trend_context.append("bullish")
        else:
            trend_context.append("bearish")

    df['trend_context'] = trend_context

    return df


def detect_volume_context(df, ma_period=10):
    """
    Detects volume context relative to a moving average of volume.
    Adds two columns:
      - 'vol_ma_{period}' (e.g., vol_ma_10)
      - 'volume_context' ('high', 'normal', 'low')

    Parameters:
    - df (pd.DataFrame): Your DataFrame with 'volume' column.
    - ma_period (int): Moving average period (default = 10 days).
    """

    vol_ma_column = f"vol_ma_{ma_period}"

    # Calculate 10-day moving average of volume
    df[vol_ma_column] = df['volume'].rolling(window=ma_period).mean()

    volume_context = []

    for idx, row in df.iterrows():
        vol = row['volume']
        vol_ma = row[vol_ma_column]

        if pd.isna(vol_ma):
            volume_context.append(None)
            continue

        if vol > 1.2 * vol_ma:
            volume_context.append("high")
        elif vol < 0.8 * vol_ma:
            volume_context.append("low")
        else:
            volume_context.append("normal")

    df['volume_context'] = volume_context

    return df


def tag_gap_fill_setups_long_and_short(df):
    """
    Tags strong gap fill setups for both longs and shorts based on:
    - gap_type + candlestick patterns + trend + volume
    Adds two columns:
      - 'gap_fill_setup_long' (True/False)
      - 'gap_fill_setup_short' (True/False)
    """

    setup_long_tags = []
    setup_short_tags = []

    # Bullish patterns for long setups (gap down fills)
    bullish_patterns = [
        "Bullish Marubozu", "Hammer", "Dragonfly Doji", "Bullish Engulfing",
        "Morning Star", "Three White Soldiers", "Piercing Pattern", "Tweezer Bottoms"
    ]

    # Bearish patterns for short setups (gap up fades)
    bearish_patterns = [
        "Bearish Marubozu", "Hanging Man", "Gravestone Doji", "Bearish Engulfing",
        "Evening Star", "Three Black Crows", "Dark Cloud Cover", "Tweezer Tops"
    ]

    for idx, row in df.iterrows():
        try:
            ## LONG SETUP CONDITIONS
            gap_down = (row['gap_type'] in ["price_gap_down", "bar_gap_down"])
            pattern_bullish = (
                (row['candle_type'] in bullish_patterns) or
                (row['two_candle_pattern'] in bullish_patterns) or
                (row['three_candle_pattern'] in bullish_patterns)
            )
            trend_ok_long = (row['trend_context'] in ["bullish", "neutral"])
            volume_ok_long = (row['volume_context'] in ["normal", "high"])

            ## SHORT SETUP CONDITIONS
            gap_up = (row['gap_type'] in ["price_gap_up", "bar_gap_up"])
            pattern_bearish = (
                (row['candle_type'] in bearish_patterns) or
                (row['two_candle_pattern'] in bearish_patterns) or
                (row['three_candle_pattern'] in bearish_patterns)
            )
            trend_ok_short = (row['trend_context'] in ["bearish", "neutral"])
            volume_ok_short = (row['volume_context'] in ["normal", "high"])

            ## FINAL TAGGING
            if gap_down and pattern_bullish and trend_ok_long and volume_ok_long:
                setup_long_tags.append(True)
            else:
                setup_long_tags.append(False)

            if gap_up and pattern_bearish and trend_ok_short and volume_ok_short:
                setup_short_tags.append(True)
            else:
                setup_short_tags.append(False)

        except:
            setup_long_tags.append(False)
            setup_short_tags.append(False)

    df['gap_fill_setup_long'] = setup_long_tags
    df['gap_fill_setup_short'] = setup_short_tags

    return df


def tag_gap_fill_and_continuation_setups(df):
    """
    Tags both gap reversion (mean reversion) setups and gap continuation setups.
    Adds four new columns:
      - 'gap_fill_setup_long'
      - 'gap_fill_setup_short'
      - 'gap_continuation_setup_long'
      - 'gap_continuation_setup_short'
    """

    setup_fill_long_tags = []
    setup_fill_short_tags = []
    setup_cont_long_tags = []
    setup_cont_short_tags = []

    # Bullish patterns for LONG gap fill (after gap DOWN)
    bullish_reversal_patterns = [
        "Bullish Marubozu", "Hammer", "Dragonfly Doji", "Bullish Engulfing",
        "Morning Star", "Three White Soldiers", "Piercing Pattern", "Tweezer Bottoms"
    ]

    # Bearish patterns for SHORT gap fill (after gap UP)
    bearish_reversal_patterns = [
        "Bearish Marubozu", "Hanging Man", "Gravestone Doji", "Bearish Engulfing",
        "Evening Star", "Three Black Crows", "Dark Cloud Cover", "Tweezer Tops"
    ]

    # Bullish patterns for CONTINUATION after GAP UP
    bullish_continuation_patterns = [
        "Bullish Marubozu", "Three White Soldiers"
    ]

    # Bearish patterns for CONTINUATION after GAP DOWN
    bearish_continuation_patterns = [
        "Bearish Marubozu", "Three Black Crows"
    ]

    for idx, row in df.iterrows():
        try:
            ### GAP DOWN SCENARIOS ###
            gap_down = (row['gap_type'] in ["price_gap_down", "bar_gap_down"])

            # Reversal Logic (gap fill long)
            bullish_reversal_detected = (
                (row['candle_type'] in bullish_reversal_patterns) or
                (row['two_candle_pattern'] in bullish_reversal_patterns) or
                (row['three_candle_pattern'] in bullish_reversal_patterns)
            )
            trend_ok_long = (row['trend_context'] in ["bullish", "neutral"])
            volume_ok_long = (row['volume_context'] in ["normal", "high"])

            if gap_down and bullish_reversal_detected and trend_ok_long and volume_ok_long:
                setup_fill_long_tags.append(True)
            else:
                setup_fill_long_tags.append(False)

            # Continuation Logic (gap continuation short)
            bearish_continuation_detected = (
                (row['candle_type'] in bearish_continuation_patterns) or
                (row['two_candle_pattern'] in bearish_continuation_patterns) or
                (row['three_candle_pattern'] in bearish_continuation_patterns)
            )
            trend_ok_cont_short = (row['trend_context'] in ["bearish", "neutral"])
            volume_ok_cont_short = (row['volume_context'] in ["normal", "high"])

            if gap_down and bearish_continuation_detected and trend_ok_cont_short and volume_ok_cont_short:
                setup_cont_short_tags.append(True)
            else:
                setup_cont_short_tags.append(False)

            ### GAP UP SCENARIOS ###
            gap_up = (row['gap_type'] in ["price_gap_up", "bar_gap_up"])

            # Reversal Logic (gap fill short)
            bearish_reversal_detected = (
                (row['candle_type'] in bearish_reversal_patterns) or
                (row['two_candle_pattern'] in bearish_reversal_patterns) or
                (row['three_candle_pattern'] in bearish_reversal_patterns)
            )
            trend_ok_short = (row['trend_context'] in ["bearish", "neutral"])
            volume_ok_short = (row['volume_context'] in ["normal", "high"])

            if gap_up and bearish_reversal_detected and trend_ok_short and volume_ok_short:
                setup_fill_short_tags.append(True)
            else:
                setup_fill_short_tags.append(False)

            # Continuation Logic (gap continuation long)
            bullish_continuation_detected = (
                (row['candle_type'] in bullish_continuation_patterns) or
                (row['two_candle_pattern'] in bullish_continuation_patterns) or
                (row['three_candle_pattern'] in bullish_continuation_patterns)
            )
            trend_ok_cont_long = (row['trend_context'] in ["bullish", "neutral"])
            volume_ok_cont_long = (row['volume_context'] in ["normal", "high"])

            if gap_up and bullish_continuation_detected and trend_ok_cont_long and volume_ok_cont_long:
                setup_cont_long_tags.append(True)
            else:
                setup_cont_long_tags.append(False)

        except:
            setup_fill_long_tags.append(False)
            setup_fill_short_tags.append(False)
            setup_cont_long_tags.append(False)
            setup_cont_short_tags.append(False)

    # Assign new columns
    df['gap_fill_setup_long'] = setup_fill_long_tags
    df['gap_fill_setup_short'] = setup_fill_short_tags
    df['gap_continuation_setup_long'] = setup_cont_long_tags
    df['gap_continuation_setup_short'] = setup_cont_short_tags

    return df




def convert_numpy(obj):
    if isinstance(obj, (np.integer, np.int64, np.int32)):
        return int(obj)
    if isinstance(obj, (np.floating, np.float64, np.float32)):
        return float(obj)
    if isinstance(obj, (np.ndarray,)):
        return obj.tolist()
    return obj



from datetime import datetime
import json

def upload_gapfiller_stats(df, cat_key, rolling_stats, grouped_days, grouped_mam, cumulative, conditional):
    latest_row = df.iloc[-1]  # Get last row of your DataFrame

    payload = {
        "date": str(latest_row["date"]),
        "open": float(latest_row["open"]),
        "previous_close": float(latest_row["previous_close"]),
        "previous_date": str(latest_row["previous_date"]),
        "gap_value": float(latest_row["gap value"]),
        "cat_key": cat_key,
        "rolling_stats": rolling_stats,
        "grouped_days_to_target": grouped_days,
        "grouped_mam": grouped_mam,
        "chartdata_cumulative": cumulative,
        "chartdata_conditional": conditional,
        "created_at": datetime.utcnow().isoformat()
    }

    # Convert NumPy types to native types
    payload = json.loads(json.dumps(payload, default=lambda o: o.item() if hasattr(o, 'item') else o))

    res = supabase.table("cat_key_stats").upsert(payload, on_conflict=["cat_key"]).execute()

    print("✅ Uploaded to Supabase:", res)



def upload_gapfiller_stats_v2(rolling_stats_cat_keys, mam_cat_keys, days_to_hit_chart):
    # === Days To Hit ===
    days_rows = []
    for cat_key, values in days_to_hit_chart.items():
        row = {"cat_key": cat_key}
        row.update({f"conditional_day_{i}": values["conditional"][i] for i in range(11)})
        row.update({f"cumulative_day_{i}": values["cumulative"][i] for i in range(11)})
        days_rows.append(row)

    res = requests.post(
        f"{SUPABASE_URL}/rest/v1/gapfiller_days_to_hit?on_conflict=cat_key",
        headers={**HEADERS, "Prefer": "resolution=merge-duplicates"},
        json=days_rows
    )
    if res.status_code >= 300:
        print(f"❌ Failed to upload gapfiller_days_to_hit: {res.status_code} - {res.text}")
    else:
        print("✅ Uploaded gapfiller_days_to_hit")

    # === Rolling Stats ===
    rolling_rows = []
    for cat_key, stats in rolling_stats_cat_keys.items():
        rolling_rows.append({
            "cat_key": cat_key,
            "median_mam": stats["median_mam"],
            "trimmed_mean_mam": stats["trimmed_mean_mam"],
            "std_mam": stats["std_mam"]
        })

    res = requests.post(
        f"{SUPABASE_URL}/rest/v1/gapfiller_rolling_stats?on_conflict=cat_key",
        headers={**HEADERS, "Prefer": "resolution=merge-duplicates"},
        json=rolling_rows
    )
    if res.status_code >= 300:
        print(f"❌ Failed to upload gapfiller_rolling_stats: {res.status_code} - {res.text}")
    else:
        print("✅ Uploaded gapfiller_rolling_stats")

    # === MaM Stats ===
    mam_rows = []
    for cat_key, stats in mam_cat_keys.items():
        mam_rows.append({
            "cat_key": cat_key,
            "mean_mam": stats["mean_mam"],
            "max_mam": stats["max_mam"],
            "min_mam": stats["min_mam"]
        })

    res = requests.post(
        f"{SUPABASE_URL}/rest/v1/gapfiller_mam_stats?on_conflict=cat_key",
        headers={**HEADERS, "Prefer": "resolution=merge-duplicates"},
        json=mam_rows
    )
    if res.status_code >= 300:
        print(f"❌ Failed to upload gapfiller_mam_stats: {res.status_code} - {res.text}")
    else:
        print("✅ Uploaded gapfiller_mam_stats")

    print("✅ Supabase summary upload complete.")

def create_days_to_hit_summary_by_gap_type_structured(file_path):
    df = pd.read_excel(file_path, sheet_name="All Data")
    df["gap_type"] = df["gap_type"].astype(str)
    df = df[df["days to target"].notna()]

    max_day = 10
    summary_rows = []

    for gap_type, group in df.groupby("gap_type"):
        total = len(group)
        counts = [(group["days to target"] == i).sum() for i in range(max_day + 1)]

        cumulative = []
        conditional = []
        cum_hits = 0

        for i in range(max_day + 1):
            day_hits = counts[i]
            cum_hits += day_hits
            cumulative.append(round(100 * cum_hits / total, 1))
            if i == 0:
                conditional.append(round(100 * day_hits / total, 1))
            else:
                remaining = total - sum(counts[:i])
                conditional.append(round(100 * day_hits / remaining, 1) if remaining > 0 else 0.0)

        row = {
            "gap_type": gap_type,
            "total_occurrences": total,
        }

        for i in range(max_day + 1):
            row[f"count_day_{i}"] = counts[i]
            row[f"conditional_day_{i}"] = conditional[i]
            row[f"cumulative_day_{i}"] = cumulative[i]

        summary_rows.append(row)

    result_df = pd.DataFrame(summary_rows)

    # 🔁 Scale % columns to 0–1 for Excel formatting
    for col in result_df.columns:
        if col.startswith("conditional_day_") or col.startswith("cumulative_day_"):
            result_df[col] = result_df[col] / 100.0

    # Write to Excel
    book = load_workbook(file_path)
    sheet_name = "%% Days to Hit by Gap Types"
    if sheet_name in book.sheetnames:
        del book[sheet_name]
    ws = book.create_sheet(title=sheet_name)

    for row in dataframe_to_rows(result_df, index=False, header=True):
        ws.append(row)

    # Apply % formatting in Excel
    percent_cols = [col for col in result_df.columns if col.startswith("conditional_") or col.startswith("cumulative_")]
    header_row = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_name in percent_cols:
            col_idx = header_row.index(col_name)
            row[col_idx].number_format = "0.0%"

    book.save(file_path)
    print(f"✅ Structured days-to-hit by gap_type written to '{sheet_name}'")



from supabase import create_client

SUPABASE_URL = "https://ebbjhrgpsdcvqnpvewlg.supabase.co"
SUPABASE_SERVICE_ROLE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImViYmpocmdwc2RjdnFucHZld2xnIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc1MDI2NjgzOCwiZXhwIjoyMDY1ODQyODM4fQ.UedMkLiUwlbaILvtZG9yDDp1KpStnilhoYhiL_7pexg"

supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)


def show_problem_cells(df):
    """
    Flags and prints rows that contain NaN, inf, or -inf values — which are invalid for JSON/Supabase.
    """
    problem_mask = df.applymap(lambda x: isinstance(x, float) and (pd.isna(x) or x in [np.inf, -np.inf]))
    problem_rows = df[problem_mask.any(axis=1)]

    if not problem_rows.empty:
        print("🚨 Found JSON-breaking values:")
        for idx in problem_rows.index:
            bad_columns = df.columns[problem_mask.loc[idx]]
            print(f"Row {idx}: bad columns -> {list(bad_columns)}")
    else:
        print("✅ No NaN/inf/-inf values detected.")

    return problem_rows

def upload_to_supabase_gaps_tables(df_all_data: pd.DataFrame, df_days_to_hit_by_gap_type: pd.DataFrame):
    import uuid
    import numpy as np
    from tqdm import tqdm

    print("✅ Cleaning df_all_data before Supabase upload")

    # === Full safe conversion: NaN, NaT, inf, -inf → None ===
    df_all_data = df_all_data.replace([float('inf'), float('-inf')], pd.NA)
    df_all_data = df_all_data.where(pd.notna(df_all_data), None)

    # === Convert datetime columns to ISO format ===
    for col in df_all_data.select_dtypes(include=["datetime64[ns]", "datetime64[ns, UTC]"]).columns:
        df_all_data[col] = df_all_data[col].apply(lambda x: x.isoformat() if x else None)

    # === Convert NumPy/NA types to native JSON-safe types ===
    def convert_numpy(obj):
        if isinstance(obj, (np.integer, np.int64, np.int32)):
            return int(obj)
        if isinstance(obj, (np.floating, np.float64, np.float32)):
            return float(obj) if not np.isnan(obj) and not np.isinf(obj) else None
        if isinstance(obj, (np.ndarray,)):
            return obj.tolist()
        return obj

    records = df_all_data.to_dict(orient="records")
    cleaned_records = []

    for record in records:
        clean = {}
        for k, v in record.items():
            clean[k] = convert_numpy(v)
        if 'id' not in clean or not clean['id']:
            clean['id'] = str(uuid.uuid4())
        cleaned_records.append(clean)

    # === Upload in batches ===
    batch_size = 500
    for i in tqdm(range(0, len(cleaned_records), batch_size)):
        batch = cleaned_records[i:i + batch_size]
        res = supabase.table("gaps_all_data").upsert(batch).execute()
        if res.get("status_code", 200) >= 400:
            print(f"❌ Failed to upload batch: {res}")
            raise ValueError("Upload failed.")

    print("✅ gaps_all_data upload complete.")

    # === Upload to gaps_days_to_hit_by_gap_type ===
    print("Uploading to Supabase: gaps_days_to_hit_by_gap_type")
    records2 = df_days_to_hit_by_gap_type.to_dict(orient="records")

    delete_res = supabase.table("gaps_days_to_hit_by_gap_type").delete().neq("gap_type", "").execute()
    if delete_res.get("status_code", 200) >= 400:
        print(f"❌ Failed to clear old data: {delete_res}")

    res = supabase.table("gaps_days_to_hit_by_gap_type").insert(records2).execute()
    if res.get("status_code", 200) >= 400:
        print(f"❌ Failed to upload to gaps_days_to_hit_by_gap_type: {res}")
    else:
        print("✅ gaps_days_to_hit_by_gap_type upload complete.")




load_dotenv()

CORRECTIONS = {
    "2004-04-19": {"open": 113.55, "high": 113.99, "low": 113.27, "close": 113.83, "volume": 28277600},
    "2004-04-27": {"open": 114.23, "high": 115.12, "low": 113.96, "close": 114.30, "volume": 43485500},
    "2004-05-05": {"open": 112.41, "high": 112.96, "low": 112.16, "close": 112.78, "volume": 34405000},
    "2006-03-03": {"open": 128.67, "high": 130.07, "low": 128.65, "close": 128.76, "volume": 73402496},
    "2006-05-18": {"open": 127.35, "high": 127.75, "low": 126.11, "close": 126.21, "volume": 37906304},
    "2006-03-23": {"open": 130.26, "high": 130.39, "low": 129.66, "close": 130.11, "volume": 46704200},
    "2007-04-02": {"open": 142.16, "high": 142.46, "low": 140.89, "close": 142.16, "volume": 79454680},
    "2007-05-25": {"open": 151.49, "high": 152.02, "low": 151.18, "close": 151.69, "volume": 33309621},
    "2007-06-29": {"open": 150.90, "high": 151.65, "low": 149.15, "close": 150.43, "volume": 39701730},
    "2007-07-23": {"open": 154.18, "high": 154.72, "low": 153.30, "close": 153.97, "volume": 21183836},
    "2007-07-26": {"open": 150.19, "high": 150.80, "low": 146.39, "close": 148.02, "volume": 57592490},
    "2007-08-10": {"open": 144.39, "high": 146.50, "low": 143.12, "close": 144.71, "volume": 11018370},
    "2022-01-24": {"open": 432.03, "high": 440.38, "low": 420.76, "close": 439.84, "volume": 52496719},
    "2022-01-26": {"open": 440.72, "high": 444.04, "low": 428.86, "close": 433.38, "volume": 36391088},
    "2022-02-24": {"open": 411.02, "high": 428.76, "low": 410.64, "close": 428.30, "volume": 13942946},
    "2022-05-04": {"open": 417.08, "high": 429.66, "low": 413.71, "close": 429.06, "volume": 44247895},
    "2022-05-05": {"open": 424.55, "high": 425.00, "low": 409.44, "close": 413.81, "volume": 72929106},
    "2022-05-20": {"open": 393.25, "high": 397.03, "low": 380.54, "close": 389.63, "volume": 31432197},
    "2022-10-13": {"open": 349.21, "high": 367.51, "low": 348.11, "close": 365.97, "volume": 47254518},
    "2024-08-01": {"open": 552.57, "high": 554.87, "low": 539.43, "close": 543.01, "volume": 76428732},
    "2024-09-11": {"open": 548.70, "high": 555.36, "low": 539.96, "close": 554.42, "volume": 75248608},
    "2024-12-18": {"open": 603.98, "high": 606.41, "low": 585.89, "close": 586.28, "volume": 38248729},
    "2025-03-03": {"open": 596.18, "high": 597.34, "low": 579.90, "close": 583.77, "volume": 74249199},
    "2025-04-04": {"open": 523.67, "high": 525.87, "low": 505.06, "close": 505.28, "volume": 17965131},
    "2025-04-07": {"open": 489.19, "high": 523.17, "low": 481.80, "close": 504.38, "volume": 56611355},
    "2025-04-08": {"open": 521.86, "high": 524.98, "low": 489.16, "close": 496.48, "volume": 55816581},
    "2025-04-09": {"open": 493.44, "high": 548.62, "low": 493.05, "close": 548.62, "volume": 11867317},
    "2025-04-10": {"open": 532.17, "high": 533.50, "low": 509.32, "close": 524.58, "volume": 52331225},
    "2025-04-11": {"open": 523.01, "high": 536.43, "low": 520.07, "close": 533.94, "volume": 37866334},
    "2025-04-30": {"open": 547.57, "high": 556.52, "low": 541.52, "close": 554.54, "volume": 33101463}
}



def run_gap_analysis(start_date="2003-01-01"):
    ib = IB()

    # === Pull SPY OHLCV ===
    df = get_ib_spy_data(start_date=start_date)

    # === Apply known corrections ===
    df["date"] = pd.to_datetime(df["date"])
    for date_str, values in CORRECTIONS.items():
        mask = df["date"] == pd.to_datetime(date_str)
        for col, val in values.items():
            df.loc[mask, col] = val

    if df is None or df.empty:
        raise ValueError("SPY DataFrame is empty or None.")

    print("✅ SPY data retrieved from IB")

    # === Compute gap data ===
    df = compute_gap_data(df)
    df["date"] = pd.to_datetime(df["date"]).dt.normalize()
    df["target_achieved_date"] = pd.to_datetime(df["target_achieved_date"]).dt.normalize()
    df = calculate_statistics(df)

    # === Generate additional summary sheets ===
    output_path = r"C:\2mdt\2mindt-site\public\spy gap analysis.xlsx"
    gap_types_summary = create_days_to_target_summary_table_gap_types(output_path)
    rolling_stats = compute_rolling_stats_by_gap_type(output_path)

    # === Save everything with formatting ===
    save_multiple_sheets_with_formatting(
        {
            "All Data": df,
            "%% Days to Target by Gap Type": gap_types_summary,
            "Rolling Stats by Gap Type": rolling_stats
        },
        output_path
    )

    ib.disconnect()



def is_market_open():
    import datetime
    now = datetime.datetime.now()
    is_weekday = now.weekday() < 5
    is_market_time = now.hour == 6 and now.minute >= 30 or (7 <= now.hour < 13)
    return is_weekday and is_market_time


###=============================================================================
###=============================================================================
###=============================================================================
# === Example Call ===
if __name__ == "__main__":
    run_gap_analysis()
###=============================================================================
###=============================================================================
###=============================================================================