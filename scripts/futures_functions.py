# === futures_functions.py ===
from ib_insync import IB, Future
from datetime import datetime
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
# --- CONFIG IMPORTS ---
from config import TICKER_MAP, CONTRACT_MAP, ETF_TO_FUTURES
from dropbox_utils import upload_file
from config import get_dropbox_path


# ---------- Helpers ----------

def _contracts_with_expiry(symbol: str):
    """Return sorted [(expiry, contract)] list of all CME futures (expired included)."""
    ib = IB()
    cid = abs(hash(("CONTRACTS", symbol))) % 9000
    print(f"[connect] IBKR (clientId={cid}) for {symbol}")
    ib.connect("127.0.0.1", 7496, clientId=cid)

    cds = ib.reqContractDetails(Future(symbol, exchange="CME", includeExpired=True))
    ib.disconnect()
    if not cds:
        raise RuntimeError(f"No contracts found for {symbol}")

    out = []
    for cd in cds:
        exp = pd.to_datetime(cd.contract.lastTradeDateOrContractMonth, errors="coerce")
        if pd.notna(exp):
            out.append((exp, cd.contract))
    out.sort(key=lambda x: x[0])
    print(f"[contracts] {len(out)} total for {symbol}")
    return out


def _detect_front_previous(contracts):
    """Identify front and previous based on today's date."""
    today = datetime.now()
    front, prev = None, None
    for i, (exp, c) in enumerate(contracts):
        if exp > today:
            front = c
            prev = contracts[i - 1][1] if i > 0 else None
            break
    if front is None and contracts:
        front = contracts[-1][1]
        prev = contracts[-2][1] if len(contracts) >= 2 else None
    print(f"[detect] Today={today.date()} → Front={front.localSymbol if front else 'None'} | "
          f"Previous={prev.localSymbol if prev else 'None'}")
    return front, prev



def _fetch_30m_bars(contract, days=30):
    """Fetch 30-minute bars for given contract."""
    ib = IB()
    cid = abs(hash(("BARS", contract.localSymbol))) % 9000
    ib.connect("127.0.0.1", 7496, clientId=cid)
    bars = ib.reqHistoricalData(
        contract,
        endDateTime="",
        durationStr=f"{days} D",
        barSizeSetting="30 mins",
        whatToShow="TRADES",
        useRTH=False,
        formatDate=1,
    )
    ib.disconnect()

    df = pd.DataFrame(bars)
    if df.empty:
        return df

    # Keep full datetime, strip tz only
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    try:
        df["date"] = df["date"].dt.tz_localize(None)
    except Exception:
        pass

    return df


# def _fetch_30m_bars(contract, days=30):
#     """Fetch 30-minute bars for given contract."""
#     ib = IB()
#     cid = abs(hash(("BARS", contract.localSymbol))) % 9000
#     ib.connect("127.0.0.1", 7496, clientId=cid)
#     bars = ib.reqHistoricalData(
#         contract,
#         endDateTime="",
#         durationStr=f"{days} D",
#         barSizeSetting="30 mins",
#         whatToShow="TRADES",
#         useRTH=False,
#         formatDate=1,
#     )
#     ib.disconnect()
#
#     df = pd.DataFrame(bars)
#     if df.empty:
#         return df
#     df["date"] = pd.to_datetime(df["date"], errors="coerce")
#     try:
#         df["date"] = df["date"].dt.tz_localize(None)
#     except Exception:
#         pass
#     return df


def _strip_tz(df):
    """Remove tzinfo from all datetime columns."""
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            try:
                out[col] = pd.to_datetime(out[col], errors="coerce").dt.tz_localize(None)
            except Exception:
                pass
    return out


def format_excel_workbook(path: str):
    """Freeze header row, bold header text, and auto-fit column widths for all sheets."""
    try:
        wb = load_workbook(path)
        for ws in wb.worksheets:
            # Freeze header
            ws.freeze_panes = "A2"

            # Bold header row
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Auto-fit column widths
            for col_cells in ws.columns:
                vals = [str(c.value) if c.value else "" for c in col_cells]
                max_len = max((len(v) for v in vals), default=0)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)

        wb.save(path)
    except Exception as e:
        print(f"[autosize] skipped: {e}")


def run_futures_volume(symbol="ES", parent_ticker=None):

    """
    Pull all futures contracts from previous→front→future until first 0-volume contract.
    Produce a timeband-style Excel with per-contract volume and barCount columns,
    total aggregates, and 20-day averages. Upload to Dropbox.
    """
    contracts = _contracts_with_expiry(symbol)
    front, prev = _detect_front_previous(contracts)
    if not front:
        print("[error] could not detect front contract")
        return None

    # --- Restrict to previous → end ---
    valid_contracts, started = [], False
    for exp, c in contracts:
        if prev and c.localSymbol == prev.localSymbol:
            started = True
        if started:
            valid_contracts.append((exp, c))
    if not started:
        valid_contracts = [(exp, c) for exp, c in contracts if exp >= datetime.now()]

    # --- Fetch sequentially until 0-volume ---
    tables = []
    for exp, c in valid_contracts:
        df = _fetch_30m_bars(c, 30)
        if df.empty or df["volume"].sum() == 0:
            print(f"[volume] {c.localSymbol}: 0 volume → stop")
            break

        role = (
            "previous" if prev and c.localSymbol == prev.localSymbol
            else "front" if front and c.localSymbol == front.localSymbol
            else "future"
        )

        # Add expiry first
        df["expiry"] = pd.to_datetime(exp, errors="coerce").date()
        df["contract"] = c.localSymbol
        df["role"] = role

        tables.append(df)
        print(
            f"[fetched] {c.localSymbol} ({role}): {len(df)} rows, vol={df['volume'].sum():,.0f}")

    if not tables:
        print("[result] No valid contracts with volume.")
        return None

    # --- Combine and clean ---
    all_df = pd.concat(tables, ignore_index=True)
    all_df = _strip_tz(all_df)

    # --- Create 30-min band grid ---
    all_df["timestamp"] = pd.to_datetime(all_df["date"], errors="coerce")
    all_df["band"] = all_df["timestamp"].dt.strftime("%H:%M") + "–" + (
        all_df["timestamp"] + pd.Timedelta(minutes=30)
    ).dt.strftime("%H:%M")
    all_df["date_only"] = all_df["timestamp"].dt.date

    # --- Base subset ---
    cols = [
        "date_only", "band", "timestamp", "contract",
        "volume", "barCount", "average"
    ]

    df_base = all_df[cols].copy()

    # --- Pivot both volume and barCount by contract ---
    # --- Pivot both volume and barCount by contract (no OHLC) ---
    pivot = (
        df_base.pivot_table(
            index=["date_only", "band", "timestamp"],
            columns="contract",
            values=["volume", "barCount"],
            aggfunc="sum",
        )
        .reset_index()
        .fillna(0)
    )
    pivot.rename(columns={"date_only": "date"}, inplace=True)

    # --- Flatten multi-index columns like ('volume','esu5') → 'esu5_volume' ---
    pivot.columns = [
        f"{str(c[1]).lower()}_{str(c[0]).lower()}" if c[1] else str(c[0]).lower()
        for c in pivot.columns
    ]

    # --- Dynamic rename using actual contract names ---
    pivot.columns = [str(c).lower().replace(" ", "_") for c in pivot.columns]
    data_cols = [c for c in pivot.columns if c not in [
        "date_only", "band", "timestamp", "open", "high", "low", "close", "average"
    ]]

    for col in data_cols:
        clean = col.replace(".", "_").replace("-", "_").replace("/", "_")
        pivot.rename(columns={col: clean}, inplace=True)

    # --- Compute totals ---
    vol_cols = [c for c in pivot.columns if c.endswith("_volume")]
    bar_cols = [c for c in pivot.columns if c.endswith("_barcount")]
    pivot["volume"] = pivot[vol_cols].sum(axis=1)
    pivot["barCount"] = pivot[bar_cols].sum(axis=1)

    # --- Add session + 20-day metrics ---
    pivot["granularity"] = "30min"
    pivot["generated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    pivot["session"] = pivot["timestamp"].dt.time.between(
        pd.to_datetime("09:30").time(), pd.to_datetime("16:00").time()
    ).map({True: "RTH", False: "ETH"})
    pivot["avg_20d"] = pivot.groupby("band")["volume"].transform(
        lambda x: x.rolling(20, min_periods=1).mean()
    )
    pivot["ratio_to_avg_20d"] = pivot["volume"] / pivot["avg_20d"]

    # --- Determine chronological contract order using expiry dates ---
    contract_exp_map = {
        c.localSymbol.lower(): pd.to_datetime(c.lastTradeDateOrContractMonth, errors="coerce")
        for _, c in valid_contracts
    }
    ordered_contracts = sorted(contract_exp_map.items(), key=lambda x: x[1] or pd.Timestamp.max)
    ordered_syms = []
    for sym, _ in ordered_contracts:
        sym_l = sym.lower()
        if any(col.startswith(sym_l) for col in pivot.columns):
            ordered_syms.append(sym_l)

    ordered_vol_cols = [f"{sym}_volume" for sym in ordered_syms if f"{sym}_volume" in pivot.columns]
    ordered_bar_cols = [f"{sym}_barcount" for sym in ordered_syms if f"{sym}_barcount" in pivot.columns]

    # --- Final column order ---
    ordered_cols = (
            ["date", "timestamp", "generated_at", "granularity", "session",
             "band"]
            + ordered_vol_cols
            + ["volume"]  # rename later as Total_volume
            + ordered_bar_cols
            + ["barCount", "avg_20d", "ratio_to_avg_20d"]
    )

    pivot = pivot[[c for c in ordered_cols if c in pivot.columns]]

    # --- Rename totals for clarity ---
    pivot.rename(columns={
        "volume": "Total_volume",
        "barCount": "Total_barCount"
    }, inplace=True)

    # --- Write Excel with all sheets ---
    ts = datetime.now().strftime("%Y-%m-%d_%H%M")
    fname = f"{symbol.lower()}_timeband_volume.xlsx"
    print(f"[save] Writing Excel → {fname}")

    with pd.ExcelWriter(fname, engine="openpyxl", mode="w") as w:
        # Main Timebands sheet
        pivot.to_excel(w, sheet_name="Timebands", index=False)

        # Individual contract sheets
        for exp, c in valid_contracts:
            name = c.localSymbol[:31]  # Excel sheet name limit
            sub = all_df[all_df["contract"] == c.localSymbol].copy()
            if sub.empty:
                continue
            sub = _strip_tz(sub)
            keep = [
                "date", "band", "open", "high", "low", "close",
                "volume", "barCount", "average", "expiry", "role"
            ]
            sub = sub[[col for col in keep if col in sub.columns]]
            sub.sort_values("date", inplace=True)
            sub["date"] = pd.to_datetime(sub["date"], errors="coerce").dt.date
            sub.to_excel(w, sheet_name=name, index=False)

    # --- Format workbook ---
    format_excel_workbook(fname)
    abs_path = os.path.abspath(fname)

    # --- Upload to Dropbox ---
    parent = (parent_ticker or symbol).lower()  # "spy" / "qqq" ...
    child = f"{symbol.lower()}-timebands-volume"  # e.g., "es-timebands-volume"
    dropbox_path = f"/{parent}/{child}/{os.path.basename(fname)}"
    upload_file(abs_path, dropbox_path)
    print(f"[Dropbox] Uploaded {symbol} → {dropbox_path}")

    # --- Clean up local ---
    try:
        os.remove(abs_path)
        print(f"[Local] Deleted temporary file: {abs_path}")
    except Exception as e:
        print(f"[Warning] Could not delete local file: {e}")

    return dropbox_path




# ---------- Multi-ETF Runner ----------

def run_all_etf_futures_volume():
    """
    Loop through all tickers in TICKER_MAP.
    For each ticker, find mapped futures in ETF_TO_FUTURES,
    and produce one Excel file per futures root.
    """
    for ticker in TICKER_MAP.keys():
        fut_syms = ETF_TO_FUTURES.get(ticker, [])
        if not fut_syms:
            print(f"[skip] {ticker}: no futures mapping found.")
            continue

        print(f"\n[map] {ticker} → {', '.join(fut_syms)}")
        for fut in fut_syms:
            try:
                print(f"[run] Fetching futures volume for {fut} (from {ticker})")
                path = run_futures_volume(fut, parent_ticker=ticker)

                if path:
                    print(f"[ok] {fut}: saved → {path}")
                else:
                    print(f"[warn] {fut}: no valid data.")
            except Exception as e:
                print(f"[error] {fut}: {e}")


# ---------- CLI ----------

if __name__ == "__main__":
    # To run all ETFs → futures
    run_all_etf_futures_volume()
