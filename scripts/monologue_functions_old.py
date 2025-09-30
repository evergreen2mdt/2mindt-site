# === Standard Library ===

import os

import tqdm
# === Third-Party Libraries ===
import pandas as pd
from openpyxl import load_workbook

from openpyxl.utils import get_column_letter
from copy import copy
# === Project Utilities ===
from options_functions import (
    format_sheet
)
from tqdm import tqdm
from options_functions import (OPTIONS_COLUMNS,
    GREEKS_COLUMNS,
    MONTE_CARLO_COLUMNS,
    TOUCH_PROBS_COLUMNS)
from options_functions import format_probabilities_in_workbook
###=============================================================================
###=======================Monologue Compare Functions===========================
###=============================================================================


def write_and_format_comparison_sheet(df, path, sheetname):

    print("formatting wkbks")

    os.makedirs(os.path.dirname(path), exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheetname)

    book = load_workbook(path)
    ws = book[sheetname]

    # Bold headers
    for cell in ws[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font

    # Freeze panes
    ws.freeze_panes = "A2"
    print("pane frozen")
    # Auto-adjust column widths
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    print("header widened")
    # Format percentages
    for col_idx, cell in enumerate(ws[1], 1):
        header = str(cell.value).lower()
        if any(kw in header for kw in ["prob", "touch", "ratio", "adj_factor", "cumulative"]) and "day" not in header:
            col_letter = get_column_letter(col_idx)
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                if isinstance(cell.value, (float, int)):
                    cell.number_format = '0.00%'
    print("percentages formatted")
    book.save(path)







# === STEP 1: Get Last 10 Days of Files ===
def get_last_10_days_files(folder, file_prefix):
    print("Scanning for recent files...", flush=True)
    files = [f for f in os.listdir(folder) if f.startswith(file_prefix) and f.endswith(".xlsx")]
    dated_files = []
    for f in tqdm(files, desc="Parsing files"):
        try:
            date_str = f[len(file_prefix):].split("_")[0]
            dt = pd.to_datetime(date_str)
            dated_files.append((dt.date(), os.path.join(folder, f)))
        except:
            continue
    dated_files.sort(reverse=True)
    grouped = {}
    for dt, path in tqdm(dated_files, desc="Grouping by date"):
        if dt not in grouped:
            grouped[dt] = []
        grouped[dt].append(path)
    print(f"Found files for {len(grouped)} recent days.", flush=True)
    return grouped

def load_strategy_sheet_from_excel(path):
    if os.path.exists(path):
        return pd.read_excel(path)
    return pd.DataFrame()


def load_sheets(filepath):
    try:
        print(f"    Loading sheets from {filepath}", flush=True)
        options = pd.read_excel(filepath, sheet_name="options and pinning")
        greeks = pd.read_excel(filepath, sheet_name="greeks")
        mc = pd.read_excel(filepath, sheet_name="monte carlo")
        touch_probs = pd.read_excel(filepath, sheet_name="touch probs")
        volume = pd.read_excel(filepath, sheet_name="etf volume")

        return options, greeks, mc, touch_probs, volume

    except Exception as e:
        print(f"    Failed to load sheets from {filepath}: {e}", flush=True)
        return None, None, None, None, None



# === STEP 3: Compare Snapshots ===

def compare_strikes(df_now, df_prev):
    print("    Comparing strikes...", flush=True)
    try:
        merged = df_now.merge(
            df_prev,
            on=["strike"],
            how="inner",
            suffixes=("_now", "_prev")
        )
    except Exception as e:
        print(f"    Merge failed: {e}", flush=True)
        return []

    significant = []
    for row in tqdm(merged.itertuples(), desc="    Evaluating merged rows"):
        strike = row.strike
        dte = getattr(row, "dte_now", 0)
        touch_now = getattr(row, "cumulative_mc_touch_now", 0)
        touch_prev = getattr(row, "cumulative_mc_touch_prev", 0)
        delta_prob = touch_now - touch_prev
        net_gamma_now = getattr(row, "net_gamma_exposure_now", 0)
        net_gamma_prev = getattr(row, "net_gamma_exposure_prev", 0)

        if abs(delta_prob) >= 0.05 or (net_gamma_now * net_gamma_prev < 0):
            significant.append({
                "strike": strike,
                "type": getattr(row, "type_now", "?"),
                "expiration": getattr(row, "expiration_now", "?"),  # ✅ Included for strategy key
                "dte": dte,
                "delta_adj_prob": delta_prob,
                "net_gamma": net_gamma_now,
                "pinning_strength": getattr(row, "pinning_strength_now", 0),
                "vega_crush": bool(getattr(row, "vega_crush_zone_now", False)),
                "theta_gravity": bool(getattr(row, "theta_gravity_now", False)),
                "gamma_regime": getattr(row, "gamma_regime_now", None),
                "spot_now": getattr(row, "spot_now", 0),
                "spot_prev": getattr(row, "spot_prev", 0)
            })
    print(f"    Found {len(significant)} significant changes.", flush=True)
    return significant


def update_strategies(date, changes, strategy_memory):
    today = str(date)
    print("Updating strategies...", flush=True)
    for ch in tqdm(changes, desc="Evaluating strategy candidates"):
        sid = f"{int(ch['strike'])}{ch['type']}_{ch['expiration']}_fade"

        if sid not in strategy_memory:
            if ch['delta_adj_prob'] > 0.1 and ch['net_gamma'] < 0 and ch['pinning_strength'] > 10000:
                strategy_memory[sid] = {
                    "created_on": today,
                    "strike": ch['strike'],
                    "type": ch['type'],
                    "status": "ongoing",
                    "days_tracked": 1,
                    "thesis": f"Fade {ch['strike']} on rising prob, short gamma, and pinning",
                    "last_note": f"Initialized on {today} with prob Δ {ch['delta_adj_prob']:.2%}"
                }
        elif strategy_memory[sid]["status"] == "ongoing":
            strategy = strategy_memory[sid]
            strategy["days_tracked"] += 1
            strategy["last_note"] = f"Day {strategy['days_tracked']} - prob Δ {ch['delta_adj_prob']:.2%}, gamma {ch['net_gamma']:.0f}"
            if strategy["days_tracked"] >= 10:
                strategy["status"] = "expired"
            elif ch['delta_adj_prob'] < -0.05:
                strategy["status"] = "failed"
            elif ch['delta_adj_prob'] >= 0.15:
                strategy["status"] = "successful"
    return strategy_memory





# === STEP 4: Write Log ===
def write_day_log(date, changes, strategies, spot_changes, gamma_summary, top_pins, v_now, gap_now, log_dir):

    lines = [f"===== {date} ====="]

    lines.append("\n[SPOT MOVEMENT]")
    for entry in spot_changes:
        lines.append(f"{entry['timestamp']}: SPY moved from {entry['spot_prev']:.2f} to {entry['spot_now']:.2f}")

    lines.append("\n[PINNING OVERVIEW]")
    for pin in top_pins:
        lines.append(f"{pin['timestamp']}: Top pin {int(pin['strike'])} with strength {pin['strength']:.0f}")

    lines.append("\n[GAMMA SUMMARY]")
    lines.extend(gamma_summary)

    lines.append("\n[ETF VOLUME]")
    try:
        vol = int(v_now["volume"].iloc[-1])
        ts = str(v_now["timestamp"].iloc[-1])
        lines.append(f"{ts}: 10-min volume = {vol:,}")
    except:
        lines.append("Volume data unavailable.")

    lines.append("\n[GAP SUPPORT]")
    try:
        support = gap_now["support"].iloc[0]
        trade_type = gap_now["trade_type"].iloc[0]
        best_strike = gap_now["best_strike"].iloc[0]
        reason = gap_now["reason"].iloc[0]
        lines.append(f"Trade Type: {trade_type}, Support: {support}, Best Strike: {best_strike}, Reason: {reason}")
    except:
        lines.append("Gap support data unavailable.")

    if strategies is not None and not strategies.empty:
        lines.append("\n[STRATEGY STATUS]")
        for _, strat in strategies.iterrows():
            sid = strat.get("id", "?")
            status = strat.get("status", "unknown")
            thesis = strat.get("thesis", "")
            last_note = strat.get("last_note", "")
            days_tracked = strat.get("days_tracked", "")

            if strat.get("created_on") == str(date):
                lines.append(f"+ New Strategy: {sid}: {thesis}")
            elif status == "ongoing":
                lines.append(
                    f"= Ongoing: {sid}: Days tracked {days_tracked} - {last_note}")
            else:
                lines.append(f"x Closed: {sid}: {status} - {last_note}")

    if changes:
        lines.append("\n[KEY CHANGES]")
        for change in changes:
            lines.append(
                f"- {change['type']} {change['strike']} ({int(change['dte'])} DTE): "
                f"prob Δ {change['delta_adj_prob']:.2%}, "
                f"gamma {change['net_gamma']:.0f}, pin {change['pinning_strength']:.0f}, "
                f"vega_crush={change['vega_crush']}, theta_gravity={change['theta_gravity']}, "
                f"regime={change['gamma_regime']}"
            )

    outpath = os.path.join(log_dir, f"monologue_{date}.txt")

    with open(outpath, "w", encoding="utf-8") as f:

        f.write("\n".join(lines))
    print(f"Log written to {outpath}", flush=True)



def save_strategy_sheet_to_excel(df, path):
    df.to_excel(path, index=False)



def extract_timestamp(filepath):
    """
    Extracts '2025-07-31_thu_12-48' from filenames like
    'spy_options_data_2025-07-31_thu_12-48.xlsx'
    and prevents 'montecarlo_2025-07-31_thu_12-48'
    """
    name = os.path.basename(filepath).replace(".xlsx", "")
    parts = name.split("_")

    for i in range(len(parts) - 3):
        if parts[i].count("-") == 2 and parts[i+1].isalpha() and "-" in parts[i+2]:
            timestamp = "_".join(parts[i:i+3])
            if timestamp.count("-") == 3:  # sanity check for correct format
                return timestamp

    # Final fallback: scan for last 3 parts and discard anything with 'montecarlo'
    last_parts = parts[-3:]
    last_clean = [p for p in last_parts if not p.startswith("montecarlo")]
    return "_".join(last_clean)

# === Options & Pinning ===
def build_wideform_options_and_pinning_comparison(folder):
    sheet_to_read = "options and pinning"
    sheet_to_write = "options_and_pinning_comparison"
    key_cols = ["strike", "type", "expiration"]
    files = sorted([
        os.path.join(folder, f) for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$") and "spy_options_data" in f
    ])

    wideform = None
    for filepath in tqdm(files, desc="Building Options & Pinning wideform"):
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_to_read)
            df = df[[col for col in OPTIONS_COLUMNS if col in df.columns]]
            timestamp = extract_timestamp(filepath)
            rename_map = {col: f"{col}_{timestamp}" for col in df.columns if col not in key_cols}
            selected = df[key_cols + list(rename_map.keys())].rename(columns=rename_map)
            selected["strike"] = selected["strike"].round().astype(int)
            if wideform is not None:
                wideform["strike"] = wideform["strike"].round().astype(int)
            wideform = selected if wideform is None else wideform.merge(selected, on=key_cols, how="outer")
        except Exception as e:
            print(f"[Options & Pinning] Failed to process {filepath}: {e}")

    if wideform is None:
        print("[Options & Pinning] No valid data.")
        return

    ordered_cols = key_cols[:]
    timestamps = sorted(set(col.split("_")[-1] for col in wideform.columns if col.startswith("bid_")))
    for base in OPTIONS_COLUMNS:
        for ts in timestamps:
            col_name = f"{base}_{ts}"
            if col_name in wideform.columns:
                ordered_cols.append(col_name)
    leftovers = [c for c in wideform.columns if c not in ordered_cols]
    wideform = wideform[ordered_cols + leftovers]

    path = os.path.join(folder, "comparison", f"{sheet_to_write}.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wideform.to_excel(writer, index=False, sheet_name=sheet_to_write)

    print(f"[Options & Pinning] Written to: {path}")


# === Monte Carlo ===
def build_wideform_monte_carlo_comparison(folder):
    sheet_to_read = "monte carlo"
    sheet_to_write = "monte_carlo_comparison"
    key_cols = ["strike"]
    files = sorted([
        os.path.join(folder, f) for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$") and "spy_options_data" in f
    ])
    wideform = None

    for filepath in tqdm(files, desc="Building Monte Carlo wideform"):
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_to_read)
            df = df[[col for col in MONTE_CARLO_COLUMNS if col in df.columns]]
            timestamp = extract_timestamp(filepath)
            rename_map = {col: f"{col}_{timestamp}" for col in df.columns if col not in key_cols}
            selected = df[key_cols + list(rename_map.keys())].rename(columns=rename_map)
            selected["strike"] = selected["strike"].round().astype(int)
            if wideform is not None:
                wideform["strike"] = wideform["strike"].round().astype(int)
            wideform = selected if wideform is None else wideform.merge(selected, on=key_cols, how="outer")
        except Exception as e:
            print(f"[Monte Carlo] Failed to process {filepath}: {e}")

    if wideform is None:
        print("[Monte Carlo] No valid data.")
        return

    ordered_cols = key_cols[:]
    timestamps = sorted(set(col.split("_")[-1] for col in wideform.columns if col.startswith("cumulative_prob_")))
    for base in MONTE_CARLO_COLUMNS:
        for ts in timestamps:
            col_name = f"{base}_{ts}"
            if col_name in wideform.columns:
                ordered_cols.append(col_name)
    leftovers = [c for c in wideform.columns if c not in ordered_cols]
    wideform = wideform[ordered_cols + leftovers]

    path = os.path.join(folder, "comparison", f"{sheet_to_write}.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wideform.to_excel(writer, index=False, sheet_name=sheet_to_write)

    print(f"[Monte Carlo] Written to: {path}")

# === Greeks ===
def build_wideform_greeks_comparison(folder):
    sheet_to_read = "greeks"
    sheet_to_write = "greeks_comparison"
    key_cols = ["strike", "type", "dte"]
    files = sorted([
        os.path.join(folder, f) for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$") and "spy_options_data" in f
    ])
    wideform = None

    for filepath in tqdm(files, desc="Building Greeks wideform"):
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_to_read)
            greeks_cols = [col for col in GREEKS_COLUMNS if col != "expiration" and col in df.columns]
            df = df[greeks_cols].dropna(subset=key_cols).drop_duplicates(subset=key_cols)
            df["strike"] = df["strike"].round().astype(int)
            df_full = pd.DataFrame([(s, t, d) for s in df["strike"].unique() for t in df["type"].unique() for d in range(31)], columns=key_cols)
            df = df_full.merge(df, on=key_cols, how="left")
            timestamp = extract_timestamp(filepath)
            rename_map = {col: f"{col}_{timestamp}" for col in df.columns if col not in key_cols}
            selected = df[key_cols + list(rename_map.keys())].rename(columns=rename_map)
            wideform = selected if wideform is None else wideform.merge(selected, on=key_cols, how="outer")
        except Exception as e:
            print(f"[Greeks] Failed to process {filepath}: {e}")

    if wideform is None:
        print("[Greeks] No valid data.")
        return

    ordered_cols = key_cols[:]
    timestamps = sorted(set(col.split("_")[-1] for col in wideform.columns if col.startswith("delta_exposure_")))
    for base in GREEKS_COLUMNS:
        if base in key_cols or base == "expiration":
            continue
        for ts in timestamps:
            col_name = f"{base}_{ts}"
            if col_name in wideform.columns:
                ordered_cols.append(col_name)
    leftovers = [c for c in wideform.columns if c not in ordered_cols]
    wideform = wideform[ordered_cols + leftovers]

    path = os.path.join(folder, "comparison", f"{sheet_to_write}.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wideform.to_excel(writer, index=False, sheet_name=sheet_to_write)

    print(f"[Greeks] Written to: {path}")


# === Touch Probabilities ===
def build_wideform_touch_probs_comparison(folder):
    sheet_to_read = "touch probs"
    sheet_to_write = "touch_probs_comparison"
    key_cols = ["strike", "type", "expiration"]
    files = sorted([
        os.path.join(folder, f) for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$") and "spy_options_data" in f
    ])
    wideform = None

    for filepath in tqdm(files, desc="Building Touch Probs wideform"):
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_to_read)
            df = df[[col for col in TOUCH_PROBS_COLUMNS if col in df.columns]]
            df["strike"] = df["strike"].round().astype(int)
            timestamp = extract_timestamp(filepath)
            rename_map = {col: f"{col}_{timestamp}" for col in df.columns if col not in key_cols}
            selected = df[key_cols + list(rename_map.keys())].rename(columns=rename_map)
            wideform = selected if wideform is None else wideform.merge(selected, on=key_cols, how="outer")
        except Exception as e:
            print(f"[Touch Probs] Failed to process {filepath}: {e}")

    if wideform is None:
        print("[Touch Probs] No valid data.")
        return

    ordered_cols = key_cols[:]
    for base in TOUCH_PROBS_COLUMNS:
        group = sorted([col for col in wideform.columns if col.startswith(base + "_")])
        ordered_cols.extend(group)
    leftovers = [c for c in wideform.columns if c not in ordered_cols]
    wideform = wideform[ordered_cols + leftovers]

    path = os.path.join(folder, "comparison", f"{sheet_to_write}.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wideform.to_excel(writer, index=False, sheet_name=sheet_to_write)

    print(f"[Touch Probs] Written to: {path}")




# === ETF Volume ===
def build_wideform_etf_volume_comparison(folder):
    sheet_to_read = "etf volume"
    sheet_to_write = "etf_volume_comparison"
    output_path = os.path.join(folder, "comparison", f"{sheet_to_write}.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    files = sorted([
        os.path.join(folder, f) for f in os.listdir(folder)
        if f.startswith("spy_options_data_") and f.endswith(".xlsx")
    ])

    all_rows = []
    for file in files:
        try:
            df = pd.read_excel(file, sheet_name=sheet_to_read)
            df = df[["timestamp", "volume", "spot", "generated_at"]]
            all_rows.append(df)
        except Exception as e:
            print(f"[ETF Volume] Skipping {file}: {e}")

    if all_rows:
        combined = pd.concat(all_rows, ignore_index=True).drop_duplicates().sort_values("timestamp")
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            combined.to_excel(writer, index=False, sheet_name=sheet_to_write)

        print(f"[ETF Volume] Written to: {output_path}")
    else:
        print("[ETF Volume] No valid data to write.")




def sort_all_comparison_sheets(comparison_dir):
    files_and_sheets = {
        "options_and_pinning_comparison": "options_and_pinning_comparison",
        "monte_carlo_comparison": "monte_carlo_comparison",
        "greeks_comparison": "greeks_comparison",
        "touch_probs_comparison": "touch_probs_comparison",
        "etf_volume_comparison": "etf_volume_comparison",
    }

    for base_name, sheet_name in files_and_sheets.items():
        filepath = os.path.join(comparison_dir, f"{base_name}.xlsx")

        if not os.path.exists(filepath):
            print(f"No file found: {filepath}")
            continue

        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name)

            # Determine sorting keys based on sheet type
            if "etf_volume" in base_name:
                sort_keys = ["timestamp"] if "timestamp" in df.columns else df.columns[:1].tolist()
            else:
                sort_keys = [col for col in ["strike", "type", "expiration"] if col in df.columns]

            df = df.sort_values(by=sort_keys)
            df.to_excel(filepath, index=False, sheet_name=sheet_name)
            print(f"Sorted: {filepath}")

        except Exception as e:
            print(f"Failed sorting for {filepath}: {e}")



import re

def reorder_comparison_columns(df: pd.DataFrame, key_cols: list) -> pd.DataFrame:
    """Reorders columns to show key_cols first, then pair all metric columns by base name + suffix."""
    all_cols = list(df.columns)

    # Find all column suffixes via pattern matching
    pattern = re.compile(r"(.+)_([\d]{4}-[\d]{2}-[\d]{2}_[a-z]{3}_[\d\-]+)$")
    suffix_map = {}  # base_name → [full column names]

    for col in all_cols:
        match = pattern.match(col)
        if match:
            base, suffix = match.groups()
            suffix_map.setdefault(base, []).append(col)

    # Determine logical order
    ordered_cols = key_cols.copy()
    for base, cols in suffix_map.items():
        ordered_cols.extend(sorted(cols))  # sort by suffix (timestamp)

    # Add any leftover columns not included
    remaining = [col for col in all_cols if col not in ordered_cols]
    final_cols = ordered_cols + remaining

    return df[final_cols]

#
# def build_and_format_all_comparison_sheets(folder):
#     build_wideform_options_and_pinning_comparison(folder)
#     build_wideform_monte_carlo_comparison(folder)
#     build_wideform_greeks_comparison(folder)
#     build_wideform_touch_probs_comparison(folder)
#     build_wideform_etf_volume_comparison(folder)
#
#     comparison_dir = os.path.join(folder, "comparison")
#
#     files = [
#         "options_and_pinning_comparison.xlsx",
#         "monte_carlo_comparison.xlsx",
#         "greeks_comparison.xlsx",
#         "touch_probs_comparison.xlsx",
#         "etf_volume_comparison.xlsx"
#     ]
#
#     for fname in files:
#         fpath = os.path.join(comparison_dir, fname)
#         if not os.path.exists(fpath):
#             continue
#
#         try:
#             book = load_workbook(fpath)
#             for sheet in book.sheetnames:
#                 format_sheet(book, sheet)
#             format_probabilities_in_workbook(book)
#             book.save(fpath)
#             print(f"Formatted: {fpath}")
#         except Exception as e:
#             print(f"Failed formatting {fpath}: {e}")

def build_and_format_all_comparison_sheets(folder: str) -> None:
    """
    Single pass:
      1) Build each wideform once
      2) Sort once
      3) Format once
    """
    # 1) Build once
    build_wideform_options_and_pinning_comparison(folder)
    build_wideform_monte_carlo_comparison(folder)
    build_wideform_greeks_comparison(folder)
    build_wideform_touch_probs_comparison(folder)
    build_wideform_etf_volume_comparison(folder)

    # 2) Sort once
    comparison_dir = os.path.join(folder, "comparison")
    sort_all_comparison_sheets(comparison_dir)

    # 3) Format once
    targets = [
        "options_and_pinning_comparison.xlsx",
        "monte_carlo_comparison.xlsx",
        "greeks_comparison.xlsx",
        "touch_probs_comparison.xlsx",
        "etf_volume_comparison.xlsx",
    ]
    for fname in targets:
        fpath = os.path.join(comparison_dir, fname)
        if not os.path.exists(fpath):
            continue
        try:
            book = load_workbook(fpath)
            for _sheet in book.sheetnames:
                format_sheet(book, _sheet)
            format_probabilities_in_workbook(book)
            book.save(fpath)
            print(f"Formatted: {fpath}")
        except Exception as e:
            print(f"Failed formatting {fpath}: {e}")


def run_monologue_analysis(
    folder: str,
    file_prefix: str,
    strategy_file: str,
    log_dir: str,
    v_now,                  # DataFrame: latest 10-min ETF volume row(s)
    gap_now,                # kept for API compatibility; can be empty
    latest_file_written: str,
) -> None:

    files_by_day = get_last_10_days_files(folder, file_prefix)
    if not files_by_day:
        print("No recent files found.")
        return

    all_filepaths = [latest_file_written] + [
        fp for paths in files_by_day.values() for fp in paths if fp != latest_file_written
    ]

    sorted_paths = sorted(all_filepaths, reverse=True)
    filepaths = sorted(
        sorted_paths[:2],
        key=lambda f: pd.to_datetime(extract_timestamp(f),
                                     format="%Y-%m-%d_%a_%H-%M")
    )

    try:
        filename = os.path.basename(filepaths[-1]).replace(".xlsx", "")
        timestamp_str = "_".join(filename.split("_")[-3:])  # e.g., '2025-08-05_tue_11-34'
        day = pd.to_datetime(timestamp_str, format="%Y-%m-%d_%a_%H-%M").date()
    except Exception:
        day = pd.Timestamp.today().date()

    strategy_memory = load_strategy_sheet_from_excel(strategy_file)
    daily_changes, spot_changes, pin_tracking, gamma_summary = [], [], [], []

    mc_now, g_now, o_now, _, _ = load_sheets(filepaths[1])
    mc_prev, g_prev, o_prev, _, _ = load_sheets(filepaths[0])
    if mc_now is None or mc_prev is None:
        return

    spot = mc_now["spot"].iloc[0]
    mc_now = mc_now[mc_now["strike"].between(spot - 10, spot + 10)]
    mc_prev = mc_prev[mc_prev["strike"].between(spot - 10, spot + 10)]

    # Merge using only 'strike' (MC is strike-level)
    g_now_deduped = g_now.drop_duplicates(subset="strike")
    g_prev_deduped = g_prev.drop_duplicates(subset="strike")
    merged_now = mc_now.merge(g_now_deduped, on="strike", how="left")
    merged_prev = mc_prev.merge(g_prev_deduped, on="strike", how="left")

    changes = compare_strikes(merged_now, merged_prev)
    daily_changes.extend(changes)

    ts = os.path.basename(filepaths[1]).split("_")[-1].replace(".xlsx", "")
    spot_changes.append({
        "timestamp": ts,
        "spot_now": merged_now.get("spot_x", pd.NA).iloc[0],
        "spot_prev": merged_prev.get("spot_x", pd.NA).iloc[0]
    })

    # FIX: check columns, not DataFrame membership
    if "pinning_strength" in o_now.columns and o_now["pinning_strength"].notna().any():
        top_pin = o_now.loc[o_now["pinning_strength"].idxmax()]
        pin_tracking.append({
            "timestamp": ts,
            "strike": top_pin["strike"],
            "strength": top_pin["pinning_strength"]
        })

    gamma_summary.append(f"{ts}: Avg Net Gamma = {g_now['net_gamma_exposure'].mean():,.0f}")

    strategy_memory = update_strategies(day, daily_changes, strategy_memory)
    write_day_log(day, daily_changes, strategy_memory, spot_changes, gamma_summary,
                  pin_tracking, v_now, gap_now, log_dir)
    save_strategy_sheet_to_excel(strategy_memory, strategy_file)

    # Single clean pass: build, sort, format
    build_and_format_all_comparison_sheets(folder)
