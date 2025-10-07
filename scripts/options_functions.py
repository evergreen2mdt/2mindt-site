# === Standard Library ===
import datetime
import logging
import os
from copy import copy
from typing import Optional


# === Core Data & Stats Libraries ===
import numpy as np
import pandas as pd
from scipy.stats import norm

# === Market & Finance Utilities ===
import yfinance as yf
import pandas_market_calendars as mcal

# === Timezones ===
from zoneinfo import ZoneInfo

# === Excel Utilities ===
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# ============================ CONTRACT MAP ===================================
# =============================================================================
from config import CONTRACT_MAP
from config import DEFAULT_START_DATE

# =============================================================================
# ======================== COLUMN LAYOUTS / CONSTANTS =========================
# =============================================================================
RAW_OPTIONS_COLUMNS = [
    "strike", "type", "dte", "expiration", "spot",
    "bid", "ask", "volume", "openInterest", "inTheMoney",
    "impliedVolatility"
]

PINNING_METRICS_COLUMNS = [
    "strike", "total_activity", "pinning_strength",
    "volume_based_pin_rank", "volume_based_pin_candidate",
    "influence_based_pinning_rank", "influence_based_pinning_candidate"
]

GREEKS_COLUMNS = [
    "strike", "type", "dte", "spot", "risk_free_rate",
    "delta (black-scholes)", "gamma (black-scholes)",
    "theta (black-scholes)", "vega (black-scholes)",
    "delta_exposure", "gamma_exposure", "theta_exposure", "vega_exposure",
    "call_delta_exposure", "put_delta_exposure", "net_delta_exposure",
    "call_gamma_exposure", "put_gamma_exposure", "net_gamma_exposure",
    "call_theta_exposure", "put_theta_exposure", "net_theta_exposure",
    "call_vega_exposure", "put_vega_exposure", "net_vega_exposure",
    "call_OI", "put_OI", "put_call_oi_ratio",
    "weighted_vega_exposure", "weighted_gamma_exposure",
    "vega_concentration", "vega_crush_zone", "theta_gravity",
    "strike_wall_tag", "oi_gap_zone", "put_call_oi_skew",
    "iv_slope", "gamma_regime", "target_vs_expected_move", "day_of_week_oi_bias"
]

MONTE_CARLO_COLUMNS = [
    "strike", "spot",
    "prob_hit_day_0","prob_hit_day_1","prob_hit_day_2","prob_hit_day_3","prob_hit_day_4",
    "prob_hit_day_5","prob_hit_day_6","prob_hit_day_7","prob_hit_day_8","prob_hit_day_9",
    "prob_hit_day_0_adjusted","prob_hit_day_1_adjusted","prob_hit_day_2_adjusted",
    "prob_hit_day_3_adjusted","prob_hit_day_4_adjusted","prob_hit_day_5_adjusted",
    "prob_hit_day_6_adjusted","prob_hit_day_7_adjusted","prob_hit_day_8_adjusted",
    "prob_hit_day_9_adjusted","adj_factor"
]

BASE_PROBS_COLUMNS = [
    "strike",
    "prob_hit_day_0","prob_hit_day_1","prob_hit_day_2","prob_hit_day_3","prob_hit_day_4",
    "prob_hit_day_5","prob_hit_day_6","prob_hit_day_7","prob_hit_day_8","prob_hit_day_9"
]


# =============================================================================
# ======================== UTILITIES: MARKET/TIMING ===========================
# =============================================================================

def is_market_open():
    now = datetime.datetime.now()
    is_weekday = now.weekday() < 5
    is_market_time = (now.hour == 9 and now.minute >= 30) or (10 <= now.hour < 16)
    return is_weekday and is_market_time


def setup_logging(today, log_dir, timestamp_str):
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f"options_log_{timestamp_str}.txt")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    logging.info(f"Today: {today.date()} (normalized)")
    return logging


def check_trading_day(today):
    nyse = mcal.get_calendar("NYSE")
    schedule = nyse.schedule(start_date=today, end_date=today)
    if schedule.empty:
        raise SystemExit(f"{today.date()} is not a trading day.")


# =============================================================================
# ======================== SPOT PRICE (YFINANCE) ==============================
# =============================================================================

def get_yf_spot_price(ticker):
    t = yf.Ticker(ticker)
    try:
        spot_price = t.fast_info.get("lastPrice")
        if spot_price is None or np.isnan(spot_price):
            raise ValueError("fast_info returned None/NaN")
    except Exception:
        try:
            spot_price = t.info.get("currentPrice")
            if spot_price is None or np.isnan(spot_price):
                raise ValueError("info['currentPrice'] None/NaN")
        except Exception as e:
            logging.warning(f"Failed to retrieve spot price for {ticker}: {e}")
            return None, None
    return t, float(spot_price)


# =============================================================================
# ======================== OPTION CHAIN (YFINANCE) ============================
# =============================================================================

def get_yf_option_chain(ticker_obj, spot_price, today, max_dte=30):
    expirations = ticker_obj.options
    valid_expirations = [
        datetime.datetime.strptime(exp, "%Y-%m-%d")
        for exp in expirations
        if 0 <= (datetime.datetime.strptime(exp, "%Y-%m-%d").date() - today.date()).days <= max_dte
    ]

    all_data = []
    for exp_date in valid_expirations:
        dte = (exp_date.date() - today.date()).days
        exp_str = exp_date.strftime("%Y-%m-%d")
        try:
            chain = ticker_obj.option_chain(exp_str)
            for opt_type, df_opt in zip(["call", "put"], [chain.calls, chain.puts]):
                df_opt = df_opt[
                    ["strike", "bid", "ask", "impliedVolatility",
                     "volume", "openInterest", "inTheMoney"]
                ].copy()
                df_opt["type"] = opt_type
                df_opt["expiration"] = exp_str
                df_opt["dte"] = dte
                df_opt = df_opt[
                    (df_opt["strike"] >= spot_price - 15) &
                    (df_opt["strike"] <= spot_price + 15)
                ]
                df_opt["spot"] = spot_price
                all_data.append(df_opt)
        except Exception as e:
            logging.warning(f"Failed to fetch option chain {exp_str}: {e}")
            continue

    if not all_data:
        return pd.DataFrame()
    df_all = pd.concat(all_data, ignore_index=True)
    df_all["generated_at"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    return df_all


# =============================================================================
# ============================ METRICS / TAGS =================================
# =============================================================================

def classify_moneyness(row):
    if abs(row["strike"] - row["spot"]) <= 1:
        return "ATM"
    elif (row["type"] == "call" and row["strike"] < row["spot"]) or \
         (row["type"] == "put" and row["strike"] > row["spot"]):
        return "ITM"
    else:
        return "OTM"


def apply_moneyness_columns(df):
    df["moneyness"] = df.apply(classify_moneyness, axis=1)
    df["abs_diff"] = (df["strike"] - df["spot"]).abs()
    return df


def apply_pinning_metrics(df, spot_price):
    df = apply_moneyness_columns(df)
    ranked = (
        df.groupby("strike")[["openInterest", "volume"]]
        .sum()
        .assign(total_activity=lambda x: x["openInterest"] + x["volume"])
        .reset_index()
    )
    epsilon = 1e-6
    ranked["pinning_strength"] = ranked["total_activity"] / (abs(ranked["strike"] - spot_price) + epsilon)
    ranked["volume_based_pin_rank"] = range(1, len(ranked) + 1)
    ranked["volume_based_pin_candidate"] = ranked["volume_based_pin_rank"] <= 10
    ranked["influence_based_pinning_rank"] = ranked["pinning_strength"].rank(method="dense", ascending=False).astype(int)
    ranked["influence_based_pinning_candidate"] = ranked["influence_based_pinning_rank"] <= 10

    df = df.merge(
        ranked[[
            "strike", "total_activity", "pinning_strength",
            "volume_based_pin_rank", "volume_based_pin_candidate",
            "influence_based_pinning_rank", "influence_based_pinning_candidate"
        ]],
        on="strike", how="left"
    )
    return df


# === Greeks & Enhancements (INTEGRATED) ===
def calculate_greek_exposures_and_tags(df: pd.DataFrame, spot_price: float) -> pd.DataFrame:
    from scipy.stats import norm
    treasury_yields = {1: 5.45, 2: 5.45, 3: 5.44, 7: 5.43, 10: 5.42, 14: 5.41, 21: 5.39, 30: 5.38}
    def get_r(dte):
        return treasury_yields[min(treasury_yields, key=lambda x: abs(x - dte))] / 100

    def greeks(row):
        S = max(float(row["spot"]), 1e-6)
        K = max(float(row["strike"]), 1e-6)
        T = max(float(row["dte"]) / 365.0, 1e-4)
        sigma = max(float(row["impliedVolatility"]), 1e-8)
        r = get_r(row["dte"])
        try:
            d1 = (np.log(S / K) + (r + 0.5 * sigma ** 2) * T) / (sigma * np.sqrt(T))
            d2 = d1 - sigma * np.sqrt(T)
            if row["type"] == "call":
                delta = norm.cdf(d1)
                theta = (-S * norm.pdf(d1) * sigma / (2 * np.sqrt(T)) - r * K * np.exp(-r * T) * norm.cdf(d2)) / 365
            else:
                delta = -norm.cdf(-d1)
                theta = (-S * norm.pdf(d1) * sigma / (2 * np.sqrt(T)) + r * K * np.exp(-r * T) * norm.cdf(-d2)) / 365
            gamma = norm.pdf(d1) / max(S * sigma * np.sqrt(T), 1e-8)
            vega = S * norm.pdf(d1) * np.sqrt(T) / 100
            return pd.Series([delta, gamma, vega, theta, r],
                             index=["delta (black-scholes)", "gamma (black-scholes)", "vega (black-scholes)",
                                    "theta (black-scholes)", "risk_free_rate"])
        except Exception:
            return pd.Series([np.nan] * 5, index=["delta (black-scholes)", "gamma (black-scholes)",
                                                  "vega (black-scholes)", "theta (black-scholes)",
                                                  "risk_free_rate"])

    df = df.copy()
    df[["delta (black-scholes)", "gamma (black-scholes)", "vega (black-scholes)",
        "theta (black-scholes)", "risk_free_rate"]] = df.apply(greeks, axis=1)

    df["gamma_exposure"] = df["openInterest"] * df["gamma (black-scholes)"] * (df["spot"] ** 2) * 0.01
    df["delta_exposure"] = df["openInterest"] * df["delta (black-scholes)"]
    df["theta_exposure"] = df["openInterest"] * df["theta (black-scholes)"]
    df["vega_exposure"] = df["openInterest"] * df["vega (black-scholes)"]

    grouped = df.groupby(["strike", "type"]).agg({
        "gamma_exposure": "sum",
        "delta_exposure": "sum",
        "theta_exposure": "sum",
        "vega_exposure": "sum",
        "openInterest": "sum"
    }).unstack(fill_value=0)

    grouped.columns = ['_'.join(col).strip() for col in grouped.columns.values]
    grouped = grouped.reset_index().rename(columns={
        "gamma_exposure_call": "call_gamma_exposure",
        "gamma_exposure_put": "put_gamma_exposure",
        "delta_exposure_call": "call_delta_exposure",
        "delta_exposure_put": "put_delta_exposure",
        "theta_exposure_call": "call_theta_exposure",
        "theta_exposure_put": "put_theta_exposure",
        "vega_exposure_call": "call_vega_exposure",
        "vega_exposure_put": "put_vega_exposure",
        "openInterest_call": "call_OI",
        "openInterest_put": "put_OI"
    })

    grouped["net_gamma_exposure"] = grouped["call_gamma_exposure"] - grouped["put_gamma_exposure"]
    grouped["net_delta_exposure"] = grouped["call_delta_exposure"] - grouped["put_delta_exposure"]
    grouped["net_theta_exposure"] = grouped["call_theta_exposure"] + grouped["put_theta_exposure"]
    grouped["net_vega_exposure"] = grouped["call_vega_exposure"] + grouped["put_vega_exposure"]
    grouped["put_call_oi_ratio"] = grouped["put_OI"] / grouped["call_OI"].replace(0, 1e-8)

    df = df.merge(grouped, on="strike", how="left")

    try:
        atm_strike = df.iloc[(df["strike"] - spot_price).abs().argsort()[:1]]["strike"].values[0]
        atm_iv = df[df["strike"] == atm_strike]["impliedVolatility"].mean()

        df["vega_concentration"] = df[df["strike"].between(spot_price - 5, spot_price + 5)]["vega_exposure"].sum()
        df["vega_crush_zone"] = (
            (df["vega_exposure"] > df["vega_exposure"].quantile(0.9)) &
            (df["dte"] <= 5) &
            (df["impliedVolatility"] > 1.25 * atm_iv)
        )

        df["theta_gravity"] = (
            (df["dte"] <= 10) &
            (df["moneyness"] == "ATM") &
            (df["theta_exposure"] > df["theta_exposure"].quantile(0.75)) &
            (pd.Timestamp.today().dayofweek in [2, 3])
        )

        df["dte_weight"] = np.exp(-df["dte"] / 10)
        df["weighted_vega_exposure"] = df["vega_exposure"] * df["dte_weight"]
        df["weighted_gamma_exposure"] = df["gamma_exposure"] * df["dte_weight"]

        target, entry = spot_price + 10, spot_price - 10
        df["strike_wall_tag"] = df["strike"].between(target - 1, target + 1) & (
            df["openInterest"] > df["openInterest"].quantile(0.90)
        )
        df["oi_gap_zone"] = df["strike"].between(target - 1, target + 1) & (
            df["openInterest"] < df["openInterest"].quantile(0.25)
        )

        call_OI = df[(df["type"] == "call") & df["strike"].between(target - 5, target + 5)]["openInterest"].sum()
        put_OI  = df[(df["type"] == "put")  & df["strike"].between(entry - 5, entry + 5)]["openInterest"].sum()
        df["put_call_oi_skew"] = call_OI / put_OI if put_OI else np.nan

        otm_calls = df[(df["type"] == "call") & (df["strike"] > spot_price)]
        df["iv_slope"] = (
            (otm_calls["impliedVolatility"].mean() - atm_iv) /
            (otm_calls["strike"].mean() - spot_price)
            if not otm_calls.empty else np.nan
        )

        df["gamma_regime"] = "positive" if df["gamma_exposure"].sum() > 0 else "negative"
        df["target_vs_expected_move"] = (target - entry) / 12
        df["day_of_week_oi_bias"] = "put_bias" if pd.Timestamp.today().dayofweek >= 3 else "call_bias"
    except Exception as e:
        logging.warning(f"Tagging enhancement failed: {e}")

    return df


# =============================================================================
# ================= Intraday Horizon & Touch Probabilities ====================
# =============================================================================
#
# def _minutes_until_rth_close(now=None, tz=ZoneInfo("America/New_York")) -> float:
#     now = pd.Timestamp.now(tz=tz) if now is None else now
#     rth_open  = now.replace(hour=9,  minute=30, second=0, microsecond=0)
#     rth_close = now.replace(hour=16, minute=0,  second=0, microsecond=0)
#     if now <= rth_open:
#         return (rth_close - rth_open).total_seconds() / 60.0
#     if now >= rth_close:
#         return 0.0
#     return (rth_close - now).total_seconds() / 60.0
#
# def _minutes_until_eth_close(now=None, tz=ZoneInfo("America/New_York")) -> float:
#     now = pd.Timestamp.now(tz=tz) if now is None else now
#     pre_open  = now.replace(hour=4,  minute=0, second=0, microsecond=0)
#     rth_open  = now.replace(hour=9,  minute=30, second=0, microsecond=0)
#     rth_close = now.replace(hour=16, minute=0, second=0, microsecond=0)
#     eth_close = now.replace(hour=20, minute=0, second=0, microsecond=0)
#     if pre_open <= now < rth_open:
#         return (rth_open - now).total_seconds() / 60.0
#     if rth_close <= now < eth_close:
#         return (eth_close - now).total_seconds() / 60.0
#     return 0.0
#
# def _active_minutes_left(now=None, tz=ZoneInfo("America/New_York")) -> float:
#     rth = _minutes_until_rth_close(now, tz)
#     if rth > 0:
#         return rth
#     return _minutes_until_eth_close(now, tz)




# # Backward-compatible name (used by old code)
# def _minutes_until_close(now=None, tz=ZoneInfo("America/New_York")) -> float:
#     return _minutes_until_rth_close(now, tz)
#
# def _intraday_T_years(dte: int, now=None) -> float:
#     if int(dte) <= 0:
#         minutes_left = _active_minutes_left(now)
#         return max(minutes_left / 390.0, 0.0) / 252.0
#     return float(dte) / 252.0

def black_scholes_touch_prob(row):
    S = float(row["spot"])
    K = float(row["strike"])
    dte = row.get("dte")
    sigma = float(row.get("iv_strike", row.get("impliedVolatility")))
    if dte is None or pd.isna(dte) or sigma <= 0 or S <= 0 or K <= 0:
        return np.nan
    T = float(dte) / 252.0
    if T <= 0:
        return np.nan
    d1 = (np.log(S / K) + 0.5 * sigma**2 * T) / (sigma * np.sqrt(T))
    prob = 2.0 * (1.0 - norm.cdf(abs(d1)))
    return float(np.clip(prob, 0.0, 1.0))


def heston_touch_prob(row, kappa=2.0, theta=None, rho=-0.7, sigma_v=0.5, num_paths=5000, random_seed=None):
    dte = row.get("dte")
    if dte is None or pd.isna(dte) or dte < 0:
        return np.nan
    T = float(dte) / 252.0
    if T <= 0:
        return np.nan
    S0 = float(row["spot"])
    K  = float(row["strike"])
    r  = float(row.get("risk_free_rate", 0.0))
    sigma = float(row.get("sigma_today", row.get("iv_strike", row.get("impliedVolatility", 0.0))))
    if S0 <= 0 or K <= 0 or sigma <= 0:
        return np.nan
    if random_seed is not None:
        np.random.seed(random_seed)
    v0 = sigma**2
    if theta is None:
        theta = v0
    steps = max(50, int(dte))
    dt = T / steps
    sqrt_dt = np.sqrt(dt)
    sqrt_1mr2 = np.sqrt(max(1.0 - rho*rho, 0.0))
    S = np.full(num_paths, S0)
    v = np.full(num_paths, v0)
    hit = np.zeros(num_paths, dtype=bool)
    for _ in range(steps):
        z1 = np.random.normal(size=num_paths)
        z2 = np.random.normal(size=num_paths)
        z2 = rho * z1 + sqrt_1mr2 * z2
        v_pos = np.maximum(v, 0.0)
        v = v + kappa * (theta - v_pos) * dt + sigma_v * np.sqrt(v_pos) * sqrt_dt * z2
        v = np.maximum(v, 0.0)
        S = S * np.exp((r - 0.5 * v_pos) * dt + np.sqrt(v_pos) * sqrt_dt * z1)
        hit |= (S >= K) if K > S0 else (S <= K)
        if hit.all():
            break
    return float(np.clip(hit.mean(), 0.0, 1.0))


def jump_diffusion_touch_prob(row, lam=0.1, mu_j=-0.02, sigma_j=0.05, max_jumps=5):
    S = float(row["spot"])
    K = float(row["strike"])
    dte = row.get("dte")
    sigma = float(row.get("iv_strike", row.get("impliedVolatility")))
    r = float(row.get("risk_free_rate", 0.0))
    if dte is None or pd.isna(dte) or sigma <= 0 or S <= 0 or K <= 0:
        return np.nan
    T = float(dte) / 252.0
    if T <= 0:
        return np.nan
    kappa = np.exp(mu_j + 0.5 * sigma_j**2) - 1.0
    drift_adj = r - lam * kappa
    prob = 0.0
    for k in range(max_jumps + 1):
        pk = np.exp(-lam*T) * (lam*T)**k / np.math.factorial(k)
        sigma_eff = np.sqrt(sigma**2 + (k * sigma_j**2) / max(T, 1e-12))
        mu_eff = drift_adj + (k * mu_j) / max(T, 1e-12)
        d1 = (np.log(S/K) + (mu_eff + 0.5*sigma_eff**2)*T) / (sigma_eff * np.sqrt(T))
        p_touch = 2.0 * (1.0 - norm.cdf(abs(d1)))
        prob += pk * np.clip(p_touch, 0.0, 1.0)
    return float(min(max(prob, 0.0), 1.0))


# def interpret_prob(model: str, prob: float) -> str:
#     if pd.isna(prob):
#         return "invalid"
#     model = model.lower()
#     if model in ["black_scholes", "heston", "jump_diffusion", "monte_carlo"]:
#         if prob >= 0.90: return "very high"
#         if prob >= 0.70: return "high"
#         if prob >= 0.10: return "moderate"
#         return "low"
#     else:
#         return "unknown model"


# =============================================================================
# ===================== Monte Carlo + Overlay / Daily Sheets ==================
# =============================================================================

def monte_carlo_probs_by_strike(df, num_paths=10000, num_days=10, random_seed=None):
    if random_seed is not None:
        np.random.seed(random_seed)
    result = []
    spot = df["spot"].iloc[0]
    df_iv = df[["strike","dte","iv_strike","risk_free_rate"]].copy()
    df_iv["d_gap"] = (df_iv["dte"] - 9).abs()
    iv_map = (df_iv.sort_values(["strike","d_gap"])
                    .groupby("strike", as_index=False)
                    .first()[["strike","iv_strike","risk_free_rate"]])
    for _, row in iv_map.iterrows():
        strike = float(row["strike"])
        sigma  = max(float(row["iv_strike"]), 0.05)
        r      = float(row.get("risk_free_rate", 0.05))
        row_result = {"strike": strike, "spot": float(spot)}
        try:
            T = 10 / 252.0
            dt = T / num_days
            drift = (r - 0.5 * sigma**2) * dt
            shock = np.random.normal(loc=drift, scale=sigma*np.sqrt(dt), size=(num_paths, num_days))
            price_paths = spot * np.exp(np.cumsum(shock, axis=1))
            hits = price_paths >= strike if strike > spot else price_paths <= strike
            day_hits = (hits.cumsum(axis=1) > 0).mean(axis=0)
            for i in range(num_days):
                row_result[f"prob_hit_day_{i}"] = float(day_hits[i])
        except Exception:
            for i in range(num_days):
                row_result[f"prob_hit_day_{i}"] = 0.0
        result.append(row_result)
    return pd.DataFrame(result)


def generate_montecarlo_sheet_full(df: pd.DataFrame) -> pd.DataFrame:
    df_monte = monte_carlo_probs_by_strike(df)
    z_features = ["pinning_strength", "net_gamma_exposure", "net_delta_exposure", "total_activity"]
    for col in z_features:
        df_monte[f"z_{col}"] = (df[col] - df[col].mean()) / (df[col].std() + 1e-8)
    weights = {"pinning_strength": 0.4, "net_gamma_exposure": 0.2, "net_delta_exposure": 0.1, "total_activity": 0.3}
    df_monte["adj_factor"] = 1 / (1 + np.exp(-sum(weights[col] * df_monte[f"z_{col}"] for col in weights)))
    for day in range(10):
        df_monte[f"prob_hit_day_{day}_adjusted"] = df_monte[f"prob_hit_day_{day}"] * df_monte["adj_factor"]
    return df_monte


def generate_daily_touch_prob_sheets(df: pd.DataFrame) -> dict:
    """
    Create one sheet per model (BS, Jump, Heston, Final),
    with prob_hit_day_0 … prob_hit_day_9 mapped directly from DTE.
    """
    models = [
        ("prob_touch_black_scholes", "black scholes probs"),
        ("prob_touch_jump_diffusion", "jump diffusion probs"),
        ("prob_touch_heston", "heston probs"),
        ("prob_touch_final", "final probs"),
    ]
    output = {}
    for col, sheet_name in models:
        rows = []
        for strike in sorted(df["strike"].unique()):
            row = {"strike": strike}
            for dte in range(10):
                sub = df[(df["strike"] == strike) & (df["dte"] == dte)]
                row[f"prob_hit_day_{dte}"] = float(sub[col].iloc[0]) if not sub.empty else np.nan
            rows.append(row)
        output[sheet_name] = pd.DataFrame(rows)[BASE_PROBS_COLUMNS]
    return output
#
# def self_test_touch_unification(df: pd.DataFrame) -> None:
#     """
#     Diagnostic checks to ensure IV unification and touch probability consistency.
#     Raises AssertionError if mismatches are found.
#     """
#
#     models = [
#         "prob_touch_black_scholes",
#         "prob_touch_heston",
#         "prob_touch_jump_diffusion"
#     ]
#
#     # Check iv_strike consistency
#     iv_chk = df.groupby(["strike", "dte"])["iv_strike"].nunique()
#     bad_iv = iv_chk[iv_chk > 1]
#     if not bad_iv.empty:
#         print("[DIAG] iv_strike mismatches by (strike,dte):")
#         print(bad_iv.head(20))
#
#     # Check model consistency across calls/puts
#     touch_chk = df.groupby(["strike", "dte"])[models].nunique()
#     bad_touch = touch_chk[(touch_chk > 1).any(axis=1)]
#     if not bad_touch.empty:
#         print("[DIAG] Touch-prob mismatches by (strike,dte):")
#         print(bad_touch.head(20))
#
#     # Check DTE=0 edge case
#     d0 = df[df["dte"] == 0]
#     if not d0.empty:
#         d0_chk = d0.groupby("strike")[models].nunique()
#         bad_d0 = d0_chk[(d0_chk > 1).any(axis=1)]
#         if not bad_d0.empty:
#             print("[DIAG] DTE=0 mismatches by strike:")
#             print(bad_d0.head(20))
#
#     # Hard stops if problems found
#     assert bad_iv.empty, "iv_strike differs across types for some (strike,dte)"
#     assert bad_touch.empty, "Touch models differ across types for some (strike,dte)"
#     if not d0.empty:
#         assert bad_d0.empty, "DTE=0 touch models differ across types for some strikes"

# =============================================================================
# ========================== EXCEL FORMATTING =================================
# =============================================================================

def format_sheet(wb, sheetname):
    ws = wb[sheetname]
    for cell in ws[1]:
        f = copy(cell.font)
        f.bold = True
        cell.font = f
    ws.freeze_panes = "A2"
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    if sheetname == "pinning metrics":
        format_pinning_candidates(ws)


def format_pinning_candidates(ws):
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for col_name in ["volume_based_pin_candidate", "influence_based_pinning_candidate"]:
        if col_name not in [cell.value for cell in ws[1]]:
            continue
        col_idx = [cell.value for cell in ws[1]].index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{ws.max_row}",
            CellIsRule(operator="equal", formula=["TRUE"], fill=fill)
        )


def format_probabilities_in_workbook(wb):
    PERCENTAGE_COLUMNS = {
        "monte carlo": [f"prob_hit_day_{i}" for i in range(10)] +
                       [f"prob_hit_day_{i}_adjusted" for i in range(10)] + ["adj_factor"],
        "black scholes probs": [f"prob_hit_day_{i}" for i in range(10)],
        "jump diffusion probs": [f"prob_hit_day_{i}" for i in range(10)],
        "heston probs": [f"prob_hit_day_{i}" for i in range(10)],
        "final probs": [f"prob_hit_day_{i}" for i in range(10)],
    }
    for sheetname, cols in PERCENTAGE_COLUMNS.items():
        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]
        header = {str(cell.value).strip(): idx+1 for idx, cell in enumerate(ws[1])}
        for col in cols:
            if col not in header:
                continue
            col_letter = get_column_letter(header[col])
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                if isinstance(cell.value, (float, int)):
                    cell.number_format = "0.00%"


# =============================================================================
# =========================== MAIN RUNNER =====================================
# =============================================================================
from dropbox_utils import upload_file
from config import get_dropbox_path

def run_options(contract_map=CONTRACT_MAP, start_date=DEFAULT_START_DATE):
    now = pd.Timestamp.now().tz_localize(None)
    today = now.normalize()
    timestamp_str = now.strftime("%Y-%m-%d_%a_%H-%M").lower()

    for ticker in contract_map:
        print(f"\n=== {ticker} ===")
        tkr_obj, spot_price = get_yf_spot_price(ticker)
        if tkr_obj is None or spot_price is None:
            print(f"Skipping {ticker}: no spot price.")
            continue

        df = get_yf_option_chain(tkr_obj, spot_price, today)
        if df.empty:
            print(f"No options data for {ticker}")
            continue

        # Map expirations to next 10 NYSE trading days
        nyse = mcal.get_calendar("NYSE")
        sched = nyse.schedule(start_date=today, end_date=today + pd.Timedelta(days=20))
        trading_days = list(sched.index.normalize())[:11]
        tmap = {d.strftime("%Y-%m-%d"): i for i, d in enumerate(trading_days)}
        df["trade_day_idx"] = df["expiration"].map(tmap)
        df = df[df["trade_day_idx"].between(0, 9)].copy()
        df["dte"] = df["trade_day_idx"].astype(int)

        # IV cleanup
        df["type"] = df["type"].astype(str).str.lower().replace({"c": "call", "p": "put"})
        df["impliedVolatility"] = pd.to_numeric(df["impliedVolatility"], errors="coerce").replace(0, np.nan)

        _eps = 1e-6
        df["bid"] = pd.to_numeric(df["bid"], errors="coerce")
        df["ask"] = pd.to_numeric(df["ask"], errors="coerce")
        df["spread"] = (df["ask"] - df["bid"]).abs().fillna(1.0).clip(lower=_eps)
        df["iv_weight"] = 1.0 / df["spread"]

        def _wavg_iv(g):
            x = pd.to_numeric(g["impliedVolatility"], errors="coerce")
            w = pd.to_numeric(g["iv_weight"], errors="coerce")
            m = x.notna() & w.notna() & (w > 0)
            return float(np.average(x[m], weights=w[m])) if m.any() else float(x.mean())

        iv_map = (
            df.groupby(["strike", "dte"], as_index=False)
              .apply(lambda g: pd.Series({"iv_strike": _wavg_iv(g)}))
        )
        df = df.merge(iv_map, on=["strike", "dte"], how="left")
        df["iv_strike"] = pd.to_numeric(df["iv_strike"], errors="coerce") \
            .fillna(pd.to_numeric(df["impliedVolatility"], errors="coerce").fillna(0.20))

        # Pinning + Greeks
        df = apply_pinning_metrics(df, spot_price)
        df = calculate_greek_exposures_and_tags(df, spot_price)

        # Realized vol blend
        def _intraday_rv_yf(symbol: str, window_min: int = 15) -> float:
            try:
                dfi = yf.download(symbol, period="1d", interval="1m", progress=False, auto_adjust=True)
                if dfi is None or dfi.empty:
                    return np.nan
                idx = dfi.index
                NY = ZoneInfo("America/New_York")
                idx = idx.tz_localize(NY) if getattr(idx, "tz", None) is None else idx.tz_convert(NY)
                ret = np.log(dfi["Close"]).diff().dropna()
                r = ret.tail(int(window_min))
                return float(r.std(ddof=0) * np.sqrt(252 * 390)) if not r.empty else np.nan
            except Exception:
                return np.nan

        rv5  = _intraday_rv_yf(ticker, 5)
        rv15 = _intraday_rv_yf(ticker, 15)
        if np.isnan(rv5) and np.isnan(rv15):
            rv5  = _intraday_rv_yf("SPY", 5)
            rv15 = _intraday_rv_yf("SPY", 15)

        df["rv_5m"]  = rv5
        df["rv_15m"] = rv15

        alpha = 0.70
        df["sigma_today"] = (
            alpha * df["iv_strike"] + (1 - alpha) * df["rv_15m"].fillna(df["iv_strike"])
        ).clip(lower=0.05, upper=3.00)

        # Touch models + ensemble
        df["prob_touch_black_scholes"] = df.apply(black_scholes_touch_prob, axis=1)
        df["prob_touch_heston"] = df.apply(heston_touch_prob, axis=1)
        df["prob_touch_jump_diffusion"] = df.apply(jump_diffusion_touch_prob, axis=1)

        p_bs   = df["prob_touch_black_scholes"].fillna(0.0)
        p_jump = df["prob_touch_jump_diffusion"].fillna(p_bs)
        p_hest = df["prob_touch_heston"].fillna(p_bs)
        p_ens  = 0.5*p_bs + 0.3*p_jump + 0.2*p_hest
        df["prob_touch_final"] = np.clip(np.maximum(p_ens, p_bs), 0.0, 0.995)

        # Monte Carlo + daily sheets
        df_monte = generate_montecarlo_sheet_full(df)
        daily_touch = generate_daily_touch_prob_sheets(df)

        df_raw_options = df[RAW_OPTIONS_COLUMNS].copy()
        df_pinning = df[PINNING_METRICS_COLUMNS].drop_duplicates(subset="strike").sort_values("strike").copy()
        df_greeks = df[GREEKS_COLUMNS].copy()
        df_monte_carlo = df_monte[MONTE_CARLO_COLUMNS].copy()

        # === Save temporary local Excel ===
        local_filename = f"{ticker.lower()}_options_data_{timestamp_str}.xlsx"
        with pd.ExcelWriter(local_filename, engine="openpyxl", mode="w") as writer:
            df_raw_options.to_excel(writer, sheet_name="raw options", index=False)
            df_pinning.to_excel(writer, sheet_name="pinning metrics", index=False)
            df_greeks.to_excel(writer, sheet_name="greeks", index=False)
            df_monte_carlo.sort_values(by=["strike"]).to_excel(writer, sheet_name="monte carlo", index=False)
            daily_touch["black scholes probs"].to_excel(writer, sheet_name="black scholes probs", index=False)
            daily_touch["jump diffusion probs"].to_excel(writer, sheet_name="jump diffusion probs", index=False)
            daily_touch["heston probs"].to_excel(writer, sheet_name="heston probs", index=False)
            daily_touch["final probs"].to_excel(writer, sheet_name="final probs", index=False)

        # Formatting
        book = load_workbook(local_filename)
        for sheet in [
            "raw options","pinning metrics","greeks","monte carlo",
            "black scholes probs","jump diffusion probs","heston probs","final probs"
        ]:
            if sheet in book.sheetnames:
                format_sheet(book, sheet)
        format_probabilities_in_workbook(book)
        book.save(local_filename)

        # === Upload to Dropbox ===
        dropbox_path = get_dropbox_path(ticker, "options", f"{ticker.lower()}_options_data_{timestamp_str}.xlsx")
        upload_file(local_filename, dropbox_path)
        print(f"[Dropbox] Uploaded {ticker} options → {dropbox_path}")
        os.remove(local_filename)
        print(f"[Local] Deleted temporary file: {local_filename}")
