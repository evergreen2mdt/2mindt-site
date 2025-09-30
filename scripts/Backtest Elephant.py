from openpyxl.formatting.rule import CellIsRule
import numpy as np
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from toolz import first


def get_50ma_mam_prior(cat_key_filter, target_date_str):
    """
    Returns the 50-day MA MAM prior to a given trade date for a specific cat_key.

    Parameters:
        cat_key_filter (str): The cat_key to filter trades.
        target_date_str (str): The date of the trade to evaluate (format: 'YYYY-MM-DD').

    Returns:
        float or None: The 50-day MA MAM prior to the target date, or None if not enough data.
    """
    file_path = r"C:\Users\colby\OneDrive\Documents\_cat spy gaps\cat_.xlsx"

    # Load sheet
    df = pd.read_excel(file_path, sheet_name="All Data")

    # Ensure date parsing and cleaning
    df["target_origination_date"] = pd.to_datetime(
        df["target_origination_date"])
    df = df[df["cat_key"] == cat_key_filter].copy()
    df = df[df["max_adverse_move"] != 0].copy()  # only non-zero MAMs
    df = df.sort_values("index")  # must be sorted by index to preserve order

    # Find the row index of the target trade
    target_date = pd.to_datetime(target_date_str)
    target_row = df[df["target_origination_date"] == target_date]

    if target_row.empty:
        print(
            f"No trade found on {target_date_str} for cat_key '{cat_key_filter}'")
        return None

    target_index = target_row.index[0]

    # Get all rows before this one (by index, not date)
    prior_rows = df.loc[:target_index - 1]

    if len(prior_rows) < 50:
        print("Not enough prior non-zero MAM values to compute 50 MA")
        return None

    last_50_mams = prior_rows.tail(50)["max_adverse_move"]
    return round(last_50_mams.mean(), 4)


#print(get_50ma_mam_prior("move_up + price_gap_down", "2008-02-01"))




# def generate_mam_cats_workbook():
#     input_path = r"C:\Users\colby\OneDrive\Documents\_cat spy gaps\cat_.xlsx"
#     output_path = r"C:\Users\colby\OneDrive\Documents\_cat spy gaps\elephant hunter\mam_cats.xlsx"
#
#     # Load main data
#     df = pd.read_excel(input_path, sheet_name='All Data')
#     df['date'] = pd.to_datetime(df['date']).dt.date
#     df['mam_date'] = pd.to_datetime(df['mam_date']).dt.date
#     df['target_origination_date'] = pd.to_datetime(df['target_origination_date']).dt.date
#     df['target_achieved_date'] = pd.to_datetime(df['target_achieved_date']).dt.date
#     df = df.dropna(subset=['cat_key', 'mam_date'])
#     df = df.reset_index(drop=True)
#     df['index'] = df.index
#
#     # Create writer
#     with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
#         for cat_key, group in df.groupby('cat_key'):
#             group = group.sort_values(by='index').reset_index(drop=True)
#             non_zero_mams = group[group['max_adverse_move'] != 0]
#
#             if len(non_zero_mams) < 50:
#                 continue
#
#             ma_values = []
#             for i, row in group.iterrows():
#                 prior_rows = non_zero_mams[non_zero_mams['index'] < row['index']]
#                 recent = prior_rows.tail(50)
#                 if len(recent) < 50:
#                     ma_values.append(np.nan)
#                 else:
#                     ma_values.append(round(recent['max_adverse_move'].mean(), 4))
#
#             group['50_ma_mam'] = ma_values
#             group['cat_key'] = cat_key
#
#             # Drop rows with MAM = 0 before writing
#             group = group[group['max_adverse_move'] != 0]
#
#             output = group[['index', 'date', 'cat_key', 'max_adverse_move', 'mam_date', 'mam_index',
#                             'days_to_mam', 'target_origination_date','target_achieved_date', '50_ma_mam']]
#             sheet_name = f"{cat_key}"[:31]
#             output.to_excel(writer, sheet_name=sheet_name, index=False)
#
#     # Reopen for formatting
#     wb = openpyxl.load_workbook(output_path)
#     for sheet in wb.sheetnames:
#         ws = wb[sheet]
#         ws.freeze_panes = "A2"
#
#         # Bold headers
#         for cell in ws[1]:
#             cell.font = Font(bold=True)
#
#         # Auto column width
#         for col in ws.columns:
#             max_len = max((len(str(cell.value)) if cell.value is not None else 0 for cell in col), default=0)
#             ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
#
#     wb.save(output_path)
#     print(f"Workbook created: {output_path}")
#     os.startfile(output_path)

# Run it
#generate_mam_cats_workbook()




#
# def generate_full_trade_workbook_separated():
#     input_path = r"C:\Users\colby\OneDrive\Documents\_cat spy gaps\cat_.xlsx"
#     output_path = r"C:\Users\colby\OneDrive\Documents\_cat spy gaps\elephant hunter\mam_full_trades_separated.xlsx"
#
#     df = pd.read_excel(input_path, sheet_name='All Data')
#     df['date'] = pd.to_datetime(df['date']).dt.date
#     df['mam_date'] = pd.to_datetime(df['mam_date']).dt.date
#     df['target_origination_date'] = pd.to_datetime(df['target_origination_date']).dt.date
#     df['target_achieved_date'] = pd.to_datetime(df['target_achieved_date']).dt.date
#
#     df = df.dropna(subset=['cat_key', 'mam_date'])
#     df['index'] = df.index
#
#     results = {}
#
#     for cat_key, group in df.groupby('cat_key'):
#         group = group.sort_values(by='index').reset_index(drop=True)
#         non_zero_trades = group[group['max_adverse_move'] != 0]
#
#         if len(non_zero_trades) < 50:
#             continue
#
#         trade_blocks = []
#
#         for trade_id, (_, trade_row) in enumerate(non_zero_trades.iterrows(), start=1):
#             start = trade_row['target_origination_date']
#             end = trade_row['target_achieved_date']
#             trade_block = df[(df['date'] >= start) & (df['date'] <= end)].copy()
#             trade_block['trade_id'] = trade_id  # Add a unique ID per trade block
#             trade_blocks.append(trade_block)
#
#         if trade_blocks:
#             full_df = pd.concat(trade_blocks).drop_duplicates().sort_values(by=['trade_id', 'index'])
#             full_df = full_df[['trade_id', 'index', 'date', 'cat_key', 'max_adverse_move', 'mam_date', 'mam_index',
#                                'days_to_mam', 'target_origination_date', 'target_achieved_date']]
#             results[f"cat_{cat_key}"] = full_df
#
#     with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
#         for sheet_name, data in results.items():
#             data.to_excel(writer, sheet_name=sheet_name[:31], index=False)
#
#     # Apply formatting
#     wb = openpyxl.load_workbook(output_path)
#     highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#
#     for sheet in wb.sheetnames:
#         ws = wb[sheet]
#         ws.freeze_panes = "A2"
#
#         # Bold headers
#         for cell in ws[1]:
#             cell.font = Font(bold=True)
#
#         # Auto column width
#         for col in ws.columns:
#             max_length = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
#             adjusted_width = max_length + 2
#             ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
#
#         # Highlight cat_key that matches the sheet name
#         cat_key_value = sheet.replace("cat_", "")
#         for idx, cell in enumerate(ws[1], start=1):
#             if cell.value == 'cat_key':
#                 cat_col_letter = get_column_letter(idx)
#                 ws.conditional_formatting.add(
#                     f"{cat_col_letter}2:{cat_col_letter}{ws.max_row}",
#                     CellIsRule(operator='equal', formula=[f'"{cat_key_value}"'], fill=highlight_fill)
#                 )
#                 break
#
#     wb.save(output_path)
#     os.startfile(output_path)
#     print(f"Workbook created at: {output_path}")

# Run it
# generate_full_trade_workbook_separated()





import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

def generate_filtered_rdd_mam_trades():
    input_path = r"C:\Users\colby\OneDrive\Documents\_cat spy gaps\cat_.xlsx"
    output_path = r"C:\Users\colby\OneDrive\Documents\_cat spy gaps\elephant hunter\full list elephants.xlsx"

    df = pd.read_excel(input_path, sheet_name='All Data')
    df['date'] = pd.to_datetime(df['date']).dt.date
    df['mam_date'] = pd.to_datetime(df['mam_date']).dt.date
    df['target_origination_date'] = pd.to_datetime(df['target_origination_date']).dt.date
    df['target_achieved_date'] = pd.to_datetime(df['target_achieved_date']).dt.date
    df = df.dropna(subset=['cat_key', 'mam_date'])
    df['index'] = df.index

    results = {}
    for cat_key in df['cat_key'].unique():
        group = df[df['cat_key'] == cat_key].sort_values(by='index').reset_index(drop=True)
        non_zero_trades = group[group['max_adverse_move'] != 0]
        if len(non_zero_trades) < 50:
            continue

        non_zero_group = group[group['max_adverse_move'] != 0]
        trade_rows = []
        trade_id_counter = 1

        for _, trade in non_zero_trades.iterrows():
            start = trade['target_origination_date']
            if start < pd.to_datetime("2009-01-01").date():
                continue

            end = trade['target_achieved_date'] if pd.notnull(trade['target_achieved_date']) else trade['mam_date']

            entry_price = trade['close/target']
            trade_index = trade['index']
            prior_trades = non_zero_group[non_zero_group['index'] < trade_index]
            recent_50 = prior_trades.tail(50)
            if len(recent_50) < 50:
                continue

            ma_mam = round(recent_50['max_adverse_move'].mean(), 4)
            in_trade_df = df[(df['date'] >= start) & (df['date'] <= end)].copy()
            in_trade_df['trade_id'] = trade_id_counter

            if 'gap_down' in cat_key.lower():
                in_trade_df['running_drawdown'] = entry_price - in_trade_df['low']
            else:
                in_trade_df['running_drawdown'] = in_trade_df['high'] - entry_price

            in_trade_df['percent_drawdown'] = (in_trade_df['running_drawdown'] / entry_price).round(4)
            in_trade_df['ma_mam_for_trade'] = ma_mam
            in_trade_df['rdd_mam'] = in_trade_df['running_drawdown'] / ma_mam
            in_trade_df.loc[in_trade_df['running_drawdown'] == 0, 'rdd_mam'] = np.nan

            in_trade_df['days_to_target'] = None
            try:
                start_idx = df[df['date'] == start].index[0]
                end_idx = df[df['date'] == end].index[0]
                in_trade_df.loc[in_trade_df['date'] == start, 'days_to_target'] = end_idx - start_idx
            except:
                pass

            trade_origin_mask = in_trade_df['date'] == start
            in_trade_df.loc[~trade_origin_mask, 'max_adverse_move'] = np.nan

            for col in ['mam_date', 'mam_index', 'days_to_mam', 'target_origination_date', 'target_achieved_date']:
                in_trade_df.loc[~trade_origin_mask, col] = np.nan

            # Only include trade if any row in it has rdd_mam >= 4
            if (in_trade_df['rdd_mam'] >= 4.0).any():
                trade_rows.append(in_trade_df)
                trade_id_counter += 1

        if trade_rows:
            combined = pd.concat(trade_rows).sort_values(by=['trade_id', 'date'])
            combined = combined[['trade_id', 'index', 'date', 'cat_key', 'running_drawdown', 'max_adverse_move',
                                 'ma_mam_for_trade', 'rdd_mam', 'mam_date', 'mam_index', 'days_to_mam',
                                 'days_to_target', 'target_origination_date', 'target_achieved_date',
                                 'open', 'high', 'low', 'close/target', 'percent_drawdown',
                                 'VIX', 'vix_regime']]
            results[cat_key] = combined

    # Write to Excel
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        for sheet_name, data in results.items():
            data.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    # Format the workbook
    wb = openpyxl.load_workbook(output_path)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == 'cat_key':
                col_letter = get_column_letter(idx)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    openpyxl.formatting.rule.CellIsRule(
                        operator='equal', formula=[f'"{sheet}"'], fill=yellow_fill
                    )
                )

            if cell.value == 'rdd_mam':
                col_letter = get_column_letter(idx)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    openpyxl.formatting.rule.CellIsRule(
                        operator='greaterThanOrEqual', formula=['4'], fill=red_fill
                    )
                )

            if cell.value == 'percent_drawdown':
                col_letter = get_column_letter(idx)
                for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                    for cell in row:
                        cell.number_format = '0.00%'

    wb.save(output_path)
    os.startfile(output_path)
    print(f"Filtered workbook created at: {output_path}")

# Uncomment to run
#generate_filtered_rdd_mam_trades()








import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def generate_cat_key_summaries():
    input_path = r"C:\Users\colby\OneDrive\Documents\_cat spy gaps\elephant hunter\full list elephants.xlsx"
    vix_path = r"C:\Users\colby\OneDrive\Documents\_cat spy gaps\cat_.xlsx"
    output_path = r"C:\Users\colby\OneDrive\Documents\_cat spy gaps\elephant hunter\summary short elephants.xlsx"

    # Load VIX data from All Data sheet
    vix_df = pd.read_excel(vix_path, sheet_name="All Data", usecols=["target_origination_date", "VIX"])
    vix_df.rename(columns={"target_origination_date": "entry_date"}, inplace=True)
    vix_df["entry_date"] = pd.to_datetime(vix_df["entry_date"]).dt.date

    def classify_vix(v):
        if pd.isna(v): return None
        elif v < 12: return "very low"
        elif v < 17: return "normal"
        elif v < 25: return "elevated"
        elif v < 35: return "high"
        else: return "extreme"

    vix_df["vix_regime"] = vix_df["VIX"].apply(classify_vix)

    xl = pd.ExcelFile(input_path)
    all_summaries = []
    writer = pd.ExcelWriter(output_path, engine="openpyxl")

    for sheet in xl.sheet_names:
        if "gap_up" in sheet.lower():
            continue

        df = xl.parse(sheet)
        if 'trade_id' not in df.columns:
            continue

        for col in ['date', 'mam_date', 'target_origination_date', 'target_achieved_date']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col]).dt.date

        trade_summaries = []

        for trade_id, trade_df in df.groupby('trade_id'):
            trade_df = trade_df.sort_values('date').copy()

            if 'running_drawdown' not in trade_df.columns or 'ma_mam_for_trade' not in trade_df.columns:
                continue

            trade_df['rdd_mam'] = trade_df['running_drawdown'] / trade_df['ma_mam_for_trade']
            trade_df.loc[trade_df['running_drawdown'] == 0, 'rdd_mam'] = None

            entry_date = pd.NaT
            entry = None
            rdd_cross = trade_df[trade_df['rdd_mam'] >= 4]

            if not rdd_cross.empty:
                idx = rdd_cross.index[0]
                next_idx = idx + 1
                if next_idx in trade_df.index:
                    entry_date = trade_df.loc[next_idx, 'date']
                    entry = trade_df.loc[next_idx, 'open']

            first_row = trade_df.iloc[0].copy()
            first_row['entry_date'] = entry_date
            first_row['entry'] = entry
            first_row['rdd_mam'] = first_row['max_adverse_move'] / first_row['ma_mam_for_trade']
            first_row['profit from entry'] = first_row['close/target'] - entry

            dd_window = trade_df[(trade_df['date'] >= entry_date) & (trade_df['date'] <= first_row['target_achieved_date'])]
            lowest_low = dd_window['low'].min()
            first_row['lowest_low'] = lowest_low
            first_row['dd from entry'] = entry - lowest_low

            drop_threshold = ((first_row['profit from entry']) * 0.45)
            first_row['45% drop'] = drop_threshold
            first_row['doable'] = "True" if drop_threshold <= first_row['dd from entry'] else "False"
            first_row['new entry'] = entry - drop_threshold

            new_entry_date_row = trade_df[trade_df['date'] > entry_date]
            new_entry_date_row = new_entry_date_row[new_entry_date_row['low'] <= first_row['new entry']]
            if not new_entry_date_row.empty:
                first_row['new entry date'] = new_entry_date_row.iloc[0]['date']
            else:
                first_row['new entry date'] = pd.NaT

            low_window = trade_df[(trade_df['date'] >= first_row['new entry date']) &
                                  (trade_df['date'] <= first_row['target_achieved_date'])] if pd.notna(first_row['new entry date']) else pd.DataFrame()
            if not low_window.empty:
                first_row['low after new entry'] = low_window['low'].min()
                first_row['new dd'] = first_row['new entry'] - first_row['low after new entry']
            else:
                first_row['low after new entry'] = None
                first_row['new dd'] = None
            first_row['new profit'] = first_row['close/target'] - first_row['new entry'] if first_row['doable'] == "True" else None

            trade_summaries.append(first_row)

        summary_df = pd.DataFrame(trade_summaries)
        summary_df = summary_df.merge(vix_df, how="left", on="entry_date")
        summary_df = summary_df.drop(columns=['running_drawdown', 'percent_drawdown'], errors='ignore')
        summary_df.to_excel(writer, sheet_name=sheet[:31], index=False)
        all_summaries.append(summary_df)

    writer.close()

    # Combine and assign clusters
    combined_df = pd.concat(all_summaries, ignore_index=True)
    if 'target_origination_date' in combined_df.columns:
        combined_df = combined_df.sort_values(by='target_origination_date')

    combined_df = combined_df.sort_values(by='entry_date').reset_index(drop=True)
    combined_df['cluster'] = 0
    current_cluster = 0
    current_end = datetime.min.date()

    for i, row in combined_df.iterrows():
        start = row['entry_date']
        end = row['target_achieved_date']
        if pd.isnull(start) or pd.isnull(end):
            combined_df.at[i, 'cluster'] = -1
            continue
        if start > current_end:
            current_cluster += 1
            current_end = end
        else:
            current_end = max(current_end, end)
        combined_df.at[i, 'cluster'] = current_cluster

    cols = combined_df.columns.tolist()
    cols.insert(0, cols.pop(cols.index('cluster')))
    combined_df = combined_df[cols]

    # Final write and formatting
    wb = openpyxl.load_workbook(output_path)
    summary_path = output_path.replace('.xlsx', '_temp.xlsx')
    with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
        combined_df.to_excel(writer, sheet_name="Summary", index=False)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            data = list(ws.iter_rows(values_only=True))
            if not data or len(data) < 2:
                continue
            df = pd.DataFrame(data[1:], columns=data[0])
            df.to_excel(writer, sheet_name=sheet[:31], index=False)

    os.replace(summary_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    date_cols = ['date', 'mam_date', 'target_origination_date', 'target_achieved_date', 'entry_date', 'new entry date']
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "A2"
        headers = [cell.value for cell in ws[1]]
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
        for col_name in date_cols:
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for cell in ws[col_letter][1:]:
                    cell.number_format = "yyyy-mm-dd"

    wb.save(output_path)
    os.startfile(output_path)
    print(f"Workbook saved at: {output_path}")

# Run it
generate_cat_key_summaries()
