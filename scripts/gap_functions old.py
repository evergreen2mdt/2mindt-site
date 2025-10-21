# === Standard Library ===
import os
from datetime import datetime

# === Third-Party Libraries ===
from tqdm import tqdm

# === Interactive Brokers API ===
from ib_insync import IB, Stock

# === Excel Handling ===

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from scipy.stats import trim_mean
import pandas as pd



def get_ib_spy_data(start_date="2003-01-01"):
    ib = IB()
    try:
        ib.connect('127.0.0.1', 7496, clientId=1)
        print('Connected to IB')
    except Exception as e:
        print(f"Connection failed: {e}")
        return None

    contract = Stock('SPY', 'SMART', 'USD')
    ib_start_date = datetime.strptime(start_date, '%Y-%m-%d')
    ib_end_date = datetime.now()
    duration_days = (ib_end_date - ib_start_date).days
    duration_years = (duration_days // 365) + 1
    duration_str = f'{duration_years} Y'
    print(f"end date: {ib_end_date}")
    bars = ib.reqHistoricalData(
        contract,
        endDateTime=ib_end_date.strftime("%Y%m%d %H:%M:%S"),
        durationStr=duration_str,
        barSizeSetting='1 day',
        whatToShow='TRADES',
        useRTH=True,
        formatDate=1
    )

    if not bars:
        print("No historical data fetched from IB.")
        return None

    all_data = []
    for bar in tqdm(bars, desc="Downloading IB SPY bars"):
        all_data.append({
            'date': bar.date,
            'open': bar.open,
            'high': bar.high,
            'low': bar.low,
            'close': bar.close,
            'volume': bar.volume
        })

    df = pd.DataFrame(all_data)
    df['date'] = pd.to_datetime(df['date']).dt.date
    print("SPY data retrieved from IB.")
    return df


CORRECTIONS = {
    "2004-04-19": {"open": 113.55, "high": 113.99, "low": 113.27, "close": 113.83, "volume": 28277600},
    "2004-04-27": {"open": 114.23, "high": 115.12, "low": 113.96, "close": 114.30, "volume": 43485500},
    "2004-05-05": {"open": 112.41, "high": 112.96, "low": 112.16, "close": 112.78, "volume": 34405000},
    "2006-03-03": {"open": 128.67, "high": 130.07, "low": 128.65, "close": 128.76, "volume": 73402496},
    "2006-05-18": {"open": 127.35, "high": 127.75, "low": 126.11, "close": 126.21, "volume": 37906304},
    "2006-03-23": {"open": 130.26, "high": 130.39, "low": 129.66, "close": 130.11, "volume": 46704200},
    "2007-04-02": {"open": 142.16, "high": 142.46, "low": 140.89, "close": 142.16, "volume": 79454680},
    "2007-05-25": {"open": 151.49, "high": 152.02, "low": 151.18, "close": 151.69, "volume": 33309621},
    "2007-06-29": {"open": 150.90, "high": 151.65, "low": 149.15, "close": 150.43, "volume": 39701730},
    "2007-07-23": {"open": 154.18, "high": 154.72, "low": 153.30, "close": 153.97, "volume": 21183836},
    "2007-07-26": {"open": 150.19, "high": 150.80, "low": 146.39, "close": 148.02, "volume": 57592490},
    "2007-08-10": {"open": 144.39, "high": 146.50, "low": 143.12, "close": 144.71, "volume": 11018370},
    "2022-01-24": {"open": 432.03, "high": 440.38, "low": 420.76, "close": 439.84, "volume": 52496719},
    "2022-01-26": {"open": 440.72, "high": 444.04, "low": 428.86, "close": 433.38, "volume": 36391088},
    "2022-02-24": {"open": 411.02, "high": 428.76, "low": 410.64, "close": 428.30, "volume": 13942946},
    "2022-05-04": {"open": 417.08, "high": 429.66, "low": 413.71, "close": 429.06, "volume": 44247895},
    "2022-05-05": {"open": 424.55, "high": 425.00, "low": 409.44, "close": 413.81, "volume": 72929106},
    "2022-05-20": {"open": 393.25, "high": 397.03, "low": 380.54, "close": 389.63, "volume": 31432197},
    "2022-10-13": {"open": 349.21, "high": 367.51, "low": 348.11, "close": 365.97, "volume": 47254518},
    "2024-08-01": {"open": 552.57, "high": 554.87, "low": 539.43, "close": 543.01, "volume": 76428732},
    "2024-09-11": {"open": 548.70, "high": 555.36, "low": 539.96, "close": 554.42, "volume": 75248608},
    "2024-12-18": {"open": 603.98, "high": 606.41, "low": 585.89, "close": 586.28, "volume": 38248729},
    "2025-03-03": {"open": 596.18, "high": 597.34, "low": 579.90, "close": 583.77, "volume": 74249199},
    "2025-04-04": {"open": 523.67, "high": 525.87, "low": 505.06, "close": 505.28, "volume": 17965131},
    "2025-04-07": {"open": 489.19, "high": 523.17, "low": 481.80, "close": 504.38, "volume": 56611355},
    "2025-04-08": {"open": 521.86, "high": 524.98, "low": 489.16, "close": 496.48, "volume": 55816581},
    "2025-04-09": {"open": 493.44, "high": 548.62, "low": 493.05, "close": 548.62, "volume": 11867317},
    "2025-04-10": {"open": 532.17, "high": 533.50, "low": 509.32, "close": 524.58, "volume": 52331225},
    "2025-04-11": {"open": 523.01, "high": 536.43, "low": 520.07, "close": 533.94, "volume": 37866334},
    "2025-04-30": {"open": 547.57, "high": 556.52, "low": 541.52, "close": 554.54, "volume": 33101463}
}

def compute_gap_data(df):
    """
    Compute previous day move types, gap type, current day move types,
    trade categories, days to target, and same-day move types.
    """
    df = df.copy()

    prev_day_move_types = []
    gap_types = []
    days_to_target = []
    target_dates = []

    df['year'] = pd.to_datetime(df['date']).dt.year
    df['dow'] = pd.to_datetime(df['date']).dt.day_name()
    df['previous_date'] = df['date'].shift(1)
    df['previous_open'] = df['open'].shift(1)
    df['previous_high'] = df['high'].shift(1)
    df['previous_low'] = df['low'].shift(1)
    df['previous_close'] = df['close'].shift(1)
    df['gap value'] = abs(df['open'] - df['previous_close'])

    data = df.to_dict('records')

    for i in tqdm(range(1, len(data)), desc="Computing gap data"):
        prev_bar = data[i - 1]
        curr_bar = data[i]

        if prev_bar['close'] == prev_bar['open']:
            prev_day_move_type = 'doji'
        elif prev_bar['close'] > prev_bar['open']:
            prev_day_move_type = 'move_up'
        else:
            prev_day_move_type = 'move_down'
        prev_day_move_types.append(prev_day_move_type)

        if prev_bar['close'] < curr_bar['open'] <= prev_bar['high']:
            gap_type = 'price_gap_up'
        elif prev_bar['close'] > curr_bar['open'] >= prev_bar['low']:
            gap_type = 'price_gap_down'
        elif curr_bar['open'] > prev_bar['high']:
            gap_type = 'bar_gap_up'
        elif curr_bar['open'] < prev_bar['low']:
            gap_type = 'bar_gap_down'
        else:
            gap_type = 'no_gap'
        gap_types.append(gap_type)

        target_achieved = False
        days_count = 0
        target_achieved_date = None

        if gap_type == 'no_gap':
            days_to_target.append(0)
            target_dates.append(prev_bar['date'])
        else:
            for j in range(i, len(data)):
                next_bar = data[j]
                days_count += 1
                if gap_type in ['price_gap_up', 'bar_gap_up'] and next_bar['low'] <= prev_bar['close']:
                    target_achieved = True
                    target_achieved_date = next_bar['date']
                    break
                elif gap_type in ['price_gap_down', 'bar_gap_down'] and next_bar['high'] >= prev_bar['close']:
                    target_achieved = True
                    target_achieved_date = next_bar['date']
                    break

            days_to_target.append(days_count if target_achieved else None)
            target_dates.append(target_achieved_date)

    df = df.iloc[1:].reset_index(drop=True).copy()
    df['prev_day_move_types'] = prev_day_move_types
    df['gap_type'] = gap_types
    df['days to target'] = days_to_target

    df['target_achieved_date'] = target_dates

    df['gap_type'] = df['gap_type']
    df['gap value'] = df['gap value']
    df['original_index'] = range(len(df))

    df['trade_direction'] = df['gap_type'].apply(
        lambda x: "not tradable" if "no_gap" in str(x).lower()
        else "long" if "down" in str(x).lower()
        else "short" if "up" in str(x).lower()
        else None
    )

    df["cat_key"] = df["prev_day_move_types"].astype(str) + " + " + df["gap_type"].astype(str)

    def classify_same_day_move(row):
        if row["trade_direction"] == "long":
            if row["open"] > row["close"]:
                return "and_go"
            elif row["open"] < row["close"]:
                return "and_comeback"
            else:
                return "none"
        elif row["trade_direction"] == "short":
            if row["open"] < row["close"]:
                return "and_go"
            elif row["open"] > row["close"]:
                return "and_comeback"
            else:
                return "none"
        else:
            return "none"

    df["same_day_move_type"] = df.apply(classify_same_day_move, axis=1)

    return df


def calculate_statistics(df, trim_percentage=0.1):
    df = df.copy()
    if 'days to target' not in df.columns:
        raise ValueError("'days to target' column is required for calculate_statistics()")

    window = 50
    tqdm.write("Calculating rolling stats...")

    df[f'MA {window} Days to Close'] = df['days to target'].rolling(window=window, min_periods=1).mean()
    df[f'Rolling Stdev {window} Days'] = df['days to target'].rolling(window=window, min_periods=1).std()

    df['Median'] = df['days to target'].rolling(window=window, min_periods=1).median()
    df['Trim Avg Days to Close 50 Days'] = df['days to target'].rolling(window=window, min_periods=1).apply(
        lambda x: trim_mean(x, proportiontocut=trim_percentage), raw=False
    )
    df['Trim Stdev to Days to Close 50 Days'] = df['days to target'].rolling(window=window, min_periods=1).std()

    df['1st Quartile (25%)'] = df['days to target'].rolling(window=window, min_periods=1).quantile(0.25)
    df['2nd Quartile (50%)'] = df['days to target'].rolling(window=window, min_periods=1).quantile(0.50)
    df['3rd Quartile (75%)'] = df['days to target'].rolling(window=window, min_periods=1).quantile(0.75)
    df['IQR'] = df['3rd Quartile (75%)'] - df['1st Quartile (25%)']
    df['Upper Bound'] = df['3rd Quartile (75%)'] + (1.5 * df['IQR'])

    tqdm.write("Statistics calculation complete.")
    return df



def create_days_to_target_summary_table_gap_types(file_path):


    # === Load and clean data ===
    df = pd.read_excel(file_path, sheet_name="All Data")
    df = df[df["days to target"].notna()]
    df["gap_type"] = df["gap_type"].astype(str)
    df["days to target"] = (
        pd.to_numeric(df["days to target"], errors="coerce")
        .fillna(0)
        .astype(int)
    )

    max_day = 10
    summary_rows = []

    for gap_type in sorted(df["gap_type"].unique()):
        sub_df = df[df["gap_type"] == gap_type]
        total = len(sub_df)
        if total == 0:
            continue

        cumulative_counts = [(sub_df["days to target"] <= i).sum() for i in range(1, max_day + 1)]

        cumulative_pct = [round(c / total, 4) for c in cumulative_counts]

        row = [gap_type] + cumulative_pct
        summary_rows.append(row)

    # === Construct output dataframe ===
    columns = ["gap_type"] + [f"Days {i}" for i in range(1, max_day + 1)]

    table_df = pd.DataFrame(summary_rows, columns=columns)

    return table_df




def compute_rolling_stats_by_gap_type(file_path):
    df = pd.read_excel(file_path, sheet_name="All Data")
    df["gap_type"] = df["gap_type"].astype(str)

    # Ensure datetime types
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["target_achieved_date"] = pd.to_datetime(df["target_achieved_date"], errors="coerce")

    # Calculate days to target (only where both dates exist)
    df["days to target"] = (df["target_achieved_date"] - df["date"]).dt.days
    df = df[df["days to target"].notna()].copy()

    # Compute grouped rolling stats
    rolling_stats = df.groupby("gap_type")["days to target"].agg([
        ("rolling_mean", "mean"),
        ("rolling_std", "std"),
        ("rolling_trim_mean", lambda x: trim_mean(x, proportiontocut=0.1)),
        ("rolling_trim_std", lambda x: x[(x >= x.quantile(0.1)) & (x <= x.quantile(0.9))].std()),
        ("rolling_q1", lambda x: x.quantile(0.25)),
        ("rolling_median", "median"),
        ("rolling_q3", lambda x: x.quantile(0.75)),
        ("rolling_upper_bound", lambda x: x.quantile(0.75) + 1.5 * (x.quantile(0.75) - x.quantile(0.25))),
        ("rolling_mad", lambda x: (x - x.median()).abs().mean()),
    ]).reset_index()

    return rolling_stats



def compute_gap_data_with_mam_and_days_to_mam(df):
    """
    Computes max adverse move (MAM), MAM date, index, and days to MAM.
    Handles:
    - 'no gap': assigns default values
    - gap + no target hit: fallback MAM scan (entry → end of dataset)
    - gap + target hit: normal MAM scan (entry → target_achieved_date or fallback)
    """
    max_adverse_moves = [None]
    mam_dates = [None]
    mam_indexes = [None]
    days_to_mam = [None]

    df['date'] = pd.to_datetime(df['date'])

    for i in tqdm(range(1, len(df)), desc="Computing MAM"):
        curr_bar = df.iloc[i]
        gap_type = str(curr_bar.get('gap_type', '')).strip().lower().replace('_', ' ')
        entry_price = curr_bar["open"]
        target_achieved_date = curr_bar.get('target_achieved_date', pd.NaT)

        # CASE 1: no gap
        if gap_type == 'no gap':
            mam_dates.append(curr_bar['date'])
            mam_index = i
            mam_indexes.append(mam_index)
            days_to_mam.append(0)
            max_adverse_moves.append(0)
            continue

        # CASE 2: gap with no target achieved (fallback)
        if pd.isnull(target_achieved_date):
            j_end = len(df) - 1
        else:
            match = df[df['date'] == target_achieved_date]
            j_end = match.index[0] if not match.empty else len(df)

        if gap_type in ['price gap down', 'bar gap down']:  # LONG
            min_low = float('inf')
            min_low_index = None
            for j in range(i, j_end + 1):
                low = df.iloc[j]['low']
                if low < min_low:
                    min_low = low
                    min_low_index = j
            mam_value = entry_price - min_low
            mam_date = df.iloc[min_low_index]['date'] if min_low_index is not None else None
            mam_index = min_low_index

        elif gap_type in ['price gap up', 'bar gap up']:  # SHORT
            max_high = -float('inf')
            max_high_index = None
            for j in range(i, j_end + 1):
                high = df.iloc[j]['high']
                if high > max_high:
                    max_high = high
                    max_high_index = j
            mam_value = max_high - entry_price
            mam_date = df.iloc[max_high_index]['date'] if max_high_index is not None else None
            mam_index = max_high_index

        else:
            mam_value = None
            mam_date = None
            mam_index = None

        max_adverse_moves.append(mam_value)
        mam_dates.append(mam_date)
        mam_indexes.append(mam_index)
        days_to_mam.append(mam_index - i + 1 if mam_index is not None else None)

    df['max_adverse_move'] = max_adverse_moves
    df['mam_date'] = mam_dates
    df['mam_index'] = mam_indexes
    df['days_to_mam'] = days_to_mam

    return df

def generate_mam_stats_by_gap_type(file_path):
    df = pd.read_excel(file_path, sheet_name="All Data")
    df["gap_type"] = df["gap_type"].astype(str)
    df = df[df["max_adverse_move"].notna()].copy()

    grouped = df.groupby("gap_type")["max_adverse_move"].agg([
        ("mean_mam", "mean"),
        ("std_mam", "std"),
        ("median_mam", "median"),
        ("q1_mam", lambda x: x.quantile(0.25)),
        ("q3_mam", lambda x: x.quantile(0.75)),
        ("mad_mam", lambda x: (x - x.median()).abs().mean())
    ]).reset_index()

    grouped["iqr_mam"] = grouped["q3_mam"] - grouped["q1_mam"]
    grouped["upper_bound"] = grouped["q3_mam"] + 1.5 * grouped["iqr_mam"]


    return grouped




    print(f"Workbook written and all sheets formatted: {file_path}")

def generate_mam_days_stats_by_gap_type(file_path):
    df = pd.read_excel(file_path, sheet_name="All Data")
    df["gap_type"] = df["gap_type"].astype(str)
    df = df[df["days_to_mam"].notna()].copy()

    grouped = df.groupby("gap_type")["days_to_mam"].agg([
        ("mean_days_to_mam", "mean"),
        ("std_days_to_mam", "std"),
        ("median_days_to_mam", "median"),
        ("q1_days_to_mam", lambda x: x.quantile(0.25)),
        ("q3_days_to_mam", lambda x: x.quantile(0.75)),
        ("mad_days_to_mam", lambda x: (x - x.median()).abs().mean())
    ]).reset_index()

    grouped["iqr_days_to_mam"] = grouped["q3_days_to_mam"] - grouped["q1_days_to_mam"]
    grouped["upper_bound"] = grouped["q3_days_to_mam"] + 1.5 * grouped["iqr_days_to_mam"]

    return grouped


def format_sheet(wb, sheet_name):
    ws = wb[sheet_name]

    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2


def format_probabilities_in_workbook(wb):
    if "Days to Target" in wb.sheetnames:
        ws = wb["Days to Target"]
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if isinstance(cell.value, (float, int)):
                    cell.number_format = '0.0%'
def save_multiple_sheets_with_formatting(sheets_dict, output_path):
    wb = Workbook()
    first_sheet = True

    for sheet_name, df in sheets_dict.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        write_index = sheet_name == "All Data"

        # Write header
        headers = []
        if write_index:
            headers.append("index")
        headers.extend(map(str, df.columns))
        ws.append(headers)

        # Write data
        for idx, row in df.iterrows():
            values = [idx] if write_index else []
            values.extend(row.tolist())
            ws.append(values)

        format_sheet(wb, sheet_name)

    format_probabilities_in_workbook(wb)
    wb.save(output_path)


def run_gap_analysis(start_date="2003-01-01"):
    ib = IB()

    # === Pull SPY OHLCV ===
    df = get_ib_spy_data(start_date=start_date)

    # === Apply known corrections ===
    df["date"] = pd.to_datetime(df["date"])
    for date_str, values in CORRECTIONS.items():
        mask = df["date"] == pd.to_datetime(date_str)
        for col, val in values.items():
            df.loc[mask, col] = val

    if df is None or df.empty:
        raise ValueError("SPY DataFrame is empty or None.")

    # === Compute gap data + MAM ===
    df = compute_gap_data(df)
    df["date"] = pd.to_datetime(df["date"]).dt.normalize()
    df["target_achieved_date"] = pd.to_datetime(df["target_achieved_date"]).dt.normalize()
    df = compute_gap_data_with_mam_and_days_to_mam(df)
    df = calculate_statistics(df)

    # === Output path ===
    output_path = r"C:\2mdt\2mindt-site\public\spy-gaps-analysis\spy gap analysis.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # === Save base sheet so file exists ===
    save_multiple_sheets_with_formatting({"All Data": df}, output_path)

    # === Now safe to read from the file ===
    rolling_stats = compute_rolling_stats_by_gap_type(output_path)
    days_to_target = create_days_to_target_summary_table_gap_types(output_path)
    mam_stats_pts = generate_mam_stats_by_gap_type(output_path)
    mam_stats_days = generate_mam_days_stats_by_gap_type(output_path)

    # === Save everything ===

    save_multiple_sheets_with_formatting(
        {
            "All Data": df,
            "Rolling Stats (DAYS)": rolling_stats,
            "Days to Target": days_to_target,
            "MAM (PTS)": mam_stats_pts,
            "MAM (DAYS)": mam_stats_days,  # ✅ new worksheet
        },
        output_path
    )

    ib.disconnect()






